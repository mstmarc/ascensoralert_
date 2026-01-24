"""
Blueprint de Reportes

Generación de reportes mensuales en formato Excel con:
- Visitas a instalaciones del mes
- Visitas a administradores del mes
- Oportunidades activas (no ganadas ni perdidas)
"""

from flask import Blueprint, render_template, request, redirect, session, Response
from datetime import datetime
import calendar
import io
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import config

# Configuración de Supabase
SUPABASE_URL = config.SUPABASE_URL
HEADERS = config.HEADERS

# Crear Blueprint
reportes_bp = Blueprint('reportes', __name__)


@reportes_bp.route('/reporte_mensual', methods=["GET", "POST"])
def reporte_mensual():
    """Generar reporte mensual en Excel con visitas y oportunidades"""
    if "usuario" not in session:
        return redirect("/")

    if request.method == "POST":
        mes = int(request.form.get("mes"))
        año = int(request.form.get("año"))

        ultimo_dia = calendar.monthrange(año, mes)[1]
        fecha_inicio = f"{año}-{mes:02d}-01"
        fecha_fin = f"{año}-{mes:02d}-{ultimo_dia}"

        query_clientes = f"fecha_visita=gte.{fecha_inicio}&fecha_visita=lte.{fecha_fin}"
        response_clientes = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?{query_clientes}&select=*", headers=HEADERS)

        response_seguimiento = requests.get(
            f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?fecha_visita=gte.{fecha_inicio}&fecha_visita=lte.{fecha_fin}&select=*,clientes(nombre_cliente,direccion,localidad)",
            headers=HEADERS
        )

        response_admin = requests.get(f"{SUPABASE_URL}/rest/v1/visitas_administradores?{query_clientes}&select=*", headers=HEADERS)

        # Obtener oportunidades activas (no cerradas: que no estén ganadas ni perdidas)
        response_oportunidades = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?estado=neq.ganada&estado=neq.perdida&select=*,clientes(direccion,localidad)",
            headers=HEADERS
        )

        if response_clientes.status_code != 200:
            return f"Error al obtener datos: {response_clientes.text}"

        clientes_mes = response_clientes.json()
        visitas_seguimiento_mes = response_seguimiento.json() if response_seguimiento.status_code == 200 else []
        visitas_admin_mes = response_admin.json() if response_admin.status_code == 200 else []
        oportunidades_activas = response_oportunidades.json() if response_oportunidades.status_code == 200 else []

        # Ordenar todas las visitas por fecha
        clientes_mes = sorted(clientes_mes, key=lambda x: x.get('fecha_visita', ''))
        visitas_seguimiento_mes = sorted(visitas_seguimiento_mes, key=lambda x: x.get('fecha_visita', ''))
        visitas_admin_mes = sorted(visitas_admin_mes, key=lambda x: x.get('fecha_visita', ''))

        # Función helper para formatear fechas a dd/mm/aaaa
        def formatear_fecha(fecha_str):
            if not fecha_str:
                return ''
            try:
                # Convertir desde formato ISO (YYYY-MM-DD) a dd/mm/aaaa
                fecha = datetime.strptime(fecha_str[:10], '%Y-%m-%d')
                return fecha.strftime('%d/%m/%Y')
            except:
                return fecha_str

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "VISITAS INSTALACIONES"

        headers = ['FECHA', 'COMUNIDAD/EMPRESA', 'DIRECCION', 'ZONA', 'OBSERVACIONES']

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 2

        for cliente in clientes_mes:
            ws1.cell(row=row, column=1, value=formatear_fecha(cliente.get('fecha_visita', '')))
            ws1.cell(row=row, column=2, value=cliente.get('nombre_cliente', ''))
            ws1.cell(row=row, column=3, value=cliente.get('direccion', ''))
            ws1.cell(row=row, column=4, value=cliente.get('localidad', ''))
            ws1.cell(row=row, column=5, value=cliente.get('observaciones', ''))

            for col in range(1, 6):
                ws1.cell(row=row, column=col).border = thin_border

            row += 1

        for visita in visitas_seguimiento_mes:
            ws1.cell(row=row, column=1, value=formatear_fecha(visita.get('fecha_visita', '')))
            ws1.cell(row=row, column=2, value=visita.get('clientes', {}).get('nombre_cliente', ''))
            ws1.cell(row=row, column=3, value=visita.get('clientes', {}).get('direccion', ''))
            ws1.cell(row=row, column=4, value=visita.get('clientes', {}).get('localidad', ''))
            ws1.cell(row=row, column=5, value=visita.get('observaciones', ''))

            for col in range(1, 6):
                ws1.cell(row=row, column=col).border = thin_border

            row += 1

        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 50
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 70

        ws2 = wb.create_sheet(title="VISITAS ADMINISTRADORES")

        headers_admin = ['FECHA', 'ADMINISTRADOR', 'PERSONA CONTACTO', 'OBSERVACIONES']

        for col, header in enumerate(headers_admin, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 2
        for visita in visitas_admin_mes:
            ws2.cell(row=row, column=1, value=formatear_fecha(visita.get('fecha_visita', '')))
            ws2.cell(row=row, column=2, value=visita.get('administrador_fincas', ''))
            ws2.cell(row=row, column=3, value=visita.get('persona_contacto', ''))
            ws2.cell(row=row, column=4, value=visita.get('observaciones', ''))

            for col in range(1, 5):
                ws2.cell(row=row, column=col).border = thin_border

            row += 1

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 70

        # Tercera pestaña: OPORTUNIDADES ACTIVAS
        ws3 = wb.create_sheet(title="OPORTUNIDADES ACTIVAS")

        headers_oportunidades = ['DIRECCION', 'LOCALIDAD', 'TIPO DE OPORTUNIDAD', 'ESTADO', 'DESCRIPCION']

        for col, header in enumerate(headers_oportunidades, 1):
            cell = ws3.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 2
        # Mapeo de estados a etiquetas legibles
        estados_map = {
            'nueva': '🆕 Nueva',
            'en_contacto': '📞 En contacto',
            'presupuesto_preparacion': '✍️ Presupuesto en preparación',
            'presupuesto_enviado': '📤 Presupuesto enviado',
            'ganada': '✅ Ganada',
            'perdida': '❌ Perdida',
            'activa': '⚡ Activa'
        }

        for oportunidad in oportunidades_activas:
            # Obtener dirección y localidad del cliente relacionado
            direccion = oportunidad.get('clientes', {}).get('direccion', '') if oportunidad.get('clientes') else ''
            localidad = oportunidad.get('clientes', {}).get('localidad', '') if oportunidad.get('clientes') else ''

            # Obtener etiqueta del estado
            estado = oportunidad.get('estado', '')
            estado_label = estados_map.get(estado, estado)

            ws3.cell(row=row, column=1, value=direccion)
            ws3.cell(row=row, column=2, value=localidad)
            ws3.cell(row=row, column=3, value=oportunidad.get('tipo', ''))
            ws3.cell(row=row, column=4, value=estado_label)
            ws3.cell(row=row, column=5, value=oportunidad.get('descripcion', ''))

            for col in range(1, 6):
                ws3.cell(row=row, column=col).border = thin_border

            row += 1

        ws3.column_dimensions['A'].width = 50
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 30
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 70

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        meses = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        filename = f"DESCARGO COMERCIAL GRAN CANARIA {meses[mes]} {año}.xlsx"

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    return render_template("reporte_mensual.html")

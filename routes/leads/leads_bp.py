"""
Blueprint para gestión de Leads (Clientes potenciales / Instalaciones)

Este módulo incluye:
- Formulario de alta de leads
- Dashboard con filtros y búsqueda
- Exportación a Excel
- Visualización de detalles
- Edición y eliminación
"""

from flask import Blueprint, render_template, request, redirect, url_for, session, send_file
import requests
import helpers
from config import config
from datetime import datetime, date
from utils.formatters import limpiar_none, calcular_color_ipo, calcular_color_contrato
from utils.pagination import get_pagination
from services.cache_service import get_administradores_cached, get_filtros_cached
from utils.messages import flash_success, flash_error
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io

# Crear Blueprint sin prefijo (las rutas mantienen su estructura original)
leads_bp = Blueprint('leads', __name__)

# Constantes de configuración
SUPABASE_URL = config.SUPABASE_URL
HEADERS = config.HEADERS


# ============================================
# FORMULARIO DE ALTA DE LEADS
# ============================================

@leads_bp.route('/formulario_lead', methods=["GET", "POST"])
def formulario():
    """Formulario para crear nuevo lead/cliente"""

    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get("administrador_id")
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        data = {
            "fecha_visita": request.form.get("fecha_visita") or None,
            "tipo_cliente": request.form.get("tipo_lead") or None,
            "direccion": request.form.get("direccion"),
            "nombre_cliente": request.form.get("nombre_lead") or None,
            "codigo_postal": request.form.get("codigo_postal") or None,
            "localidad": request.form.get("localidad"),
            "zona": request.form.get("zona") or None,
            "persona_contacto": request.form.get("persona_contacto") or None,
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "administrador_id": administrador_id,
            "empresa_mantenedora": request.form.get("empresa_mantenedora") or None,
            "numero_ascensores": request.form.get("numero_ascensores") or None,
            "fecha_fin_contrato": request.form.get("fecha_fin_contrato") or None,
            "paradas": request.form.get("paradas") or None,
            "viviendas_por_planta": request.form.get("viviendas_por_planta") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Solo dirección y localidad son obligatorios
        required = [data["direccion"], data["localidad"]]
        if any(not field for field in required):
            return "Datos del lead inválidos - Dirección y Localidad son obligatorios", 400

        response = requests.post(f"{SUPABASE_URL}/rest/v1/clientes?select=id", json=data, headers=HEADERS)
        if response.status_code in [200, 201]:
            cliente_id = response.json()[0]["id"]
            return redirect(f"/nuevo_equipo?lead_id={cliente_id}")
        else:
            return f"<h3 style='color:red;'>Error al registrar lead</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

    # GET - Obtener lista de administradores (usando caché)
    fecha_hoy = date.today().strftime('%Y-%m-%d')
    administradores = get_administradores_cached()

    return render_template("formulario_lead.html", fecha_hoy=fecha_hoy, administradores=administradores)


# ============================================
# DASHBOARD DE LEADS
# ============================================

@leads_bp.route('/leads')
def leads_redirect():
    """Redirect de /leads a /leads_dashboard para compatibilidad"""
    return redirect(url_for('leads.dashboard'))


@leads_bp.route('/leads_dashboard')
def dashboard():
    """Dashboard principal de leads con filtros y búsqueda"""

    # Usar helper de paginación
    pagination = get_pagination(per_page_default=25)
    page = pagination.page
    per_page = pagination.per_page
    offset = pagination.offset

    filtro_localidad = request.args.get("localidad", "")
    filtro_empresa = request.args.get("empresa", "")
    # Aceptar tanto 'search' (desde home) como 'buscar_direccion' (desde dashboard)
    buscar_direccion = request.args.get("search", "") or request.args.get("buscar_direccion", "")

    # Si hay búsqueda de texto, usar RPC para búsqueda sin acentos
    if buscar_direccion:
        # Usar función RPC para búsqueda sin acentos
        rpc_url = f"{SUPABASE_URL}/rest/v1/rpc/buscar_clientes_sin_acentos"

        rpc_params = {
            "termino_busqueda": buscar_direccion,
            "filtro_localidad": filtro_localidad,
            "filtro_empresa": filtro_empresa,
            "limite": per_page,
            "desplazamiento": offset
        }

        response = requests.post(rpc_url, json=rpc_params, headers=HEADERS)

        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al buscar leads</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

        leads_base = response.json()

        # Obtener total_count del primer resultado si existe
        total_registros = leads_base[0].get('total_count', 0) if leads_base else 0
        total_pages = max(1, (total_registros + per_page - 1) // per_page)

    else:
        # Búsqueda normal con filtros (sin texto de búsqueda)
        query_params = []

        if filtro_localidad:
            query_params.append(f"localidad=eq.{filtro_localidad}")

        if filtro_empresa:
            query_params.append(f"empresa_mantenedora=eq.{filtro_empresa}")

        query_string = "&".join(query_params) if query_params else ""

        # OPTIMIZACIÓN: Para count solo necesitamos id, no todos los campos
        count_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id"
        if query_string:
            count_url += f"&{query_string}"

        count_response = requests.get(count_url, headers={**HEADERS, "Prefer": "count=exact"})
        total_registros = int(count_response.headers.get("Content-Range", "0").split("/")[-1])
        total_pages = max(1, (total_registros + per_page - 1) // per_page)

        # OPTIMIZACIÓN: Usar join de Supabase + selección específica de campos
        data_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id,direccion,nombre_cliente,localidad,empresa_mantenedora,numero_ascensores,created_at&limit={per_page}&offset={offset}"
        if query_string:
            data_url += f"&{query_string}"

        response = requests.get(data_url, headers=HEADERS)

        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al obtener leads</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

        leads_base = response.json()

    # Ahora obtener equipos para cada cliente encontrado
    # Esto es necesario porque RPC no devuelve relaciones anidadas
    cliente_ids = [lead['id'] for lead in leads_base]

    equipos_por_cliente = {}
    if cliente_ids:
        # Obtener todos los equipos de estos clientes en una sola query
        equipos_url = f"{SUPABASE_URL}/rest/v1/equipos?select=cliente_id,ipo_proxima,fecha_vencimiento_contrato&cliente_id=in.({','.join(map(str, cliente_ids))})"
        equipos_response = requests.get(equipos_url, headers=HEADERS)

        if equipos_response.status_code == 200:
            equipos_data = equipos_response.json()
            # Agrupar equipos por cliente_id
            for equipo in equipos_data:
                cliente_id = equipo['cliente_id']
                if cliente_id not in equipos_por_cliente:
                    equipos_por_cliente[cliente_id] = []
                equipos_por_cliente[cliente_id].append(equipo)

    rows = []

    # OPTIMIZACIÓN: Usar caché de filtros (TTL: 30 minutos)
    # Reduce de 2 queries a 0 queries en cargas subsecuentes
    localidades_disponibles, empresas_disponibles = get_filtros_cached()

    # Procesar leads y sus equipos
    for lead in leads_base:
        lead_id = lead["id"]

        # Obtener equipos de este cliente
        equipos = equipos_por_cliente.get(lead_id, [])

        direccion = lead.get("direccion", "-")
        nombre_cliente = lead.get("nombre_cliente", "")
        localidad = lead.get("localidad", "-")
        empresa_mantenedora = lead.get("empresa_mantenedora", "-")
        total_equipos = len(equipos) if equipos else lead.get("numero_ascensores", 0)
        created_at = lead.get("created_at")

        ipo_proxima = None
        contrato_vence = None

        # Procesar equipos si existen
        if equipos:
            for equipo in equipos:
                ipo_equipo = equipo.get("ipo_proxima")
                if ipo_equipo:
                    try:
                        ipo_date = datetime.strptime(ipo_equipo, "%Y-%m-%d")
                        if ipo_proxima is None or ipo_date < ipo_proxima:
                            ipo_proxima = ipo_date
                    except:
                        pass

                contrato_equipo = equipo.get("fecha_vencimiento_contrato")
                if contrato_equipo:
                    try:
                        contrato_date = datetime.strptime(contrato_equipo, "%Y-%m-%d")
                        if contrato_vence is None or contrato_date < contrato_vence:
                            contrato_vence = contrato_date
                    except:
                        pass

        # Crear fila SIEMPRE, tenga o no equipos
        ipo_proxima_str = ipo_proxima.strftime("%d/%m/%Y") if ipo_proxima else "-"
        contrato_vence_str = contrato_vence.strftime("%d/%m/%Y") if contrato_vence else "-"

        color_ipo = calcular_color_ipo(ipo_proxima_str)
        color_contrato = calcular_color_contrato(contrato_vence_str)

        row = {
            "lead_id": lead_id,
            "direccion": direccion,
            "nombre_cliente": nombre_cliente,
            "localidad": localidad,
            "total_equipos": total_equipos,
            "empresa_mantenedora": empresa_mantenedora,
            "ipo_proxima": ipo_proxima_str,
            "ipo_fecha_original": ipo_proxima,
            "contrato_vence": contrato_vence_str,
            "contrato_fecha_original": contrato_vence,
            "color_ipo": color_ipo,
            "color_contrato": color_contrato,
            "created_at": created_at
        }

        rows.append(row)

    # Ordenar por fecha de creación: más reciente primero
    rows.sort(key=lambda x: x.get("created_at") or "", reverse=True)

    # Los filtros ya vienen ordenados del caché, no es necesario volver a ordenar
    # localidades_disponibles y empresas_disponibles ya están ordenados

    return render_template("dashboard.html",
        rows=rows,
        localidades=localidades_disponibles,
        empresas=empresas_disponibles,
        filtro_localidad=filtro_localidad,
        filtro_empresa=filtro_empresa,
        buscar_direccion=buscar_direccion,
        page=page,
        total_pages=total_pages,
        total_registros=total_registros,
        per_page=per_page
    )


# ============================================
# EXPORTAR LEADS A EXCEL
# ============================================

@leads_bp.route('/exportar_leads')
@helpers.login_required
def exportar():
    """Exportar leads a formato Excel con filtros aplicados"""

    # Obtener los mismos filtros que el dashboard
    filtro_localidad = request.args.get("localidad", "")
    filtro_empresa = request.args.get("empresa", "")
    buscar_direccion = request.args.get("buscar_direccion", "")

    # Construir query (sin límite de paginación para exportar todo)
    if buscar_direccion:
        # Usar RPC para búsqueda sin acentos (sin límite)
        rpc_url = f"{SUPABASE_URL}/rest/v1/rpc/buscar_clientes_sin_acentos"
        rpc_params = {
            "termino_busqueda": buscar_direccion,
            "filtro_localidad": filtro_localidad,
            "filtro_empresa": filtro_empresa,
            "limite": 10000,  # Límite alto para exportación
            "desplazamiento": 0
        }
        response = requests.post(rpc_url, json=rpc_params, headers=HEADERS)
        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al buscar leads</h3>"
        leads_base = response.json()
    else:
        # Query normal con filtros
        query_params = []
        if filtro_localidad:
            query_params.append(f"localidad=eq.{filtro_localidad}")
        if filtro_empresa:
            query_params.append(f"empresa_mantenedora=eq.{filtro_empresa}")

        query_string = "&".join(query_params) if query_params else ""
        data_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id,direccion,nombre_cliente,localidad,empresa_mantenedora,numero_ascensores,telefono,email,persona_contacto,administrador_fincas"
        if query_string:
            data_url += f"&{query_string}"

        response = requests.get(data_url, headers=HEADERS)
        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al obtener leads</h3>"
        leads_base = response.json()

    # Obtener equipos para calcular IPOs
    cliente_ids = [lead['id'] for lead in leads_base]
    equipos_por_cliente = {}
    if cliente_ids:
        equipos_url = f"{SUPABASE_URL}/rest/v1/equipos?select=cliente_id,ipo_proxima&cliente_id=in.({','.join(map(str, cliente_ids))})"
        equipos_response = requests.get(equipos_url, headers=HEADERS)
        if equipos_response.status_code == 200:
            equipos_data = equipos_response.json()
            for equipo in equipos_data:
                cliente_id = equipo['cliente_id']
                if cliente_id not in equipos_por_cliente:
                    equipos_por_cliente[cliente_id] = []
                equipos_por_cliente[cliente_id].append(equipo)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Instalaciones"

    # Estilo de cabecera
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras
    headers = ["Dirección", "Nombre", "Población", "Teléfono", "Email", "Contacto",
               "Empresa Mantenedora", "Administrador", "Nº Ascensores", "Próxima IPO"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Datos
    row_num = 2
    for lead in leads_base:
        equipos = equipos_por_cliente.get(lead['id'], [])

        # Calcular próxima IPO
        ipo_proxima = None
        if equipos:
            for equipo in equipos:
                ipo_equipo = equipo.get("ipo_proxima")
                if ipo_equipo:
                    try:
                        ipo_date = datetime.strptime(ipo_equipo, "%Y-%m-%d")
                        if ipo_proxima is None or ipo_date < ipo_proxima:
                            ipo_proxima = ipo_date
                    except:
                        pass

        ipo_str = ipo_proxima.strftime("%d/%m/%Y") if ipo_proxima else "-"

        ws.cell(row=row_num, column=1, value=lead.get('direccion', ''))
        ws.cell(row=row_num, column=2, value=lead.get('nombre_cliente', ''))
        ws.cell(row=row_num, column=3, value=lead.get('localidad', ''))
        ws.cell(row=row_num, column=4, value=lead.get('telefono', ''))
        ws.cell(row=row_num, column=5, value=lead.get('email', ''))
        ws.cell(row=row_num, column=6, value=lead.get('persona_contacto', ''))
        ws.cell(row=row_num, column=7, value=lead.get('empresa_mantenedora', ''))
        ws.cell(row=row_num, column=8, value=lead.get('administrador_fincas', ''))
        ws.cell(row=row_num, column=9, value=lead.get('numero_ascensores', 0))
        ws.cell(row=row_num, column=10, value=ipo_str)
        row_num += 1

    # Ajustar ancho de columnas
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generar nombre de archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"instalaciones_{timestamp}.xlsx"

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================
# VER DETALLE DE LEAD
# ============================================

@leads_bp.route('/ver_lead/<int:lead_id>')
@helpers.login_required
def ver(lead_id):
    """Ver detalle completo del lead con equipos, oportunidades y visitas"""

    response = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}", headers=HEADERS)
    if response.status_code != 200 or not response.json():
        return f"<h3 style='color:red;'>Error al obtener Lead</h3><pre>{response.text}</pre><a href='{url_for('leads.dashboard')}'>Volver</a>"

    lead = response.json()[0]

    equipos_response = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?cliente_id=eq.{lead_id}", headers=HEADERS)
    equipos = []
    if equipos_response.status_code == 200:
        equipos_raw = equipos_response.json()
        for equipo in equipos_raw:
            fecha_venc = equipo.get("fecha_vencimiento_contrato", "-")
            if fecha_venc and fecha_venc != "-":
                try:
                    partes = fecha_venc.split("-")
                    if len(partes) == 3:
                        fecha_venc = f"{partes[2]}/{partes[1]}/{partes[0]}"
                except:
                    pass

            ipo = equipo.get("ipo_proxima", "-")
            if ipo and ipo != "-":
                try:
                    partes = ipo.split("-")
                    if len(partes) == 3:
                        ipo = f"{partes[2]}/{partes[1]}/{partes[0]}"
                except:
                    pass

            equipos.append({
                "id": equipo.get("id"),
                "tipo_equipo": equipo.get("tipo_equipo", "-"),
                "identificacion": equipo.get("identificacion", "-"),
                "rae": equipo.get("rae", "-"),
                "ipo_proxima": ipo,
                "fecha_vencimiento_contrato": fecha_venc,
                "descripcion": equipo.get("descripcion", "-")
            })

    oportunidades_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?cliente_id=eq.{lead_id}&order=fecha_creacion.desc",
        headers=HEADERS
    )
    oportunidades = []
    if oportunidades_response.status_code == 200:
        oportunidades = oportunidades_response.json()

    visitas_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?cliente_id=eq.{lead_id}&select=*,oportunidades(tipo)",
        headers=HEADERS
    )
    visitas_seguimiento = []
    if visitas_response.status_code == 200:
        visitas_seguimiento = visitas_response.json()

    # Combinar todas las visitas (inicial + seguimiento) y ordenar por fecha descendente
    todas_visitas = []

    # Agregar la visita inicial del lead
    if lead.get('fecha_visita'):
        todas_visitas.append({
            'fecha_visita': lead.get('fecha_visita'),
            'observaciones': lead.get('observaciones'),
            'tipo': 'inicial',
            'oportunidades': None
        })

    # Agregar las visitas de seguimiento
    for visita in visitas_seguimiento:
        todas_visitas.append({
            'fecha_visita': visita.get('fecha_visita'),
            'observaciones': visita.get('observaciones'),
            'tipo': 'seguimiento',
            'oportunidades': visita.get('oportunidades')
        })

    # Ordenar todas las visitas por fecha descendente (más reciente primero)
    todas_visitas.sort(key=lambda x: x.get('fecha_visita', ''), reverse=True)

    # Obtener datos del administrador si existe relación
    administrador = None
    if lead.get('administrador_id'):
        admin_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{lead['administrador_id']}",
            headers=HEADERS
        )
        if admin_response.status_code == 200 and admin_response.json():
            administrador = admin_response.json()[0]

    # Obtener tareas comerciales del cliente (abiertas y cerradas)
    tareas_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{lead_id}&order=fecha_creacion.desc",
        headers=HEADERS
    )
    tareas_comerciales = []
    if tareas_response.status_code == 200:
        tareas_comerciales = tareas_response.json()

    return render_template("ver_lead.html",
        lead=lead,
        equipos=equipos,
        oportunidades=oportunidades,
        todas_visitas=todas_visitas,
        administrador=administrador,
        tareas_comerciales=tareas_comerciales
    )


# ============================================
# EDITAR LEAD
# ============================================

@leads_bp.route('/editar_lead/<int:lead_id>', methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('clientes', 'write')
def editar(lead_id):
    """Editar información de un lead existente"""

    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get("administrador_id")
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        data = {
            "fecha_visita": request.form.get("fecha_visita") or None,
            "tipo_cliente": request.form.get("tipo_lead") or None,
            "direccion": request.form.get("direccion"),
            "nombre_cliente": request.form.get("nombre_lead") or None,
            "codigo_postal": request.form.get("codigo_postal") or None,
            "localidad": request.form.get("localidad"),
            "zona": request.form.get("zona") or None,
            "persona_contacto": request.form.get("persona_contacto") or None,
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "administrador_id": administrador_id,
            "empresa_mantenedora": request.form.get("empresa_mantenedora") or None,
            "numero_ascensores": request.form.get("numero_ascensores") or None,
            "fecha_fin_contrato": request.form.get("fecha_fin_contrato") or None,
            "paradas": request.form.get("paradas") or None,
            "viviendas_por_planta": request.form.get("viviendas_por_planta") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Solo dirección y localidad son obligatorios
        if not data["direccion"] or not data["localidad"]:
            flash_error("Dirección y Localidad son obligatorios")
            return redirect(request.referrer)

        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}",
            json=data,
            headers=HEADERS
        )
        if res.status_code in [200, 204]:
            return redirect(url_for('leads.ver', lead_id=lead_id))
        else:
            return f"<h3 style='color:red;'>Error al actualizar Lead</h3><pre>{res.text}</pre><a href='{url_for('leads.dashboard')}'>Volver</a>"

    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}",
        headers=HEADERS
    )
    if response.status_code == 200 and response.json():
        lead = response.json()[0]
        # LIMPIAR VALORES NONE PARA NO MOSTRAR "none" EN EL FORMULARIO
        lead = limpiar_none(lead)
    else:
        return f"<h3 style='color:red;'>Error al obtener Lead</h3><pre>{response.text}</pre><a href='{url_for('leads.dashboard')}'>Volver</a>"

    # Obtener lista de administradores (usando caché)
    administradores = get_administradores_cached()

    return render_template("editar_lead.html", lead=lead, administradores=administradores)


# ============================================
# ELIMINAR LEAD
# ============================================

@leads_bp.route('/eliminar_lead/<int:lead_id>')
@helpers.login_required
@helpers.requiere_permiso('clientes', 'delete')
def eliminar(lead_id):
    """Eliminar lead y todos sus datos relacionados"""

    requests.delete(f"{SUPABASE_URL}/rest/v1/equipos?cliente_id=eq.{lead_id}", headers=HEADERS)
    requests.delete(f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?cliente_id=eq.{lead_id}", headers=HEADERS)
    requests.delete(f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{lead_id}", headers=HEADERS)

    response = requests.delete(f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}", headers=HEADERS)

    if response.status_code in [200, 204]:
        flash_success("Lead eliminado correctamente")
        return redirect(url_for('leads.dashboard'))
    else:
        flash_error(f"Error al eliminar lead: {response.text}")
        return redirect(url_for('leads.ver', lead_id=lead_id))

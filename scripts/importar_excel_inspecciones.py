#!/usr/bin/env python3
"""
Script de Importación de Inspecciones desde Excel
AscensorAlert - Fedes Ascensores

Importa datos desde FICHERO_IPO_GLOBAL.xlsx a las tablas de Supabase:
- Hoja 1 (Principal): Inspecciones
- Hoja 2 (CORTINAS): Materiales especiales tipo CORTINA
- Hoja 3 (PESACARGAS): Materiales especiales tipo PESACARGA

USO:
    python scripts/importar_excel_inspecciones.py ruta/al/FICHERO_IPO_GLOBAL.xlsx

REQUISITOS:
    pip install openpyxl requests python-dateutil
"""

import sys
import os
import requests
from openpyxl import load_workbook
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

# Configuración de Supabase (usar variables de entorno en producción)
SUPABASE_URL = "https://hvkifqguxsgegzaxwcmj.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if not SUPABASE_KEY:
    print("❌ ERROR: Variable de entorno SUPABASE_KEY no configurada")
    sys.exit(1)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}


def parse_fecha(valor):
    """Convierte un valor de Excel en fecha formato YYYY-MM-DD"""
    if not valor or valor == "-":
        return None

    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')

    if isinstance(valor, str):
        # Intentar parsear diferentes formatos
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
            try:
                fecha = datetime.strptime(valor.strip(), fmt)
                return fecha.strftime('%Y-%m-%d')
            except ValueError:
                continue

    return None


def limpiar_texto(valor):
    """Limpia y normaliza valores de texto"""
    if not valor or valor == "-" or valor == "":
        return None

    if isinstance(valor, str):
        return valor.strip()

    return str(valor).strip() if valor else None


def get_or_create_oca(nombre_oca):
    """Obtiene o crea un OCA y devuelve su ID"""
    if not nombre_oca:
        return None

    nombre_limpio = limpiar_texto(nombre_oca)
    if not nombre_limpio:
        return None

    # Buscar OCA existente
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?nombre=eq.{nombre_limpio}",
        headers=HEADERS
    )

    if response.status_code == 200 and response.json():
        return response.json()[0]['id']

    # Crear nuevo OCA
    data = {
        "nombre": nombre_limpio,
        "activo": True
    }

    response = requests.post(
        f"{SUPABASE_URL}/rest/v1/ocas?select=id",
        json=data,
        headers=HEADERS
    )

    if response.status_code in [200, 201]:
        return response.json()[0]['id']

    print(f"⚠️  No se pudo crear OCA '{nombre_limpio}': {response.text}")
    return None


def importar_hoja_principal(wb):
    """
    Importa inspecciones de la hoja principal

    Columnas esperadas:
    - FECHA INSPECCION
    - PRESUPUESTO
    - ESTADO
    - Estado del material
    - Cliente
    - Instalación
    - Máquina
    - OCA
    """
    print("\n" + "="*60)
    print("📋 IMPORTANDO HOJA PRINCIPAL: Inspecciones")
    print("="*60)

    ws = wb.active  # Primera hoja

    # Leer encabezados (asumiendo que están en la primera fila)
    headers = [cell.value for cell in ws[1]]
    print(f"Encabezados encontrados: {headers}")

    inspecciones_importadas = 0
    errores = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Mapear valores (ajustar índices según tu Excel)
        # NOTA: Ajusta estos índices según la estructura real de tu Excel
        try:
            fecha_inspeccion_raw = row[0] if len(row) > 0 else None
            estado_presupuesto = row[2] if len(row) > 2 else "PENDIENTE"
            cliente = row[4] if len(row) > 4 else None
            instalacion = row[5] if len(row) > 5 else None
            maquina = row[6] if len(row) > 6 else None
            oca_nombre = row[7] if len(row) > 7 else None

            # Validar datos mínimos
            if not cliente or not instalacion:
                print(f"⏭️  Fila {idx}: Datos insuficientes (cliente o instalación vacíos)")
                continue

            fecha_inspeccion = parse_fecha(fecha_inspeccion_raw)
            if not fecha_inspeccion:
                print(f"⚠️  Fila {idx}: Fecha de inspección inválida, usando hoy")
                fecha_inspeccion = datetime.now().strftime('%Y-%m-%d')

            # Obtener o crear OCA
            oca_id = get_or_create_oca(oca_nombre) if oca_nombre else None

            # Normalizar estado de presupuesto
            estado_map = {
                "ACEPTADO": "ACEPTADO",
                "ENVIADO": "ENVIADO",
                "PREPARANDO": "PREPARANDO",
                None: "PENDIENTE",
                "": "PENDIENTE"
            }
            estado_presupuesto_normalizado = estado_map.get(estado_presupuesto, "PENDIENTE")

            # Crear inspección
            data = {
                "rae": maquina or f"RAE-IMPORT-{idx}",
                "fecha_inspeccion": fecha_inspeccion,
                "titular_nombre": limpiar_texto(cliente),
                "direccion_instalacion": limpiar_texto(instalacion),
                "oca_id": oca_id,
                "resultado": "Desfavorable",
                "tiene_defectos": True,
                "estado_presupuesto": estado_presupuesto_normalizado,
                "estado_trabajo": "PENDIENTE" if estado_presupuesto_normalizado != "ACEPTADO" else "EN_EJECUCION",
                "empresa_conservadora": "FEDES ASCENSORES",
                "observaciones": f"Importado desde Excel. Fila original: {idx}",
                "created_by": "IMPORT_SCRIPT"
            }

            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/inspecciones",
                json=data,
                headers=HEADERS
            )

            if response.status_code in [200, 201]:
                inspecciones_importadas += 1
                print(f"✅ Fila {idx}: {cliente} - {instalacion}")
            else:
                errores += 1
                print(f"❌ Fila {idx}: Error al importar - {response.text}")

        except Exception as e:
            errores += 1
            print(f"❌ Fila {idx}: Excepción - {str(e)}")

    print(f"\n📊 Resumen Hoja Principal:")
    print(f"   ✅ Inspecciones importadas: {inspecciones_importadas}")
    print(f"   ❌ Errores: {errores}")

    return inspecciones_importadas


def importar_cortinas(wb):
    """
    Importa cortinas fotoeléctricas de la hoja CORTINAS

    Columnas esperadas:
    - CLIENTE
    - CANTIDAD
    - FECHA LIMITE
    """
    print("\n" + "="*60)
    print("📋 IMPORTANDO HOJA: CORTINAS")
    print("="*60)

    try:
        ws = wb['CORTINAS']
    except KeyError:
        print("⚠️  Hoja 'CORTINAS' no encontrada en el Excel")
        return 0

    materiales_importados = 0
    errores = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            cliente = row[0] if len(row) > 0 else None
            cantidad = row[1] if len(row) > 1 else 1
            fecha_limite_raw = row[2] if len(row) > 2 else None

            if not cliente:
                print(f"⏭️  Fila {idx}: Cliente vacío")
                continue

            fecha_limite = parse_fecha(fecha_limite_raw)
            if not fecha_limite:
                # Si no hay fecha límite, calcular 6 meses desde hoy
                fecha_limite = (datetime.now() + relativedelta(months=6)).strftime('%Y-%m-%d')
                print(f"⚠️  Fila {idx}: Fecha límite calculada (6 meses desde hoy)")

            # Intentar convertir cantidad a int
            try:
                cantidad_int = int(cantidad) if cantidad else 1
            except (ValueError, TypeError):
                cantidad_int = 1

            data = {
                "tipo": "CORTINA",
                "cliente_nombre": limpiar_texto(cliente),
                "cantidad": cantidad_int,
                "fecha_limite": fecha_limite,
                "estado": "PENDIENTE",
                "observaciones": f"Importado desde Excel hoja CORTINAS. Fila: {idx}"
            }

            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/materiales_especiales",
                json=data,
                headers=HEADERS
            )

            if response.status_code in [200, 201]:
                materiales_importados += 1
                print(f"✅ Fila {idx}: CORTINA - {cliente} (x{cantidad_int})")
            else:
                errores += 1
                print(f"❌ Fila {idx}: Error - {response.text}")

        except Exception as e:
            errores += 1
            print(f"❌ Fila {idx}: Excepción - {str(e)}")

    print(f"\n📊 Resumen CORTINAS:")
    print(f"   ✅ Materiales importados: {materiales_importados}")
    print(f"   ❌ Errores: {errores}")

    return materiales_importados


def importar_pesacargas(wb):
    """
    Importa pesacargas de la hoja PESACARGAS

    Columnas esperadas:
    - CLIENTE
    - CANTIDAD
    - FECHA LIMITE
    """
    print("\n" + "="*60)
    print("📋 IMPORTANDO HOJA: PESACARGAS")
    print("="*60)

    try:
        ws = wb['PESACARGAS']
    except KeyError:
        print("⚠️  Hoja 'PESACARGAS' no encontrada en el Excel")
        return 0

    materiales_importados = 0
    errores = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            cliente = row[0] if len(row) > 0 else None
            cantidad = row[1] if len(row) > 1 else 1
            fecha_limite_raw = row[2] if len(row) > 2 else None

            if not cliente:
                print(f"⏭️  Fila {idx}: Cliente vacío")
                continue

            fecha_limite = parse_fecha(fecha_limite_raw)
            if not fecha_limite:
                # Si no hay fecha límite, calcular 6 meses desde hoy
                fecha_limite = (datetime.now() + relativedelta(months=6)).strftime('%Y-%m-%d')
                print(f"⚠️  Fila {idx}: Fecha límite calculada (6 meses desde hoy)")

            # Intentar convertir cantidad a int
            try:
                cantidad_int = int(cantidad) if cantidad else 1
            except (ValueError, TypeError):
                cantidad_int = 1

            data = {
                "tipo": "PESACARGA",
                "cliente_nombre": limpiar_texto(cliente),
                "cantidad": cantidad_int,
                "fecha_limite": fecha_limite,
                "estado": "PENDIENTE",
                "observaciones": f"Importado desde Excel hoja PESACARGAS. Fila: {idx}"
            }

            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/materiales_especiales",
                json=data,
                headers=HEADERS
            )

            if response.status_code in [200, 201]:
                materiales_importados += 1
                print(f"✅ Fila {idx}: PESACARGA - {cliente} (x{cantidad_int})")
            else:
                errores += 1
                print(f"❌ Fila {idx}: Error - {response.text}")

        except Exception as e:
            errores += 1
            print(f"❌ Fila {idx}: Excepción - {str(e)}")

    print(f"\n📊 Resumen PESACARGAS:")
    print(f"   ✅ Materiales importados: {materiales_importados}")
    print(f"   ❌ Errores: {errores}")

    return materiales_importados


def main():
    """Función principal de importación"""
    print("\n" + "🚀 "*30)
    print("   SCRIPT DE IMPORTACIÓN DE INSPECCIONES")
    print("   AscensorAlert - Fedes Ascensores")
    print("🚀 "*30 + "\n")

    # Validar argumentos
    if len(sys.argv) < 2:
        print("❌ ERROR: Debes proporcionar la ruta al archivo Excel")
        print("\nUso:")
        print(f"    python {sys.argv[0]} ruta/al/FICHERO_IPO_GLOBAL.xlsx")
        sys.exit(1)

    archivo_excel = sys.argv[1]

    # Validar que el archivo existe
    if not os.path.exists(archivo_excel):
        print(f"❌ ERROR: El archivo '{archivo_excel}' no existe")
        sys.exit(1)

    print(f"📁 Archivo a importar: {archivo_excel}\n")

    # Cargar workbook
    try:
        print("📂 Cargando archivo Excel...")
        wb = load_workbook(archivo_excel)
        print(f"✅ Excel cargado correctamente")
        print(f"   Hojas encontradas: {wb.sheetnames}\n")
    except Exception as e:
        print(f"❌ ERROR al cargar Excel: {str(e)}")
        sys.exit(1)

    # Importar cada hoja
    total_inspecciones = importar_hoja_principal(wb)
    total_cortinas = importar_cortinas(wb)
    total_pesacargas = importar_pesacargas(wb)

    # Resumen final
    print("\n" + "="*60)
    print("📊 RESUMEN GENERAL DE LA IMPORTACIÓN")
    print("="*60)
    print(f"✅ Inspecciones importadas: {total_inspecciones}")
    print(f"✅ Cortinas importadas: {total_cortinas}")
    print(f"✅ Pesacargas importadas: {total_pesacargas}")
    print(f"✅ TOTAL: {total_inspecciones + total_cortinas + total_pesacargas} registros")
    print("="*60)

    print("\n✅ Importación completada!")
    print("\n🌐 Accede a AscensorAlert para ver los datos importados:")
    print("   👉 /inspecciones")
    print("   👉 /materiales_especiales\n")


if __name__ == "__main__":
    main()

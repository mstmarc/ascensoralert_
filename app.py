from flask import Flask, request, render_template, redirect, session, Response, url_for, flash, send_file
import requests
import os
import urllib.parse
from datetime import date, datetime, timedelta
import calendar
import io
import resend
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
import pdfplumber
import logging
import sys
import helpers
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas

# Configurar logging para que aparezca en Render
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)

# Silenciar logs verbosos de pdfplumber/pdfminer
logging.getLogger('pdfminer').setLevel(logging.WARNING)
logging.getLogger('pdfplumber').setLevel(logging.WARNING)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY")
if not app.secret_key:
    raise RuntimeError("SECRET_KEY environment variable is not set")

# Filtro Jinja2 para formatear fechas a dd/mm/yyyy
@app.template_filter('format_fecha')
def format_fecha_filter(fecha_str):
    """Formatea fechas al formato dd/mm/yyyy para mostrar en templates"""
    if not fecha_str or fecha_str == "-":
        return "-"
    try:
        # Manejar timestamps ISO con T y timezone
        fecha_limpia = fecha_str.split('T')[0] if 'T' in str(fecha_str) else str(fecha_str)
        # Intentar parsear en formato ISO (YYYY-MM-DD)
        fecha = datetime.strptime(fecha_limpia, '%Y-%m-%d')
        return fecha.strftime('%d/%m/%Y')
    except:
        # Si ya está en formato dd/mm/yyyy o no se puede parsear, retornar como está
        return str(fecha_str) if fecha_str else "-"

# Datos de Supabase
SUPABASE_URL = "https://hvkifqguxsgegzaxwcmj.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")  # Service role key para storage

if not SUPABASE_KEY:
    raise RuntimeError("SUPABASE_KEY environment variable is not set")

# Headers para operaciones de base de datos (usa anon key)
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

# Headers para operaciones de storage (usa service key si está disponible, sino anon key)
STORAGE_KEY = SUPABASE_SERVICE_KEY if SUPABASE_SERVICE_KEY else SUPABASE_KEY
STORAGE_HEADERS = {
    "apikey": STORAGE_KEY,
    "Authorization": f"Bearer {STORAGE_KEY}",
}

# Configuración de Resend para emails
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "onboarding@resend.dev")
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# ============================================
# FUNCIONES DE CONTEXTO PARA TEMPLATES
# ============================================
# Registra funciones de permisos para que estén disponibles en Jinja2

@app.context_processor
def inject_permisos():
    """Inyecta funciones de control de acceso en todos los templates"""
    import json

    # Obtener perfil del usuario actual
    perfil_actual = helpers.obtener_perfil_usuario()

    # Construir diccionario de permisos para JavaScript
    permisos_js = {}
    if perfil_actual in helpers.PERMISOS_POR_PERFIL:
        for modulo, permisos in helpers.PERMISOS_POR_PERFIL[perfil_actual].items():
            permisos_js[modulo] = permisos

    return {
        'tiene_permiso': helpers.tiene_permiso,
        'puede_escribir': helpers.puede_escribir,
        'puede_eliminar': helpers.puede_eliminar,
        'obtener_perfil_usuario': helpers.obtener_perfil_usuario,
        'obtener_modulos_permitidos': helpers.obtener_modulos_permitidos,
        'perfil_usuario': perfil_actual,
        'permisos_usuario_json': json.dumps(permisos_js)
    }

# ============================================
# SISTEMA DE CACHÉ OPTIMIZADO
# ============================================
# Evita consultas repetidas a Supabase, mejorando el rendimiento

# Caché para administradores (5 min)
cache_administradores = {
    'data': [],
    'timestamp': None
}

# Caché para métricas del dashboard home (5 min)
cache_metricas_home = {
    'data': None,
    'timestamp': None
}

# Caché para filtros de localidades y empresas (30 min - cambian poco)
cache_filtros = {
    'localidades': [],
    'empresas': [],
    'timestamp': None
}

# Caché para últimas instalaciones (10 min)
cache_ultimas_instalaciones = {
    'data': [],
    'timestamp': None
}

# Caché para últimas oportunidades (10 min)
cache_ultimas_oportunidades = {
    'data': [],
    'timestamp': None
}

def get_administradores_cached():
    """
    Obtiene la lista de administradores usando caché.
    Se renueva automáticamente cada 5 minutos.
    Esto reduce drásticamente las consultas a Supabase.
    """
    now = datetime.now()

    # Si no hay caché o pasaron 5 minutos, renovar
    if not cache_administradores['timestamp'] or \
       (now - cache_administradores['timestamp']) > timedelta(minutes=5):

        try:
            print(f"🔄 Consultando administradores desde Supabase...")
            response = requests.get(
                f"{SUPABASE_URL}/rest/v1/administradores?select=id,nombre_empresa&order=nombre_empresa.asc",
                headers=HEADERS,
                timeout=10  # Aumentado a 10 segundos
            )

            print(f"📡 Respuesta de Supabase - Status: {response.status_code}")

            if response.status_code == 200:
                data = response.json()
                cache_administradores['data'] = data
                cache_administradores['timestamp'] = now
                print(f"✅ Caché de administradores actualizado: {len(data)} registros")
            else:
                print(f"⚠️ Error al actualizar caché de administradores: {response.status_code}")
                print(f"📄 Respuesta: {response.text[:200]}")
                # Si falla pero hay caché previo, usarlo
                if cache_administradores['data']:
                    print(f"ℹ️ Usando caché anterior: {len(cache_administradores['data'])} registros")

        except requests.exceptions.Timeout:
            print(f"⏱️ Timeout al consultar Supabase (>10s)")
        except Exception as e:
            print(f"❌ Excepción al actualizar caché de administradores: {type(e).__name__}: {str(e)}")
            # Si falla, devolver lo que haya en caché (aunque esté desactualizado)

    return cache_administradores['data']

def get_metricas_home_cached():
    """
    Obtiene las métricas del dashboard home usando caché.
    Se renueva cada 5 minutos.
    """
    now = datetime.now()

    if not cache_metricas_home['timestamp'] or \
       (now - cache_metricas_home['timestamp']) > timedelta(minutes=5):

        try:
            print(f"🔄 Consultando métricas del home desde Supabase...")

            # Obtener todas las métricas en paralelo sería ideal, pero las hacemos secuenciales
            metricas = {}

            # Total clientes
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?select=id", headers=HEADERS, timeout=10)
            metricas['total_clientes'] = len(resp.json()) if resp.ok else 0

            # Total equipos
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?select=id", headers=HEADERS, timeout=10)
            metricas['total_equipos'] = len(resp.json()) if resp.ok else 0

            # Total oportunidades
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/oportunidades?select=id", headers=HEADERS, timeout=10)
            metricas['total_oportunidades'] = len(resp.json()) if resp.ok else 0

            # IPOs de hoy
            hoy = datetime.now().strftime("%Y-%m-%d")
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?select=id&ipo_proxima=eq.{hoy}", headers=HEADERS, timeout=10)
            metricas['ipos_hoy'] = len(resp.json()) if resp.ok else 0

            # Contratos por vencer (próximos 30 días)
            fecha_inicio = datetime.now().strftime("%Y-%m-%d")
            fecha_fin = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/equipos?select=id&fecha_vencimiento_contrato=gte.{fecha_inicio}&fecha_vencimiento_contrato=lte.{fecha_fin}",
                headers=HEADERS, timeout=10
            )
            metricas['contratos_vencer'] = len(resp.json()) if resp.ok else 0

            # IPOs de esta semana
            fecha_fin_semana = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/equipos?select=id&ipo_proxima=gte.{fecha_inicio}&ipo_proxima=lte.{fecha_fin_semana}",
                headers=HEADERS, timeout=10
            )
            metricas['ipos_semana'] = len(resp.json()) if resp.ok else 0

            # Oportunidades pendientes
            resp = requests.get(
                f"{SUPABASE_URL}/rest/v1/oportunidades?select=id&estado=eq.activa",
                headers=HEADERS, timeout=10
            )
            metricas['oportunidades_pendientes'] = len(resp.json()) if resp.ok else 0

            cache_metricas_home['data'] = metricas
            cache_metricas_home['timestamp'] = now
            print(f"✅ Caché de métricas home actualizado")

        except Exception as e:
            print(f"❌ Error al actualizar caché de métricas: {type(e).__name__}: {str(e)}")
            # Si falla, devolver lo que haya en caché

    return cache_metricas_home['data']

def get_filtros_cached():
    """
    Obtiene los filtros (localidades y empresas) usando caché.
    Se renueva cada 30 minutos (cambian poco).
    """
    now = datetime.now()

    if not cache_filtros['timestamp'] or \
       (now - cache_filtros['timestamp']) > timedelta(minutes=30):

        try:
            print(f"🔄 Consultando filtros desde Supabase...")

            localidades = set()
            empresas = set()

            # Localidades
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?select=localidad", headers=HEADERS, timeout=10)
            if resp.ok:
                for item in resp.json():
                    if item.get("localidad"):
                        localidades.add(item["localidad"])

            # Empresas
            resp = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?select=empresa_mantenedora", headers=HEADERS, timeout=10)
            if resp.ok:
                for item in resp.json():
                    if item.get("empresa_mantenedora"):
                        empresas.add(item["empresa_mantenedora"])

            cache_filtros['localidades'] = sorted(list(localidades))
            cache_filtros['empresas'] = sorted(list(empresas))
            cache_filtros['timestamp'] = now
            print(f"✅ Caché de filtros actualizado: {len(localidades)} localidades, {len(empresas)} empresas")

        except Exception as e:
            print(f"❌ Error al actualizar caché de filtros: {type(e).__name__}: {str(e)}")

    return cache_filtros['localidades'], cache_filtros['empresas']

def get_ultimas_instalaciones_cached():
    """
    Obtiene las últimas instalaciones usando caché.
    Se renueva cada 10 minutos.
    """
    now = datetime.now()

    if not cache_ultimas_instalaciones['timestamp'] or \
       (now - cache_ultimas_instalaciones['timestamp']) > timedelta(minutes=10):

        try:
            print(f"🔄 Consultando últimas instalaciones desde Supabase...")

            # OPTIMIZACIÓN: Seleccionar solo campos necesarios en lugar de *
            response = requests.get(
                f"{SUPABASE_URL}/rest/v1/clientes?select=id,direccion,nombre_cliente,localidad,empresa_mantenedora,numero_ascensores,equipos(id)&order=fecha_visita.desc&limit=5",
                headers=HEADERS,
                timeout=10
            )

            if response.ok:
                leads_data = response.json()
                instalaciones = []

                for lead in leads_data:
                    equipos_data = lead.get('equipos', [])
                    num_equipos = len(equipos_data) if equipos_data else lead.get('numero_ascensores', 0)
                    empresa_mantenedora = lead.get('empresa_mantenedora', '-')

                    instalaciones.append({
                        'id': lead['id'],
                        'direccion': lead.get('direccion', 'Sin dirección'),
                        'nombre_cliente': lead.get('nombre_cliente', ''),
                        'localidad': lead.get('localidad', '-'),
                        'num_equipos': num_equipos,
                        'empresa_mantenedora': empresa_mantenedora
                    })

                cache_ultimas_instalaciones['data'] = instalaciones
                cache_ultimas_instalaciones['timestamp'] = now
                print(f"✅ Caché de últimas instalaciones actualizado: {len(instalaciones)} registros")

        except Exception as e:
            print(f"❌ Error al actualizar caché de instalaciones: {type(e).__name__}: {str(e)}")

    return cache_ultimas_instalaciones['data']

def get_ultimas_oportunidades_cached():
    """
    Obtiene las últimas oportunidades usando caché.
    Se renueva cada 10 minutos.
    """
    now = datetime.now()

    if not cache_ultimas_oportunidades['timestamp'] or \
       (now - cache_ultimas_oportunidades['timestamp']) > timedelta(minutes=10):

        try:
            print(f"🔄 Consultando últimas oportunidades desde Supabase...")

            # OPTIMIZACIÓN: Seleccionar solo campos necesarios en lugar de *
            response = requests.get(
                f"{SUPABASE_URL}/rest/v1/oportunidades?select=id,tipo,estado,clientes(nombre_cliente,direccion)&order=fecha_creacion.desc&limit=5",
                headers=HEADERS,
                timeout=10
            )

            if response.ok:
                cache_ultimas_oportunidades['data'] = response.json()
                cache_ultimas_oportunidades['timestamp'] = now
                print(f"✅ Caché de últimas oportunidades actualizado: {len(cache_ultimas_oportunidades['data'])} registros")

        except Exception as e:
            print(f"❌ Error al actualizar caché de oportunidades: {type(e).__name__}: {str(e)}")

    return cache_ultimas_oportunidades['data']

# FUNCIONES AUXILIARES

def limpiar_none(data):
    """Convierte valores None a strings vacíos para evitar mostrar 'none' en formularios"""
    if isinstance(data, dict):
        return {k: (v if v is not None else '') for k, v in data.items()}
    return data

def calcular_color_ipo(fecha_ipo_str):
    """Calcula el color de fondo para la celda de IPO según la urgencia"""
    if not fecha_ipo_str or fecha_ipo_str == "-":
        return ""
    
    try:
        fecha_ipo = datetime.strptime(fecha_ipo_str, "%d/%m/%Y")
        hoy = datetime.now()
        diferencia = (fecha_ipo - hoy).days
        
        if -15 <= diferencia < 0:
            return "background-color: #FFF59D;"
        
        if diferencia >= 0 and diferencia <= 30:
            return "background-color: #FFCDD2;"
        
        return ""
    except:
        return ""

def calcular_color_contrato(fecha_contrato_str):
    """Calcula el color de fondo para la celda de contrato según vencimiento"""
    if not fecha_contrato_str or fecha_contrato_str == "-":
        return ""
    
    try:
        fecha_contrato = datetime.strptime(fecha_contrato_str, "%d/%m/%Y")
        hoy = datetime.now()
        diferencia = (fecha_contrato - hoy).days
        
        if diferencia <= 30:
            return "background-color: #FFCDD2;"
        
        if 30 < diferencia <= 90:
            return "background-color: #FFF59D;"
        
        return ""
    except:
        return ""

def enviar_avisos_email(config):
    """Función que revisa fechas y envía emails"""
    
    if not RESEND_API_KEY:
        return "Error: No se ha configurado RESEND_API_KEY"
    
    # Procesar múltiples emails
    emails_raw = config['email_destinatario']
    emails_destino = [email.strip() for email in emails_raw.split(',')] if ',' in emails_raw else [emails_raw.strip()]
    primer_aviso_ipo = config['primer_aviso_despues_ipo']
    segundo_aviso_ipo = config['segundo_aviso_despues_ipo']
    dias_contrato = config['dias_aviso_antes_contrato']
    
    fecha_hoy = datetime.now().date()
    
    # Obtener equipos con IPO próxima
    equipos_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?select=*,clientes(nombre_cliente,direccion)&ipo_proxima=not.is.null",
        headers=HEADERS
    )
    
    alertas_ipo_primer_aviso = []
    alertas_ipo_segundo_aviso = []
    alertas_contrato = []
    
    if equipos_response.status_code == 200:
        equipos = equipos_response.json()
        
        for equipo in equipos:
            # Revisar IPOs - SOLO DESPUÉS de la fecha
            if equipo.get('ipo_proxima'):
                try:
                    fecha_ipo = datetime.strptime(equipo['ipo_proxima'], '%Y-%m-%d').date()
                    dias_desde_ipo = (fecha_hoy - fecha_ipo).days
                    
                    cliente = equipo.get('clientes', {})
                    if isinstance(cliente, list) and cliente:
                        cliente = cliente[0]
                    
                    equipo_data = {
                        'cliente': cliente.get('nombre_cliente', 'Sin nombre') if cliente else 'Sin nombre',
                        'direccion': cliente.get('direccion', 'Sin dirección') if cliente else 'Sin dirección',
                        'identificacion': equipo.get('identificacion', 'N/A'),
                        'fecha': fecha_ipo.strftime('%d/%m/%Y'),
                        'dias_desde_ipo': dias_desde_ipo
                    }
                    
                    # Primer aviso (ej: día 15)
                    if dias_desde_ipo == primer_aviso_ipo:
                        alertas_ipo_primer_aviso.append(equipo_data)
                    
                    # Segundo aviso (ej: día 30)
                    if dias_desde_ipo == segundo_aviso_ipo:
                        alertas_ipo_segundo_aviso.append(equipo_data)
                        
                except:
                    pass
            
            # Revisar contratos - X días ANTES del vencimiento
            if equipo.get('fecha_vencimiento_contrato'):
                try:
                    fecha_contrato = datetime.strptime(equipo['fecha_vencimiento_contrato'], '%Y-%m-%d').date()
                    dias_restantes = (fecha_contrato - fecha_hoy).days
                    
                    if 0 <= dias_restantes <= dias_contrato:
                        cliente = equipo.get('clientes', {})
                        if isinstance(cliente, list) and cliente:
                            cliente = cliente[0]
                        
                        alertas_contrato.append({
                            'cliente': cliente.get('nombre_cliente', 'Sin nombre') if cliente else 'Sin nombre',
                            'direccion': cliente.get('direccion', 'Sin dirección') if cliente else 'Sin dirección',
                            'identificacion': equipo.get('identificacion', 'N/A'),
                            'fecha': fecha_contrato.strftime('%d/%m/%Y'),
                            'dias_restantes': dias_restantes
                        })
                except:
                    pass
    
    # Si no hay alertas, no enviar email
    if not alertas_ipo_primer_aviso and not alertas_ipo_segundo_aviso and not alertas_contrato:
        return "No hay avisos pendientes"
    
    # Construir contenido del email
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Montserrat', Arial, sans-serif; background: white; margin: 0; padding: 20px; }}
            .container {{ max-width: 800px; margin: 0 auto; background: white; }}
            h1 {{ color: #366092; border-bottom: 3px solid #366092; padding-bottom: 10px; }}
            h2 {{ color: #366092; margin-top: 30px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background: #366092; color: white; padding: 12px; text-align: left; }}
            td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
            .urgente {{ color: #dc3545; font-weight: bold; }}
            .proximo {{ color: #ffc107; font-weight: bold; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔔 Avisos AscensorAlert - Fedes Ascensores</h1>
            <p><strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y')}</p>
    """
    
    if alertas_ipo_primer_aviso:
        html_content += f"""
            <h2>🔍 IPOs - Primer Aviso ({primer_aviso_ipo} días después)</h2>
            <p style="color: #666;">Es momento de visitar para verificar la subsanación de defectos</p>
            <table>
                <tr>
                    <th>Cliente</th>
                    <th>Dirección</th>
                    <th>Equipo</th>
                    <th>Fecha IPO</th>
                    <th>Días Transcurridos</th>
                </tr>
        """
        for alerta in alertas_ipo_primer_aviso:
            html_content += f"""
                <tr>
                    <td>{alerta['cliente']}</td>
                    <td>{alerta['direccion']}</td>
                    <td>{alerta['identificacion']}</td>
                    <td>{alerta['fecha']}</td>
                    <td class="proximo">{alerta['dias_desde_ipo']} días</td>
                </tr>
            """
        html_content += "</table>"
    
    if alertas_ipo_segundo_aviso:
        html_content += f"""
            <h2>⚠️ IPOs - Segundo Aviso ({segundo_aviso_ipo} días después)</h2>
            <p style="color: #dc3545; font-weight: bold;">¡URGENTE! Últimos días para verificar subsanación</p>
            <table>
                <tr>
                    <th>Cliente</th>
                    <th>Dirección</th>
                    <th>Equipo</th>
                    <th>Fecha IPO</th>
                    <th>Días Transcurridos</th>
                </tr>
        """
        for alerta in alertas_ipo_segundo_aviso:
            html_content += f"""
                <tr>
                    <td>{alerta['cliente']}</td>
                    <td>{alerta['direccion']}</td>
                    <td>{alerta['identificacion']}</td>
                    <td>{alerta['fecha']}</td>
                    <td class="urgente">{alerta['dias_desde_ipo']} días</td>
                </tr>
            """
        html_content += "</table>"
    
    if alertas_contrato:
        html_content += """
            <h2>📋 Contratos por Vencer</h2>
            <table>
                <tr>
                    <th>Cliente</th>
                    <th>Dirección</th>
                    <th>Equipo</th>
                    <th>Vencimiento</th>
                    <th>Días Restantes</th>
                </tr>
        """
        for alerta in alertas_contrato:
            clase = 'urgente' if alerta['dias_restantes'] <= 15 else 'proximo'
            html_content += f"""
                <tr>
                    <td>{alerta['cliente']}</td>
                    <td>{alerta['direccion']}</td>
                    <td>{alerta['identificacion']}</td>
                    <td>{alerta['fecha']}</td>
                    <td class="{clase}">{alerta['dias_restantes']} días</td>
                </tr>
            """
        html_content += "</table>"
    
    html_content += """
            <p style="margin-top: 30px; color: #666; font-size: 14px;">
                Este es un email automático generado por AscensorAlert.<br>
                Para gestionar tus alertas, accede a la configuración en la aplicación.
            </p>
        </div>
    </body>
    </html>
    """
    
    total_ipos = len(alertas_ipo_primer_aviso) + len(alertas_ipo_segundo_aviso)
    
    # Enviar email con Resend
    try:
        params = {
            "from": EMAIL_FROM,
            "to": emails_destino,
            "subject": f"🔔 Avisos AscensorAlert - {total_ipos} IPOs y {len(alertas_contrato)} Contratos",
            "html": html_content
        }
        
        email = resend.Emails.send(params)
        return f"Email enviado correctamente a {len(emails_destino)} destinatario(s): {total_ipos} IPOs, {len(alertas_contrato)} contratos"
    except Exception as e:
        return f"Error al enviar email: {str(e)}"

# ============================================
# RUTAS
# ============================================

# Login
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario")
        contrasena = request.form.get("contrasena")
        if not usuario or not contrasena:
            return render_template("login.html", error="Usuario y contraseña requeridos")
        encoded_user = urllib.parse.quote(usuario, safe="")
        query = f"?nombre_usuario=eq.{encoded_user}"
        response = requests.get(f"{SUPABASE_URL}/rest/v1/usuarios{query}", headers=HEADERS)

        if response.status_code == 200 and len(response.json()) == 1:
            user = response.json()[0]
            if user.get("contrasena", "") == contrasena:
                session["usuario"] = usuario
                session["usuario_id"] = user.get("id")
                session["email"] = user.get("email", "")
                # Cargar perfil del usuario (por defecto 'visualizador' si no existe)
                session["perfil"] = user.get("perfil", "visualizador")
                return redirect("/home")
        return render_template("login.html", error="Usuario o contraseña incorrectos")
    return render_template("login.html", error=None)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# Home - ACTUALIZADO: Dashboard en desktop, menú simple en móvil
@app.route("/home")
def home():
    """Homepage - Dashboard responsive para todos los dispositivos"""
    if "usuario" not in session:
        return redirect("/")

    # Dashboard responsive (funciona en desktop, tablet y móvil)

    # Variables de fecha usadas en varias secciones
    hoy = date.today().isoformat()
    fin_semana = (date.today() + timedelta(days=7)).isoformat()

    # ========== MÉTRICAS Y ALERTAS (OPTIMIZADO CON CACHÉ) ==========
    # Usar caché de métricas (TTL: 5 minutos)
    # Reduce de 10+ queries a 0 queries en cargas subsecuentes
    metricas_cached = get_metricas_home_cached()

    if metricas_cached:
        metricas = {
            'total_comunidades': metricas_cached.get('total_clientes', 0),
            'total_equipos': metricas_cached.get('total_equipos', 0),
            'total_oportunidades': metricas_cached.get('total_oportunidades', 0),
            'ipos_hoy': metricas_cached.get('ipos_hoy', 0)
        }

        alertas = {
            'contratos_criticos': metricas_cached.get('contratos_vencer', 0),
            'ipos_semana': metricas_cached.get('ipos_semana', 0),
            'oportunidades_pendientes': metricas_cached.get('oportunidades_pendientes', 0)
        }
    else:
        # Fallback si el caché falla (no debería pasar)
        metricas = {'total_comunidades': 0, 'total_equipos': 0, 'total_oportunidades': 0, 'ipos_hoy': 0}
        alertas = {'contratos_criticos': 0, 'ipos_semana': 0, 'oportunidades_pendientes': 0}

    # ========== ÚLTIMAS INSTALACIONES (OPTIMIZADO CON CACHÉ) ==========
    # Usar caché de instalaciones (TTL: 10 minutos)
    # Reduce de 1 query a 0 queries en cargas subsecuentes
    ultimas_instalaciones = get_ultimas_instalaciones_cached()

    # ========== ÚLTIMAS OPORTUNIDADES (OPTIMIZADO CON CACHÉ) ==========
    # Usar caché de oportunidades (TTL: 10 minutos)
    # Reduce de 1 query a 0 queries en cargas subsecuentes
    oportunidades_cached = get_ultimas_oportunidades_cached()

    ultimas_oportunidades = []
    if oportunidades_cached:
        for op in oportunidades_cached:
            # Obtener nombre y dirección del cliente de la relación
            cliente_info = op.get('clientes', {})
            if isinstance(cliente_info, list) and len(cliente_info) > 0:
                cliente_info = cliente_info[0]

            nombre_cliente = cliente_info.get('nombre_cliente', 'Sin nombre') if cliente_info else 'Sin nombre'
            direccion_cliente = cliente_info.get('direccion', 'Sin dirección') if cliente_info else 'Sin dirección'

            ultimas_oportunidades.append({
                'id': op['id'],
                'nombre_cliente': nombre_cliente,
                'direccion': direccion_cliente,
                'tipo': op.get('tipo', '-'),
                'estado': op.get('estado', '-')
            })
    
    # ========== PRÓXIMAS IPOs ESTA SEMANA ==========
    # OPTIMIZACIÓN: Seleccionar solo campos necesarios en lugar de *
    response_ipos = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?select=cliente_id,ipo_proxima,clientes(direccion,localidad)&ipo_proxima=gte.{hoy}&ipo_proxima=lte.{fin_semana}&order=ipo_proxima.asc",
        headers=HEADERS
    )
    
    proximas_ipos = []
    if response_ipos.ok:
        for equipo in response_ipos.json():
            fecha_ipo_str = equipo.get('ipo_proxima', '')
            if fecha_ipo_str:
                try:
                    fecha_ipo = datetime.strptime(fecha_ipo_str, '%Y-%m-%d').date()
                    dias_restantes = (fecha_ipo - date.today()).days
                    
                    # Obtener dirección del lead
                    lead_info = equipo.get('clientes', {})
                    if isinstance(lead_info, list) and len(lead_info) > 0:
                        lead_info = lead_info[0]
                    
                    proximas_ipos.append({
                        'lead_id': equipo.get('cliente_id'),
                        'direccion': lead_info.get('direccion', 'Sin dirección') if lead_info else 'Sin dirección',
                        'localidad': lead_info.get('localidad', '-') if lead_info else '-',
                        'fecha_ipo': fecha_ipo.strftime('%d/%m/%Y'),
                        'dias_restantes': dias_restantes
                    })
                except Exception as e:
                    print(f"Error procesando IPO: {e}")
                    continue
    
    return render_template(
        'home.html',
        metricas=metricas,
        alertas=alertas,
        ultimas_instalaciones=ultimas_instalaciones,
        ultimas_oportunidades=ultimas_oportunidades,
        proximas_ipos=proximas_ipos
    )

# Alta de Lead CON FECHA DE VISITA
@app.route("/formulario_lead", methods=["GET", "POST"])
def formulario_lead():
    if "usuario" not in session:
        return redirect("/")
    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get("administrador_id")
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        data = {
            "fecha_visita": request.form.get("fecha_visita") or None,
            "tipo_cliente": request.form.get("tipo_lead") or None,
            "direccion": request.form.get("direccion"),
            "nombre_cliente": request.form.get("nombre_lead") or None,
            "codigo_postal": request.form.get("codigo_postal") or None,
            "localidad": request.form.get("localidad"),
            "zona": request.form.get("zona") or None,
            "persona_contacto": request.form.get("persona_contacto") or None,
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "administrador_id": administrador_id,
            "empresa_mantenedora": request.form.get("empresa_mantenedora") or None,
            "numero_ascensores": request.form.get("numero_ascensores") or None,
            "fecha_fin_contrato": request.form.get("fecha_fin_contrato") or None,
            "paradas": request.form.get("paradas") or None,
            "viviendas_por_planta": request.form.get("viviendas_por_planta") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Solo dirección y localidad son obligatorios
        required = [data["direccion"], data["localidad"]]
        if any(not field for field in required):
            return "Datos del lead inválidos - Dirección y Localidad son obligatorios", 400

        response = requests.post(f"{SUPABASE_URL}/rest/v1/clientes?select=id", json=data, headers=HEADERS)
        if response.status_code in [200, 201]:
            cliente_id = response.json()[0]["id"]
            return redirect(f"/nuevo_equipo?lead_id={cliente_id}")
        else:
            return f"<h3 style='color:red;'>Error al registrar lead</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

    # GET - Obtener lista de administradores (usando caché)
    fecha_hoy = date.today().strftime('%Y-%m-%d')
    administradores = get_administradores_cached()

    return render_template("formulario_lead.html", fecha_hoy=fecha_hoy, administradores=administradores)

# Visita a Administrador
@app.route("/visita_administrador", methods=["GET", "POST"])
def visita_administrador():
    if "usuario" not in session:
        return redirect("/")
    
    # Puede venir con oportunidad_id desde la vista de oportunidad
    oportunidad_id = request.args.get("oportunidad_id")
    oportunidad_data = None
    
    if oportunidad_id:
        # Obtener datos de la oportunidad
        oportunidad_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=*,clientes(*)",
            headers=HEADERS
        )
        if oportunidad_response.status_code == 200 and oportunidad_response.json():
            oportunidad_data = oportunidad_response.json()[0]

    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get('administrador_id')
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        # Validación: fecha y administrador son obligatorios
        if not request.form.get("fecha_visita") or not administrador_id:
            flash("Fecha y Administrador son obligatorios", "error")
            return redirect(request.referrer)

        # Buscar el nombre del administrador para el campo administrador_fincas (NOT NULL en BD)
        administrador_nombre = None
        if administrador_id:
            admin_response = requests.get(
                f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{administrador_id}&select=nombre_empresa",
                headers=HEADERS
            )
            if admin_response.status_code == 200 and admin_response.json():
                administrador_nombre = admin_response.json()[0].get("nombre_empresa")
            else:
                flash(f"No se encontró el administrador seleccionado", "error")
                return redirect(request.referrer)

        data = {
            "fecha_visita": request.form.get("fecha_visita"),
            "administrador_id": administrador_id,
            "administrador_fincas": administrador_nombre,  # Campo NOT NULL en BD
            "persona_contacto": request.form.get("persona_contacto") or None,
            "observaciones": request.form.get("observaciones") or None,
            "oportunidad_id": int(request.form.get("oportunidad_id")) if request.form.get("oportunidad_id") else None
        }

        response = requests.post(f"{SUPABASE_URL}/rest/v1/visitas_administradores", json=data, headers=HEADERS)

        if response.status_code in [200, 201]:
            # Si viene de una oportunidad, volver a la oportunidad
            if data["oportunidad_id"]:
                flash("Visita a administrador registrada correctamente", "success")
                return redirect(url_for('ver_oportunidad', oportunidad_id=data["oportunidad_id"]))
            else:
                return render_template("visita_admin_success.html")
        else:
            flash(f"Error al registrar visita: {response.text}", "error")
            return redirect(request.referrer)

    # GET - Obtener lista de administradores (usando caché)
    fecha_hoy = date.today().strftime('%Y-%m-%d')
    administradores = get_administradores_cached()

    return render_template("visita_administrador.html",
                         fecha_hoy=fecha_hoy,
                         oportunidad=oportunidad_data,
                         administradores=administradores)

# Dashboard de visitas a administradores
@app.route("/visitas_administradores_dashboard")
def visitas_administradores_dashboard():
    if "usuario" not in session:
        return redirect("/")
    
    page = int(request.args.get("page", 1))
    per_page = 25
    offset = (page - 1) * per_page

    # OPTIMIZACIÓN: Para count solo necesitamos id, no todos los campos
    count_url = f"{SUPABASE_URL}/rest/v1/visitas_administradores?select=id"
    count_response = requests.get(count_url, headers={**HEADERS, "Prefer": "count=exact"})
    total_registros = int(count_response.headers.get("Content-Range", "0").split("/")[-1])
    total_pages = max(1, (total_registros + per_page - 1) // per_page)

    # OPTIMIZACIÓN: Seleccionar campos específicos con JOIN a administradores
    data_url = f"{SUPABASE_URL}/rest/v1/visitas_administradores?select=id,fecha_visita,administrador_id,administradores(nombre_empresa),persona_contacto,observaciones,oportunidad_id&order=fecha_visita.desc&limit={per_page}&offset={offset}"
    response = requests.get(data_url, headers=HEADERS)
    
    if response.status_code != 200:
        return f"<h3 style='color:red;'>Error al obtener visitas</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"
    
    visitas = response.json()
    
    return render_template("visitas_admin_dashboard.html",
        visitas=visitas,
        page=page,
        total_pages=total_pages,
        total_registros=total_registros
    )

@app.route("/ver_visita_admin/<int:visita_id>")
def ver_visita_admin(visita_id):
    if "usuario" not in session:
        return redirect("/")

    # Obtener visita con JOIN a administradores
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/visitas_administradores?id=eq.{visita_id}&select=*,administradores(nombre_empresa)",
        headers=HEADERS
    )
    if response.status_code != 200 or not response.json():
        flash("Visita no encontrada", "error")
        return redirect("/visitas_administradores_dashboard")

    visita = limpiar_none(response.json()[0])
    return render_template("ver_visita_admin.html", visita=visita)

@app.route("/editar_visita_admin/<int:visita_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('visitas', 'write')
def editar_visita_admin(visita_id):

    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get("administrador_id")
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        # Buscar el nombre del administrador para el campo administrador_fincas (NOT NULL en BD)
        administrador_nombre = None
        if administrador_id:
            admin_response = requests.get(
                f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{administrador_id}&select=nombre_empresa",
                headers=HEADERS
            )
            if admin_response.status_code == 200 and admin_response.json():
                administrador_nombre = admin_response.json()[0].get("nombre_empresa")

        data = {
            "fecha_visita": request.form.get("fecha_visita"),
            "administrador_id": administrador_id,
            "administrador_fincas": administrador_nombre,  # Campo NOT NULL en BD
            "persona_contacto": request.form.get("persona_contacto") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/visitas_administradores?id=eq.{visita_id}",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 204]:
            flash("Visita actualizada correctamente", "success")
            return redirect("/visitas_administradores_dashboard")
        else:
            flash("Error al actualizar visita", "error")

    response = requests.get(f"{SUPABASE_URL}/rest/v1/visitas_administradores?id=eq.{visita_id}", headers=HEADERS)
    if response.status_code != 200 or not response.json():
        flash("Visita no encontrada", "error")
        return redirect("/visitas_administradores_dashboard")

    visita = limpiar_none(response.json()[0])

    # Obtener lista de administradores (usando caché)
    administradores = get_administradores_cached()

    return render_template("editar_visita_admin.html", visita=visita, administradores=administradores)

@app.route("/eliminar_visita_admin/<int:visita_id>")
@helpers.login_required
@helpers.requiere_permiso('visitas', 'delete')
def eliminar_visita_admin(visita_id):
    
    response = requests.delete(f"{SUPABASE_URL}/rest/v1/visitas_administradores?id=eq.{visita_id}", headers=HEADERS)
    
    if response.status_code in [200, 204]:
        flash("Visita eliminada correctamente", "success")
    else:
        flash("Error al eliminar visita", "error")
    
    return redirect("/visitas_administradores_dashboard")

# Alta de Equipo (ahora requiere lead_id)
@app.route("/nuevo_equipo", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('equipos', 'write')
def nuevo_equipo():
    
    # VALIDAR que viene con lead_id
    lead_id = request.args.get("lead_id")
    
    if not lead_id:
        flash("Debes añadir equipos desde un lead específico", "error")
        return redirect("/leads_dashboard")
    
    # Verificar que el lead existe
    lead_url = f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}"
    lead_response = requests.get(lead_url, headers=HEADERS)
    
    if lead_response.status_code != 200 or not lead_response.json():
        flash("Lead no encontrado", "error")
        return redirect("/leads_dashboard")
    
    lead_data = lead_response.json()[0]
    # Limpiar valores None para evitar mostrar "none"
    lead_data = limpiar_none(lead_data)
    
    if request.method == "POST":
        equipo_data = {
            "cliente_id": int(lead_id),
            "tipo_equipo": request.form.get("tipo_equipo"),
            "identificacion": request.form.get("identificacion") or None,
            "descripcion": request.form.get("observaciones") or None,
            "fecha_vencimiento_contrato": request.form.get("fecha_vencimiento_contrato") or None,
            "rae": request.form.get("rae") or None,
            "ipo_proxima": request.form.get("ipo_proxima") or None
        }

        # Solo tipo_equipo es obligatorio
        if not equipo_data["tipo_equipo"]:
            return render_template("nuevo_equipo.html", 
                                 lead=lead_data, 
                                 error="Tipo de equipo es obligatorio")

        res = requests.post(f"{SUPABASE_URL}/rest/v1/equipos", json=equipo_data, headers=HEADERS)
        if res.status_code in [200, 201]:
            flash("Equipo añadido correctamente", "success")
            return redirect(f"/ver_lead/{lead_id}")
        else:
            return render_template("nuevo_equipo.html", 
                                 lead=lead_data, 
                                 error="Error al crear el equipo")

    return render_template("nuevo_equipo.html", lead=lead_data)

# Crear visita de seguimiento
@app.route("/crear_visita_seguimiento/<int:cliente_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('visitas', 'write')
def crear_visita_seguimiento(cliente_id):
    
    oportunidad_id = request.args.get("oportunidad_id")
    
    if request.method == "POST":
        data = {
            "cliente_id": cliente_id,
            "oportunidad_id": int(request.form.get("oportunidad_id")) if request.form.get("oportunidad_id") else None,
            "fecha_visita": request.form.get("fecha_visita"),
            "observaciones": request.form.get("observaciones")
        }
        
        if not data["fecha_visita"]:
            flash("La fecha de visita es obligatoria", "error")
            return redirect(request.referrer)
        
        response = requests.post(f"{SUPABASE_URL}/rest/v1/visitas_seguimiento", json=data, headers=HEADERS)
        
        if response.status_code in [200, 201]:
            flash("Visita de seguimiento registrada correctamente", "success")
            # Si viene de una oportunidad, volver a la oportunidad
            if data["oportunidad_id"]:
                return redirect(url_for('ver_oportunidad', oportunidad_id=data["oportunidad_id"]))
            else:
                return redirect(f"/ver_lead/{cliente_id}")
        else:
            flash("Error al registrar visita", "error")
    
    response_cliente = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{cliente_id}", headers=HEADERS)
    response_oportunidades = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?cliente_id=eq.{cliente_id}&estado=eq.activa",
        headers=HEADERS
    )
    
    if response_cliente.status_code != 200 or not response_cliente.json():
        flash("Cliente no encontrado", "error")
        return redirect("/leads_dashboard")
    
    cliente = response_cliente.json()[0]
    oportunidades = response_oportunidades.json() if response_oportunidades.status_code == 200 else []
    
    fecha_hoy = date.today().strftime('%Y-%m-%d')
    
    return render_template("crear_visita_seguimiento.html",
        cliente=cliente,
        oportunidades=oportunidades,
        oportunidad_id_preseleccionada=oportunidad_id,
        fecha_hoy=fecha_hoy
    )

# DESCARGO COMERCIAL MENSUAL
@app.route("/reporte_mensual", methods=["GET", "POST"])
def reporte_mensual():
    if "usuario" not in session:
        return redirect("/")
    
    if request.method == "POST":
        mes = int(request.form.get("mes"))
        año = int(request.form.get("año"))
        
        ultimo_dia = calendar.monthrange(año, mes)[1]
        fecha_inicio = f"{año}-{mes:02d}-01"
        fecha_fin = f"{año}-{mes:02d}-{ultimo_dia}"
        
        query_clientes = f"fecha_visita=gte.{fecha_inicio}&fecha_visita=lte.{fecha_fin}"
        response_clientes = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?{query_clientes}&select=*", headers=HEADERS)
        
        response_seguimiento = requests.get(
            f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?fecha_visita=gte.{fecha_inicio}&fecha_visita=lte.{fecha_fin}&select=*,clientes(nombre_cliente,direccion,localidad)",
            headers=HEADERS
        )
        
        response_admin = requests.get(f"{SUPABASE_URL}/rest/v1/visitas_administradores?{query_clientes}&select=*", headers=HEADERS)

        # Obtener oportunidades activas (no cerradas: que no estén ganadas ni perdidas)
        response_oportunidades = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?estado=neq.ganada&estado=neq.perdida&select=*,clientes(direccion,localidad)",
            headers=HEADERS
        )

        if response_clientes.status_code != 200:
            return f"Error al obtener datos: {response_clientes.text}"

        clientes_mes = response_clientes.json()
        visitas_seguimiento_mes = response_seguimiento.json() if response_seguimiento.status_code == 200 else []
        visitas_admin_mes = response_admin.json() if response_admin.status_code == 200 else []
        oportunidades_activas = response_oportunidades.json() if response_oportunidades.status_code == 200 else []

        # Ordenar todas las visitas por fecha
        clientes_mes = sorted(clientes_mes, key=lambda x: x.get('fecha_visita', ''))
        visitas_seguimiento_mes = sorted(visitas_seguimiento_mes, key=lambda x: x.get('fecha_visita', ''))
        visitas_admin_mes = sorted(visitas_admin_mes, key=lambda x: x.get('fecha_visita', ''))

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        # Función helper para formatear fechas a dd/mm/aaaa
        def formatear_fecha(fecha_str):
            if not fecha_str:
                return ''
            try:
                # Convertir desde formato ISO (YYYY-MM-DD) a dd/mm/aaaa
                fecha = datetime.strptime(fecha_str[:10], '%Y-%m-%d')
                return fecha.strftime('%d/%m/%Y')
            except:
                return fecha_str

        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "VISITAS INSTALACIONES"
        
        headers = ['FECHA', 'COMUNIDAD/EMPRESA', 'DIRECCION', 'ZONA', 'OBSERVACIONES']
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2

        for cliente in clientes_mes:
            ws1.cell(row=row, column=1, value=formatear_fecha(cliente.get('fecha_visita', '')))
            ws1.cell(row=row, column=2, value=cliente.get('nombre_cliente', ''))
            ws1.cell(row=row, column=3, value=cliente.get('direccion', ''))
            ws1.cell(row=row, column=4, value=cliente.get('localidad', ''))
            ws1.cell(row=row, column=5, value=cliente.get('observaciones', ''))

            for col in range(1, 6):
                ws1.cell(row=row, column=col).border = thin_border

            row += 1

        for visita in visitas_seguimiento_mes:
            ws1.cell(row=row, column=1, value=formatear_fecha(visita.get('fecha_visita', '')))
            ws1.cell(row=row, column=2, value=visita.get('clientes', {}).get('nombre_cliente', ''))
            ws1.cell(row=row, column=3, value=visita.get('clientes', {}).get('direccion', ''))
            ws1.cell(row=row, column=4, value=visita.get('clientes', {}).get('localidad', ''))
            ws1.cell(row=row, column=5, value=visita.get('observaciones', ''))

            for col in range(1, 6):
                ws1.cell(row=row, column=col).border = thin_border

            row += 1
        
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 50
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 70
        
        ws2 = wb.create_sheet(title="VISITAS ADMINISTRADORES")
        
        headers_admin = ['FECHA', 'ADMINISTRADOR', 'PERSONA CONTACTO', 'OBSERVACIONES']
        
        for col, header in enumerate(headers_admin, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for visita in visitas_admin_mes:
            ws2.cell(row=row, column=1, value=formatear_fecha(visita.get('fecha_visita', '')))
            ws2.cell(row=row, column=2, value=visita.get('administrador_fincas', ''))
            ws2.cell(row=row, column=3, value=visita.get('persona_contacto', ''))
            ws2.cell(row=row, column=4, value=visita.get('observaciones', ''))

            for col in range(1, 5):
                ws2.cell(row=row, column=col).border = thin_border

            row += 1
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 70

        # Tercera pestaña: OPORTUNIDADES ACTIVAS
        ws3 = wb.create_sheet(title="OPORTUNIDADES ACTIVAS")

        headers_oportunidades = ['DIRECCION', 'LOCALIDAD', 'TIPO DE OPORTUNIDAD', 'ESTADO', 'DESCRIPCION']

        for col, header in enumerate(headers_oportunidades, 1):
            cell = ws3.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 2
        # Mapeo de estados a etiquetas legibles
        estados_map = {
            'nueva': '🆕 Nueva',
            'en_contacto': '📞 En contacto',
            'presupuesto_preparacion': '✍️ Presupuesto en preparación',
            'presupuesto_enviado': '📤 Presupuesto enviado',
            'ganada': '✅ Ganada',
            'perdida': '❌ Perdida',
            'activa': '⚡ Activa'
        }

        for oportunidad in oportunidades_activas:
            # Obtener dirección y localidad del cliente relacionado
            direccion = oportunidad.get('clientes', {}).get('direccion', '') if oportunidad.get('clientes') else ''
            localidad = oportunidad.get('clientes', {}).get('localidad', '') if oportunidad.get('clientes') else ''

            # Obtener etiqueta del estado
            estado = oportunidad.get('estado', '')
            estado_label = estados_map.get(estado, estado)

            ws3.cell(row=row, column=1, value=direccion)
            ws3.cell(row=row, column=2, value=localidad)
            ws3.cell(row=row, column=3, value=oportunidad.get('tipo', ''))
            ws3.cell(row=row, column=4, value=estado_label)
            ws3.cell(row=row, column=5, value=oportunidad.get('descripcion', ''))

            for col in range(1, 6):
                ws3.cell(row=row, column=col).border = thin_border

            row += 1

        ws3.column_dimensions['A'].width = 50
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 30
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 70

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        meses = ['', 'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 
                'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        filename = f"DESCARGO COMERCIAL GRAN CANARIA {meses[mes]} {año}.xlsx"
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )
    
    return render_template("reporte_mensual.html")

# DASHBOARD MEJORADO CON PAGINACIÓN Y BÚSQUEDA
@app.route("/leads_dashboard")
def leads_dashboard():
    if "usuario" not in session:
        return redirect("/")

    page = int(request.args.get("page", 1))
    per_page = 25
    offset = (page - 1) * per_page

    filtro_localidad = request.args.get("localidad", "")
    filtro_empresa = request.args.get("empresa", "")
    # Aceptar tanto 'search' (desde home) como 'buscar_direccion' (desde dashboard)
    buscar_direccion = request.args.get("search", "") or request.args.get("buscar_direccion", "")

    # Si hay búsqueda de texto, usar RPC para búsqueda sin acentos
    if buscar_direccion:
        # Usar función RPC para búsqueda sin acentos
        rpc_url = f"{SUPABASE_URL}/rest/v1/rpc/buscar_clientes_sin_acentos"

        rpc_params = {
            "termino_busqueda": buscar_direccion,
            "filtro_localidad": filtro_localidad,
            "filtro_empresa": filtro_empresa,
            "limite": per_page,
            "desplazamiento": offset
        }

        response = requests.post(rpc_url, json=rpc_params, headers=HEADERS)

        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al buscar leads</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

        leads_base = response.json()

        # Obtener total_count del primer resultado si existe
        total_registros = leads_base[0].get('total_count', 0) if leads_base else 0
        total_pages = max(1, (total_registros + per_page - 1) // per_page)

    else:
        # Búsqueda normal con filtros (sin texto de búsqueda)
        query_params = []

        if filtro_localidad:
            query_params.append(f"localidad=eq.{filtro_localidad}")

        if filtro_empresa:
            query_params.append(f"empresa_mantenedora=eq.{filtro_empresa}")

        query_string = "&".join(query_params) if query_params else ""

        # OPTIMIZACIÓN: Para count solo necesitamos id, no todos los campos
        count_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id"
        if query_string:
            count_url += f"&{query_string}"

        count_response = requests.get(count_url, headers={**HEADERS, "Prefer": "count=exact"})
        total_registros = int(count_response.headers.get("Content-Range", "0").split("/")[-1])
        total_pages = max(1, (total_registros + per_page - 1) // per_page)

        # OPTIMIZACIÓN: Usar join de Supabase + selección específica de campos
        data_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id,direccion,nombre_cliente,localidad,empresa_mantenedora,numero_ascensores,created_at&limit={per_page}&offset={offset}"
        if query_string:
            data_url += f"&{query_string}"

        response = requests.get(data_url, headers=HEADERS)

        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al obtener leads</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

        leads_base = response.json()

    # Ahora obtener equipos para cada cliente encontrado
    # Esto es necesario porque RPC no devuelve relaciones anidadas
    cliente_ids = [lead['id'] for lead in leads_base]

    equipos_por_cliente = {}
    if cliente_ids:
        # Obtener todos los equipos de estos clientes en una sola query
        equipos_url = f"{SUPABASE_URL}/rest/v1/equipos?select=cliente_id,ipo_proxima,fecha_vencimiento_contrato&cliente_id=in.({','.join(map(str, cliente_ids))})"
        equipos_response = requests.get(equipos_url, headers=HEADERS)

        if equipos_response.status_code == 200:
            equipos_data = equipos_response.json()
            # Agrupar equipos por cliente_id
            for equipo in equipos_data:
                cliente_id = equipo['cliente_id']
                if cliente_id not in equipos_por_cliente:
                    equipos_por_cliente[cliente_id] = []
                equipos_por_cliente[cliente_id].append(equipo)
    rows = []

    # OPTIMIZACIÓN: Usar caché de filtros (TTL: 30 minutos)
    # Reduce de 2 queries a 0 queries en cargas subsecuentes
    localidades_disponibles, empresas_disponibles = get_filtros_cached()

    # Procesar leads y sus equipos
    for lead in leads_base:
        lead_id = lead["id"]

        # Obtener equipos de este cliente
        equipos = equipos_por_cliente.get(lead_id, [])

        direccion = lead.get("direccion", "-")
        nombre_cliente = lead.get("nombre_cliente", "")
        localidad = lead.get("localidad", "-")
        empresa_mantenedora = lead.get("empresa_mantenedora", "-")
        total_equipos = len(equipos) if equipos else lead.get("numero_ascensores", 0)
        created_at = lead.get("created_at")

        ipo_proxima = None
        contrato_vence = None

        # Procesar equipos si existen
        if equipos:
            for equipo in equipos:
                ipo_equipo = equipo.get("ipo_proxima")
                if ipo_equipo:
                    try:
                        ipo_date = datetime.strptime(ipo_equipo, "%Y-%m-%d")
                        if ipo_proxima is None or ipo_date < ipo_proxima:
                            ipo_proxima = ipo_date
                    except:
                        pass

                contrato_equipo = equipo.get("fecha_vencimiento_contrato")
                if contrato_equipo:
                    try:
                        contrato_date = datetime.strptime(contrato_equipo, "%Y-%m-%d")
                        if contrato_vence is None or contrato_date < contrato_vence:
                            contrato_vence = contrato_date
                    except:
                        pass

        # Crear fila SIEMPRE, tenga o no equipos
        ipo_proxima_str = ipo_proxima.strftime("%d/%m/%Y") if ipo_proxima else "-"
        contrato_vence_str = contrato_vence.strftime("%d/%m/%Y") if contrato_vence else "-"

        color_ipo = calcular_color_ipo(ipo_proxima_str)
        color_contrato = calcular_color_contrato(contrato_vence_str)

        row = {
            "lead_id": lead_id,
            "direccion": direccion,
            "nombre_cliente": nombre_cliente,
            "localidad": localidad,
            "total_equipos": total_equipos,
            "empresa_mantenedora": empresa_mantenedora,
            "ipo_proxima": ipo_proxima_str,
            "ipo_fecha_original": ipo_proxima,
            "contrato_vence": contrato_vence_str,
            "contrato_fecha_original": contrato_vence,
            "color_ipo": color_ipo,
            "color_contrato": color_contrato,
            "created_at": created_at
        }

        rows.append(row)
    
    # Ordenar por fecha de creación: más reciente primero
    rows.sort(key=lambda x: x.get("created_at") or "", reverse=True)

    # Los filtros ya vienen ordenados del caché, no es necesario volver a ordenar
    # localidades_disponibles y empresas_disponibles ya están ordenados

    return render_template("dashboard.html",
        rows=rows,
        localidades=localidades_disponibles,
        empresas=empresas_disponibles,
        filtro_localidad=filtro_localidad,
        filtro_empresa=filtro_empresa,
        buscar_direccion=buscar_direccion,
        page=page,
        total_pages=total_pages,
        total_registros=total_registros,
        per_page=per_page
    )

# Exportar leads a Excel
@app.route("/exportar_leads")
def exportar_leads():
    if "usuario" not in session:
        return redirect("/")

    # Obtener los mismos filtros que el dashboard
    filtro_localidad = request.args.get("localidad", "")
    filtro_empresa = request.args.get("empresa", "")
    buscar_direccion = request.args.get("buscar_direccion", "")

    # Construir query (sin límite de paginación para exportar todo)
    if buscar_direccion:
        # Usar RPC para búsqueda sin acentos (sin límite)
        rpc_url = f"{SUPABASE_URL}/rest/v1/rpc/buscar_clientes_sin_acentos"
        rpc_params = {
            "termino_busqueda": buscar_direccion,
            "filtro_localidad": filtro_localidad,
            "filtro_empresa": filtro_empresa,
            "limite": 10000,  # Límite alto para exportación
            "desplazamiento": 0
        }
        response = requests.post(rpc_url, json=rpc_params, headers=HEADERS)
        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al buscar leads</h3>"
        leads_base = response.json()
    else:
        # Query normal con filtros
        query_params = []
        if filtro_localidad:
            query_params.append(f"localidad=eq.{filtro_localidad}")
        if filtro_empresa:
            query_params.append(f"empresa_mantenedora=eq.{filtro_empresa}")

        query_string = "&".join(query_params) if query_params else ""
        data_url = f"{SUPABASE_URL}/rest/v1/clientes?select=id,direccion,nombre_cliente,localidad,empresa_mantenedora,numero_ascensores,telefono,email,persona_contacto,administrador_fincas"
        if query_string:
            data_url += f"&{query_string}"

        response = requests.get(data_url, headers=HEADERS)
        if response.status_code != 200:
            return f"<h3 style='color:red;'>Error al obtener leads</h3>"
        leads_base = response.json()

    # Obtener equipos para calcular IPOs
    cliente_ids = [lead['id'] for lead in leads_base]
    equipos_por_cliente = {}
    if cliente_ids:
        equipos_url = f"{SUPABASE_URL}/rest/v1/equipos?select=cliente_id,ipo_proxima&cliente_id=in.({','.join(map(str, cliente_ids))})"
        equipos_response = requests.get(equipos_url, headers=HEADERS)
        if equipos_response.status_code == 200:
            equipos_data = equipos_response.json()
            for equipo in equipos_data:
                cliente_id = equipo['cliente_id']
                if cliente_id not in equipos_por_cliente:
                    equipos_por_cliente[cliente_id] = []
                equipos_por_cliente[cliente_id].append(equipo)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Instalaciones"

    # Estilo de cabecera
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras
    headers = ["Dirección", "Nombre", "Población", "Teléfono", "Email", "Contacto",
               "Empresa Mantenedora", "Administrador", "Nº Ascensores", "Próxima IPO"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Datos
    row_num = 2
    for lead in leads_base:
        equipos = equipos_por_cliente.get(lead['id'], [])

        # Calcular próxima IPO
        ipo_proxima = None
        if equipos:
            for equipo in equipos:
                ipo_equipo = equipo.get("ipo_proxima")
                if ipo_equipo:
                    try:
                        ipo_date = datetime.strptime(ipo_equipo, "%Y-%m-%d")
                        if ipo_proxima is None or ipo_date < ipo_proxima:
                            ipo_proxima = ipo_date
                    except:
                        pass

        ipo_str = ipo_proxima.strftime("%d/%m/%Y") if ipo_proxima else "-"

        ws.cell(row=row_num, column=1, value=lead.get('direccion', ''))
        ws.cell(row=row_num, column=2, value=lead.get('nombre_cliente', ''))
        ws.cell(row=row_num, column=3, value=lead.get('localidad', ''))
        ws.cell(row=row_num, column=4, value=lead.get('telefono', ''))
        ws.cell(row=row_num, column=5, value=lead.get('email', ''))
        ws.cell(row=row_num, column=6, value=lead.get('persona_contacto', ''))
        ws.cell(row=row_num, column=7, value=lead.get('empresa_mantenedora', ''))
        ws.cell(row=row_num, column=8, value=lead.get('administrador_fincas', ''))
        ws.cell(row=row_num, column=9, value=lead.get('numero_ascensores', 0))
        ws.cell(row=row_num, column=10, value=ipo_str)
        row_num += 1

    # Ajustar ancho de columnas
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    # Guardar en memoria
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Generar nombre de archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"instalaciones_{timestamp}.xlsx"

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Ver detalle completo del Lead
@app.route("/ver_lead/<int:lead_id>")
def ver_lead(lead_id):
    if "usuario" not in session:
        return redirect("/")
    
    response = requests.get(f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}", headers=HEADERS)
    if response.status_code != 200 or not response.json():
        return f"<h3 style='color:red;'>Error al obtener Lead</h3><pre>{response.text}</pre><a href='/leads_dashboard'>Volver</a>"
    
    lead = response.json()[0]
    
    equipos_response = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?cliente_id=eq.{lead_id}", headers=HEADERS)
    equipos = []
    if equipos_response.status_code == 200:
        equipos_raw = equipos_response.json()
        for equipo in equipos_raw:
            fecha_venc = equipo.get("fecha_vencimiento_contrato", "-")
            if fecha_venc and fecha_venc != "-":
                try:
                    partes = fecha_venc.split("-")
                    if len(partes) == 3:
                        fecha_venc = f"{partes[2]}/{partes[1]}/{partes[0]}"
                except:
                    pass
            
            ipo = equipo.get("ipo_proxima", "-")
            if ipo and ipo != "-":
                try:
                    partes = ipo.split("-")
                    if len(partes) == 3:
                        ipo = f"{partes[2]}/{partes[1]}/{partes[0]}"
                except:
                    pass
            
            equipos.append({
                "id": equipo.get("id"),
                "tipo_equipo": equipo.get("tipo_equipo", "-"),
                "identificacion": equipo.get("identificacion", "-"),
                "rae": equipo.get("rae", "-"),
                "ipo_proxima": ipo,
                "fecha_vencimiento_contrato": fecha_venc,
                "descripcion": equipo.get("descripcion", "-")
            })
    
    oportunidades_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?cliente_id=eq.{lead_id}&order=fecha_creacion.desc",
        headers=HEADERS
    )
    oportunidades = []
    if oportunidades_response.status_code == 200:
        oportunidades = oportunidades_response.json()
    
    visitas_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?cliente_id=eq.{lead_id}&select=*,oportunidades(tipo)",
        headers=HEADERS
    )
    visitas_seguimiento = []
    if visitas_response.status_code == 200:
        visitas_seguimiento = visitas_response.json()

    # Combinar todas las visitas (inicial + seguimiento) y ordenar por fecha descendente
    todas_visitas = []

    # Agregar la visita inicial del lead
    if lead.get('fecha_visita'):
        todas_visitas.append({
            'fecha_visita': lead.get('fecha_visita'),
            'observaciones': lead.get('observaciones'),
            'tipo': 'inicial',
            'oportunidades': None
        })

    # Agregar las visitas de seguimiento
    for visita in visitas_seguimiento:
        todas_visitas.append({
            'fecha_visita': visita.get('fecha_visita'),
            'observaciones': visita.get('observaciones'),
            'tipo': 'seguimiento',
            'oportunidades': visita.get('oportunidades')
        })

    # Ordenar todas las visitas por fecha descendente (más reciente primero)
    todas_visitas.sort(key=lambda x: x.get('fecha_visita', ''), reverse=True)

    # Obtener datos del administrador si existe relación
    administrador = None
    if lead.get('administrador_id'):
        admin_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{lead['administrador_id']}",
            headers=HEADERS
        )
        if admin_response.status_code == 200 and admin_response.json():
            administrador = admin_response.json()[0]

    # Obtener tareas comerciales del cliente (abiertas y cerradas)
    tareas_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{lead_id}&order=fecha_creacion.desc",
        headers=HEADERS
    )
    tareas_comerciales = []
    if tareas_response.status_code == 200:
        tareas_comerciales = tareas_response.json()

    return render_template("ver_lead.html",
        lead=lead,
        equipos=equipos,
        oportunidades=oportunidades,
        todas_visitas=todas_visitas,
        administrador=administrador,
        tareas_comerciales=tareas_comerciales
    )

# Eliminar Lead
@app.route("/eliminar_lead/<int:lead_id>")
@helpers.login_required
@helpers.requiere_permiso('clientes', 'delete')
def eliminar_lead(lead_id):
    
    requests.delete(f"{SUPABASE_URL}/rest/v1/equipos?cliente_id=eq.{lead_id}", headers=HEADERS)
    requests.delete(f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?cliente_id=eq.{lead_id}", headers=HEADERS)
    
    response = requests.delete(f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}", headers=HEADERS)
    
    if response.status_code in [200, 204]:
        return redirect("/leads_dashboard")
    else:
        return f"<h3 style='color:red;'>Error al eliminar Lead</h3><pre>{response.text}</pre><a href='/leads_dashboard'>Volver</a>"

# Eliminar Equipo
@app.route("/eliminar_equipo/<int:equipo_id>")
@helpers.login_required
@helpers.requiere_permiso('equipos', 'delete')
def eliminar_equipo(equipo_id):
    
    equipo_response = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}", headers=HEADERS)
    if equipo_response.status_code == 200 and equipo_response.json():
        cliente_id = equipo_response.json()[0].get("cliente_id")
        
        response = requests.delete(f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}", headers=HEADERS)
        
        if response.status_code in [200, 204]:
            return redirect(f"/ver_lead/{cliente_id}")
        else:
            return f"<h3 style='color:red;'>Error al eliminar Equipo</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"
    else:
        return f"<h3 style='color:red;'>Error al obtener Equipo</h3><a href='/home'>Volver</a>"

# Ver detalle de equipo (accesible solo desde ver_lead)
@app.route("/ver_equipo/<int:equipo_id>")
def ver_equipo(equipo_id):
    if "usuario" not in session:
        return redirect("/")

    # Obtener equipo con JOIN a cliente
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}&select=*,cliente:clientes(id,direccion,localidad,nombre_cliente)",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Equipo no encontrado", "error")
        return redirect("/leads_dashboard")

    equipo = limpiar_none(response.json()[0])

    # El cliente viene como un objeto dentro de equipo
    cliente = equipo.pop('cliente', None) if 'cliente' in equipo else None

    return render_template("ver_equipo.html", equipo=equipo, cliente=cliente)

# Dashboard de Oportunidades Post-IPO
@app.route("/oportunidades_post_ipo", methods=["GET"])
def oportunidades_post_ipo():
    """Seguimiento Comercial - Sistema de tareas automáticas"""
    if "usuario" not in session:
        return redirect("/")

    # Determinar pestaña activa (abiertas o futuras)
    tab = request.args.get("tab", "abiertas")
    hoy = datetime.now().date()

    try:
        # === 1. OBTENER EQUIPOS CON IPO ===
        equipos_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/equipos?select=id,ipo_proxima,rae,cliente_id,clientes(direccion,localidad,telefono,persona_contacto,empresa_mantenedora)&ipo_proxima=not.is.null",
            headers=HEADERS,
            timeout=10
        )
        equipos_data = equipos_response.json() if equipos_response.status_code == 200 else []

        # === 2. PROCESAR IPOs Y CREAR TAREAS AUTOMÁTICAS ===
        # Agrupar equipos por cliente (solo el más próximo)
        clientes_con_ipo = {}
        for equipo in equipos_data:
            if not equipo.get('ipo_proxima'):
                continue

            try:
                ipo_date = datetime.strptime(equipo['ipo_proxima'], '%Y-%m-%d').date()
            except:
                continue

            dias_desde_ipo = (hoy - ipo_date).days
            cliente_id = equipo['cliente_id']

            # Solo guardar el equipo con IPO más reciente por cliente
            if cliente_id not in clientes_con_ipo or dias_desde_ipo < clientes_con_ipo[cliente_id]['dias_desde_ipo']:
                clientes_con_ipo[cliente_id] = {
                    'cliente_id': cliente_id,
                    'equipo_id': equipo['id'],
                    'ipo_date': ipo_date,
                    'dias_desde_ipo': dias_desde_ipo,
                    'dias_hasta_ipo': abs(dias_desde_ipo) if dias_desde_ipo < 0 else 0,
                    'rae': equipo.get('rae'),
                    'cliente': equipo.get('clientes', {})
                }

        # === 3. CREAR TAREAS AUTOMÁTICAS (IPO >= 15 días) ===
        for cliente_id, data in clientes_con_ipo.items():
            if data['dias_desde_ipo'] >= 15:
                # Verificar si ya existe tarea abierta para este cliente
                tarea_existe = requests.get(
                    f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{cliente_id}&estado=eq.abierta",
                    headers=HEADERS
                ).json()

                # Verificar si el cliente tiene tareas descartadas (para no crear nuevas)
                tarea_descartada = requests.get(
                    f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{cliente_id}&estado=eq.cerrada&tipo_cierre=not.is.null",
                    headers=HEADERS
                ).json()

                if not tarea_existe and not tarea_descartada:
                    # Crear tarea automáticamente solo si no hay tareas abiertas ni descartadas
                    nueva_tarea = {
                        'cliente_id': cliente_id,
                        'equipo_id': data['equipo_id'],
                        'estado': 'abierta',
                        'motivo_creacion': 'ipo_15_dias',
                        'dias_desde_ipo': data['dias_desde_ipo'],
                        'creado_por': 'sistema'
                    }
                    requests.post(
                        f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas",
                        headers=HEADERS,
                        json=nueva_tarea
                    )

        # === 3B. OBTENER CLIENTES CON FECHA_FIN_CONTRATO Y CREAR TAREAS AUTOMÁTICAS ===
        clientes_fin_contrato_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/clientes?select=id,fecha_fin_contrato,direccion,localidad,telefono,persona_contacto,empresa_mantenedora&fecha_fin_contrato=not.is.null",
            headers=HEADERS,
            timeout=10
        )
        clientes_fin_contrato_data = clientes_fin_contrato_response.json() if clientes_fin_contrato_response.status_code == 200 else []

        clientes_con_fin_contrato = {}
        for cliente in clientes_fin_contrato_data:
            if not cliente.get('fecha_fin_contrato'):
                continue

            try:
                fecha_fin = datetime.strptime(cliente['fecha_fin_contrato'], '%Y-%m-%d').date()
            except:
                continue

            dias_hasta_fin = (fecha_fin - hoy).days
            cliente_id = cliente['id']

            clientes_con_fin_contrato[cliente_id] = {
                'cliente_id': cliente_id,
                'fecha_fin_contrato': fecha_fin,
                'dias_hasta_fin': dias_hasta_fin,
                'cliente': cliente
            }

            # Crear tarea automática si faltan 120 días o menos
            if dias_hasta_fin <= 120 and dias_hasta_fin >= 0:
                # Verificar si ya existe tarea abierta para este cliente
                tarea_existe = requests.get(
                    f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{cliente_id}&estado=eq.abierta",
                    headers=HEADERS
                ).json()

                # Verificar si el cliente tiene tareas descartadas (para no crear nuevas)
                tarea_descartada = requests.get(
                    f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?cliente_id=eq.{cliente_id}&estado=eq.cerrada&tipo_cierre=not.is.null",
                    headers=HEADERS
                ).json()

                if not tarea_existe and not tarea_descartada:
                    # Crear tarea automáticamente solo si no hay tareas abiertas ni descartadas
                    nueva_tarea = {
                        'cliente_id': cliente_id,
                        'equipo_id': None,
                        'estado': 'abierta',
                        'motivo_creacion': 'fin_contrato_120_dias',
                        'dias_desde_ipo': None,
                        'creado_por': 'sistema'
                    }
                    requests.post(
                        f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas",
                        headers=HEADERS,
                        json=nueva_tarea
                    )

        # === 4. OBTENER TAREAS EXISTENTES CON DATOS DEL CLIENTE ===
        tareas_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?select=*,clientes(direccion,localidad,telefono,persona_contacto,empresa_mantenedora)&estado=eq.abierta&order=fecha_creacion.asc",
            headers=HEADERS
        )
        tareas_data = tareas_response.json() if tareas_response.status_code == 200 else []

        # === 5. CLASIFICAR TAREAS ===
        tareas_abiertas = []
        tareas_aplazadas = []

        for tarea in tareas_data:
            # Obtener datos del cliente
            cliente_info = tarea.get('clientes', {})
            cliente_id = tarea['cliente_id']

            # Buscar días desde IPO actualizado
            dias_desde_ipo = clientes_con_ipo.get(cliente_id, {}).get('dias_desde_ipo', tarea.get('dias_desde_ipo', 0))

            tarea_enriched = {
                'id': tarea['id'],
                'cliente_id': cliente_id,
                'direccion': cliente_info.get('direccion', 'Sin dirección'),
                'localidad': cliente_info.get('localidad', ''),
                'telefono': cliente_info.get('telefono'),
                'persona_contacto': cliente_info.get('persona_contacto'),
                'empresa_mantenedora': cliente_info.get('empresa_mantenedora'),
                'dias_desde_ipo': dias_desde_ipo,
                'fecha_creacion': tarea['fecha_creacion'],
                'aplazada_hasta': tarea.get('aplazada_hasta'),
                'motivo_aplazamiento': tarea.get('motivo_aplazamiento'),
                'notas': tarea.get('notas', [])
            }

            # Clasificar: ABIERTAS vs FUTURAS (aplazadas)
            if tarea.get('aplazada_hasta'):
                try:
                    fecha_aplazada = datetime.strptime(tarea['aplazada_hasta'], '%Y-%m-%d').date()
                    if fecha_aplazada > hoy:
                        # Aún aplazada
                        tarea_enriched['dias_para_reabrir'] = (fecha_aplazada - hoy).days
                        tareas_aplazadas.append(tarea_enriched)
                    else:
                        # Aplazamiento venció, pasa a abiertas
                        tareas_abiertas.append(tarea_enriched)
                except:
                    tareas_abiertas.append(tarea_enriched)
            else:
                tareas_abiertas.append(tarea_enriched)

        # === 6. FUTURAS - PRÓXIMAS AUTOMÁTICAS (IPO próximos 30 días + hace 0-14 días) ===
        proximas_automaticas = []
        for cliente_id, data in clientes_con_ipo.items():
            # IPO en próximos 30 días O ya ocurrió hace 0-14 días (antes de crear tarea)
            if -30 <= data['dias_desde_ipo'] < 15:
                # Verificar que no tenga tarea
                tiene_tarea = any(t['cliente_id'] == cliente_id for t in tareas_abiertas + tareas_aplazadas)
                if not tiene_tarea:
                    # Determinar si es futura o pasada
                    es_futura = data['dias_desde_ipo'] < 0

                    if es_futura:
                        # IPO aún no ocurrió
                        dias_hasta_ipo = abs(data['dias_desde_ipo'])
                        dias_para_activar = dias_hasta_ipo + 15
                    else:
                        # IPO ya ocurrió, esperando llegar a 15 días
                        dias_desde_ipo = data['dias_desde_ipo']
                        dias_hasta_ipo = None  # No aplica
                        dias_para_activar = 15 - dias_desde_ipo

                    proximas_automaticas.append({
                        'cliente_id': cliente_id,
                        'direccion': data['cliente'].get('direccion', 'Sin dirección'),
                        'localidad': data['cliente'].get('localidad', ''),
                        'telefono': data['cliente'].get('telefono'),
                        'es_futura': es_futura,
                        'dias_hasta_ipo': dias_hasta_ipo,  # Solo para futuras
                        'dias_desde_ipo': dias_desde_ipo if not es_futura else None,  # Solo para pasadas
                        'dias_para_activar': dias_para_activar,
                        'rae': data['rae'],
                        'motivo': 'IPO'
                    })

        # === 6B. FUTURAS - PRÓXIMAS POR FIN DE CONTRATO (121-150 días) ===
        for cliente_id, data in clientes_con_fin_contrato.items():
            # Fin de contrato entre 121 y 150 días
            if 121 <= data['dias_hasta_fin'] <= 150:
                # Verificar que no tenga tarea
                tiene_tarea = any(t['cliente_id'] == cliente_id for t in tareas_abiertas + tareas_aplazadas)
                if not tiene_tarea:
                    proximas_automaticas.append({
                        'cliente_id': cliente_id,
                        'direccion': data['cliente'].get('direccion', 'Sin dirección'),
                        'localidad': data['cliente'].get('localidad', ''),
                        'telefono': data['cliente'].get('telefono'),
                        'es_futura': True,
                        'dias_hasta_fin_contrato': data['dias_hasta_fin'],
                        'dias_para_activar': data['dias_hasta_fin'] - 120,  # Días hasta que se active (120 días antes)
                        'fecha_fin_contrato': data['fecha_fin_contrato'].strftime('%d/%m/%Y'),
                        'motivo': 'Fin de Contrato'
                    })

        # Ordenar
        tareas_abiertas.sort(key=lambda x: x['dias_desde_ipo'], reverse=True)  # Más urgente primero
        tareas_aplazadas.sort(key=lambda x: x['aplazada_hasta'])
        # Ordenar próximas: primero las que ya pasaron (más urgentes), luego las futuras
        proximas_automaticas.sort(key=lambda x: x['dias_para_activar'])

        return render_template(
            "oportunidades_post_ipo.html",
            tab=tab,
            tareas_abiertas=tareas_abiertas,
            tareas_aplazadas=tareas_aplazadas,
            proximas_automaticas=proximas_automaticas,
            abiertas_count=len(tareas_abiertas),
            aplazadas_count=len(tareas_aplazadas),
            proximas_count=len(proximas_automaticas)
        )

    except Exception as e:
        print(f"Error en seguimiento comercial: {str(e)}")
        flash(f"Error al cargar seguimiento comercial: {str(e)}", "error")
        return redirect("/home")


# === RUTAS DE GESTIÓN DE TAREAS COMERCIALES ===

@app.route("/tarea_comercial_aplazar/<int:tarea_id>", methods=["POST"])
def tarea_comercial_aplazar(tarea_id):
    """Aplazar una tarea comercial"""
    if "usuario" not in session:
        return {"error": "No autorizado"}, 401

    try:
        dias_aplazar = int(request.json.get("dias", 7))
        motivo = request.json.get("motivo", "")

        fecha_aplazada = (datetime.now().date() + timedelta(days=dias_aplazar)).isoformat()

        data = {
            "aplazada_hasta": fecha_aplazada,
            "motivo_aplazamiento": motivo,
            "motivo_creacion": "aplazada_vuelve"
        }

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?id=eq.{tarea_id}",
            headers=HEADERS,
            json=data
        )

        if response.status_code in [200, 204]:
            return {"success": True, "fecha_aplazada": fecha_aplazada}, 200
        else:
            return {"error": "Error al aplazar tarea"}, 500

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/tarea_comercial_descartar/<int:tarea_id>", methods=["POST"])
def tarea_comercial_descartar(tarea_id):
    """Descartar una tarea comercial (cerrarla sin crear oportunidad)"""
    if "usuario" not in session:
        return {"error": "No autorizado"}, 401

    try:
        tipo_descarte = request.json.get("tipo", "descartada_sin_interes")
        motivo = request.json.get("motivo", "")

        data = {
            "estado": "cerrada",
            "tipo_cierre": tipo_descarte,
            "motivo_cierre": motivo,
            "fecha_cierre": datetime.now().isoformat()
        }

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?id=eq.{tarea_id}",
            headers=HEADERS,
            json=data
        )

        if response.status_code in [200, 204]:
            return {"success": True}, 200
        else:
            return {"error": "Error al descartar tarea"}, 500

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/tarea_comercial_convertir/<int:tarea_id>", methods=["POST"])
def tarea_comercial_convertir(tarea_id):
    """Marcar tarea como convertida (redirecciona a crear oportunidad)"""
    if "usuario" not in session:
        return redirect("/")

    try:
        # Obtener datos de la tarea
        tarea_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?id=eq.{tarea_id}",
            headers=HEADERS
        )

        if tarea_response.status_code == 200:
            tarea = tarea_response.json()[0]
            cliente_id = tarea['cliente_id']

            # Redirigir a crear oportunidad
            # La tarea se cerrará cuando se cree la oportunidad exitosamente
            return redirect(url_for("crear_oportunidad", cliente_id=cliente_id, tarea_id=tarea_id))
        else:
            flash("Error al obtener datos de la tarea", "error")
            return redirect("/oportunidades_post_ipo")

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect("/oportunidades_post_ipo")


@app.route("/tarea_comercial_agregar_nota/<int:tarea_id>", methods=["POST"])
def tarea_comercial_agregar_nota(tarea_id):
    """Agregar una nota a la tarea"""
    if "usuario" not in session:
        return {"error": "No autorizado"}, 401

    try:
        texto_nota = request.json.get("nota", "")
        if not texto_nota:
            return {"error": "Nota vacía"}, 400

        # Obtener notas actuales
        tarea_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?id=eq.{tarea_id}",
            headers=HEADERS
        )

        if tarea_response.status_code == 200:
            tarea = tarea_response.json()[0]
            notas = tarea.get('notas', [])

            # Agregar nueva nota
            nueva_nota = {
                "fecha": datetime.now().isoformat(),
                "usuario": session.get("usuario", "Usuario"),
                "texto": texto_nota
            }
            notas.append(nueva_nota)

            # Actualizar
            response = requests.patch(
                f"{SUPABASE_URL}/rest/v1/seguimiento_comercial_tareas?id=eq.{tarea_id}",
                headers=HEADERS,
                json={"notas": notas}
            )

            if response.status_code in [200, 204]:
                return {"success": True, "nota": nueva_nota}, 200
            else:
                return {"error": "Error al guardar nota"}, 500
        else:
            return {"error": "Tarea no encontrada"}, 404

    except Exception as e:
        return {"error": str(e)}, 500


# Editar Lead - CON LIMPIEZA DE NONE
@app.route("/editar_lead/<int:lead_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('clientes', 'write')
def editar_lead(lead_id):

    if request.method == "POST":
        # Obtener administrador_id y convertir a int o None
        administrador_id = request.form.get("administrador_id")
        if administrador_id and administrador_id.strip():
            try:
                administrador_id = int(administrador_id)
            except ValueError:
                administrador_id = None
        else:
            administrador_id = None

        data = {
            "fecha_visita": request.form.get("fecha_visita") or None,
            "tipo_cliente": request.form.get("tipo_lead") or None,
            "direccion": request.form.get("direccion"),
            "nombre_cliente": request.form.get("nombre_lead") or None,
            "codigo_postal": request.form.get("codigo_postal") or None,
            "localidad": request.form.get("localidad"),
            "zona": request.form.get("zona") or None,
            "persona_contacto": request.form.get("persona_contacto") or None,
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "administrador_id": administrador_id,
            "empresa_mantenedora": request.form.get("empresa_mantenedora") or None,
            "numero_ascensores": request.form.get("numero_ascensores") or None,
            "fecha_fin_contrato": request.form.get("fecha_fin_contrato") or None,
            "paradas": request.form.get("paradas") or None,
            "viviendas_por_planta": request.form.get("viviendas_por_planta") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Solo dirección y localidad son obligatorios
        if not data["direccion"] or not data["localidad"]:
            flash("Dirección y Localidad son obligatorios", "error")
            return redirect(request.referrer)

        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}",
            json=data,
            headers=HEADERS
        )
        if res.status_code in [200, 204]:
            return redirect(f"/ver_lead/{lead_id}")
        else:
            return f"<h3 style='color:red;'>Error al actualizar Lead</h3><pre>{res.text}</pre><a href='/leads_dashboard'>Volver</a>"

    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{lead_id}",
        headers=HEADERS
    )
    if response.status_code == 200 and response.json():
        lead = response.json()[0]
        # LIMPIAR VALORES NONE PARA NO MOSTRAR "none" EN EL FORMULARIO
        lead = limpiar_none(lead)
    else:
        return f"<h3 style='color:red;'>Error al obtener Lead</h3><pre>{response.text}</pre><a href='/leads_dashboard'>Volver</a>"

    # Obtener lista de administradores (usando caché)
    administradores = get_administradores_cached()

    return render_template("editar_lead.html", lead=lead, administradores=administradores)

# Editar Equipo - CON LIMPIEZA DE NONE
@app.route("/editar_equipo/<int:equipo_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('equipos', 'write')
def editar_equipo(equipo_id):

    response = requests.get(f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}", headers=HEADERS)
    if response.status_code != 200 or not response.json():
        return f"<h3 style='color:red;'>Error al obtener equipo</h3><pre>{response.text}</pre><a href='/home'>Volver</a>"

    equipo = response.json()[0]

    if request.method == "POST":
        data = {
            "tipo_equipo": request.form.get("tipo_equipo"),
            "identificacion": request.form.get("identificacion") or None,
            "descripcion": request.form.get("observaciones") or None,
            "fecha_vencimiento_contrato": request.form.get("fecha_vencimiento_contrato") or None,
            "rae": request.form.get("rae") or None,
            "ipo_proxima": request.form.get("ipo_proxima") or None
        }

        # Solo tipo_equipo es obligatorio
        if not data["tipo_equipo"]:
            flash("Tipo de equipo es obligatorio", "error")
            return redirect(request.referrer)

        update_url = f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}"
        res = requests.patch(update_url, json=data, headers=HEADERS)
        if res.status_code in [200, 204]:
            # Obtener el cliente_id del equipo para volver a su vista
            cliente_id = equipo.get("cliente_id")
            return redirect(f"/ver_lead/{cliente_id}")
        else:
            return f"<h3 style='color:red;'>Error al actualizar equipo</h3><pre>{res.text}</pre><a href='/home'>Volver</a>"

    # LIMPIAR VALORES NONE PARA NO MOSTRAR "none" EN EL FORMULARIO
    equipo = limpiar_none(equipo)
    return render_template("editar_equipo.html", equipo=equipo)

# ============================================
# MÓDULO DE OPORTUNIDADES - ACTUALIZADO
# ============================================

@app.route("/oportunidades")
def oportunidades():
    if "usuario" not in session:
        return redirect("/")
    
    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?"
            f"select=*,clientes(nombre_cliente,direccion,localidad)"
            f"&order=fecha_creacion.desc",
            headers=HEADERS
        )
        
        if response.status_code == 200:
            oportunidades_list = response.json()

            # Contar por estado
            nuevas = sum(1 for o in oportunidades_list if o['estado'] == 'nueva')
            presupuesto_preparacion = sum(1 for o in oportunidades_list if o['estado'] == 'presupuesto_preparacion')
            presupuesto_enviado = sum(1 for o in oportunidades_list if o['estado'] == 'presupuesto_enviado')
            ganadas = sum(1 for o in oportunidades_list if o['estado'] == 'ganada')
            perdidas = sum(1 for o in oportunidades_list if o['estado'] == 'perdida')

            return render_template("oportunidades.html",
                                        oportunidades=oportunidades_list,
                                        nuevas=nuevas,
                                        presupuesto_preparacion=presupuesto_preparacion,
                                        presupuesto_enviado=presupuesto_enviado,
                                        ganadas=ganadas,
                                        perdidas=perdidas)
        else:
            flash("Error al cargar oportunidades", "error")
            return redirect(url_for("home"))
            
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("home"))


@app.route("/mi_agenda")
def mi_agenda():
    """Dashboard personal - Mi Agenda Comercial con pipeline de oportunidades"""
    if "usuario" not in session:
        return redirect("/")

    try:
        # Obtener TODAS las oportunidades que NO estén ganadas ni perdidas
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?"
            f"select=*,clientes(nombre_cliente,direccion,localidad,telefono,email,persona_contacto)"
            f"&estado=not.in.(ganada,perdida)"
            f"&order=fecha_creacion.desc",
            headers=HEADERS
        )

        if response.status_code == 200:
            oportunidades_list = response.json()

            # Agrupar por estado
            por_estado = {
                'nueva': [],
                'en_contacto': [],
                'presupuesto_preparacion': [],
                'presupuesto_enviado': [],
                'activa': []  # Para compatibilidad con oportunidades antiguas
            }

            for opp in oportunidades_list:
                estado = opp.get('estado', 'activa')
                if estado in por_estado:
                    por_estado[estado].append(opp)
                else:
                    # Si tiene un estado que no conocemos, lo ponemos en activa
                    por_estado['activa'].append(opp)

            # Contadores
            total_activas = len(oportunidades_list)

            return render_template("mi_agenda.html",
                                 por_estado=por_estado,
                                 total_activas=total_activas)
        else:
            flash(f"Error al cargar oportunidades: {response.status_code} - {response.text}", "error")
            return redirect(url_for("home"))

    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("home"))


@app.route("/cambiar_estado_oportunidad/<int:oportunidad_id>", methods=["POST"])
def cambiar_estado_oportunidad(oportunidad_id):
    """Endpoint para cambio rápido de estado desde Mi Agenda"""
    if "usuario" not in session:
        return {"error": "No autorizado"}, 401

    try:
        nuevo_estado = request.json.get("estado")
        if not nuevo_estado:
            return {"error": "Estado requerido"}, 400

        # Validar estados permitidos
        estados_validos = ["nueva", "en_contacto", "presupuesto_preparacion",
                          "presupuesto_enviado", "ganada", "perdida"]
        if nuevo_estado not in estados_validos:
            return {"error": "Estado no válido"}, 400

        data = {"estado": nuevo_estado}

        # Si se marca como ganada o perdida, agregar fecha de cierre
        if nuevo_estado in ["ganada", "perdida"]:
            data["fecha_cierre"] = datetime.now().isoformat()

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
            headers=HEADERS,
            json=data
        )

        if response.status_code in [200, 204]:
            return {"success": True, "estado": nuevo_estado}, 200
        else:
            return {"error": "Error al actualizar"}, 500

    except Exception as e:
        return {"error": str(e)}, 500


@app.route("/crear_oportunidad/<int:cliente_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('oportunidades', 'write')
def crear_oportunidad(cliente_id):
    
    if request.method == "POST":
        try:
            data = {
                "cliente_id": cliente_id,
                "tipo": request.form["tipo"],
                "descripcion": request.form.get("descripcion", ""),
                "valor_estimado": request.form.get("valor_estimado") or None,
                "observaciones": request.form.get("observaciones", ""),
                "estado": "nueva"
            }
            
            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/oportunidades",
                headers=HEADERS,
                json=data
            )
            
            if response.status_code == 201:
                flash("Oportunidad creada exitosamente!", "success")
                return redirect(url_for("ver_lead", lead_id=cliente_id))
            else:
                flash("Error al crear oportunidad", "error")
                
        except Exception as e:
            flash(f"Error: {str(e)}", "error")
    
    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/clientes?id=eq.{cliente_id}",
            headers=HEADERS
        )
        if response.status_code == 200:
            cliente = response.json()[0]
            return render_template("crear_oportunidad.html", cliente=cliente)
    except:
        flash("Error al cargar datos del cliente", "error")
        return redirect(url_for("leads_dashboard"))


# ACTUALIZADO: editar_oportunidad con nuevos campos
@app.route("/editar_oportunidad/<int:oportunidad_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('oportunidades', 'write')
def editar_oportunidad(oportunidad_id):
    
    if request.method == "POST":
        try:
            data = {
                "tipo": request.form["tipo"],
                "descripcion": request.form.get("descripcion", ""),
                "estado": request.form["estado"],
                "valor_estimado": request.form.get("valor_estimado") or None,
                "notas": request.form.get("notas", ""),
                "estado_presupuesto": request.form.get("estado_presupuesto", "No")
            }
            
            if data["estado"] in ["ganada", "perdida"]:
                data["fecha_cierre"] = datetime.now().isoformat()
            
            response = requests.patch(
                f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
                headers=HEADERS,
                json=data
            )
            
            if response.status_code in [200, 204]:
                flash("Oportunidad actualizada correctamente", "success")
                return redirect(url_for("ver_oportunidad", oportunidad_id=oportunidad_id))
            else:
                flash("Error al actualizar", "error")
                
        except Exception as e:
            flash(f"Error: {str(e)}", "error")
    
    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=*,clientes(nombre_cliente,direccion)",
            headers=HEADERS
        )
        if response.status_code == 200:
            oportunidad = response.json()[0]
            return render_template("editar_oportunidad.html", oportunidad=oportunidad)
    except:
        flash("Error al cargar oportunidad", "error")
        return redirect(url_for("oportunidades"))


# ACTUALIZADO: ver_oportunidad con visitas y acciones
@app.route("/ver_oportunidad/<int:oportunidad_id>")
def ver_oportunidad(oportunidad_id):
    if "usuario" not in session:
        return redirect("/")
    
    try:
        # Obtener oportunidad con datos del cliente
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=*,clientes(nombre_cliente,direccion,localidad)",
            headers=HEADERS
        )
        
        if response.status_code == 200 and response.json():
            oportunidad = response.json()[0]
            
            # Asegurar que acciones sea una lista
            if not oportunidad.get('acciones') or not isinstance(oportunidad.get('acciones'), list):
                oportunidad['acciones'] = []
            
            # Obtener visitas de seguimiento asociadas a esta oportunidad
            visitas_seguimiento_response = requests.get(
                f"{SUPABASE_URL}/rest/v1/visitas_seguimiento?oportunidad_id=eq.{oportunidad_id}&select=*,clientes(nombre_cliente,direccion)&order=fecha_visita.desc",
                headers=HEADERS
            )
            
            visitas_seguimiento = []
            if visitas_seguimiento_response.status_code == 200:
                for v in visitas_seguimiento_response.json():
                    visitas_seguimiento.append({
                        'id': v.get('id'),
                        'fecha_visita': v.get('fecha_visita'),
                        'tipo_visita': 'Visita a Instalación',
                        'tipo': 'instalacion',
                        'observaciones': v.get('observaciones', ''),
                        'cliente_nombre': v.get('clientes', {}).get('nombre_cliente', 'Sin nombre') if v.get('clientes') else 'Sin nombre'
                    })
            
            # Obtener visitas a administradores asociadas a esta oportunidad
            visitas_admin_response = requests.get(
                f"{SUPABASE_URL}/rest/v1/visitas_administradores?oportunidad_id=eq.{oportunidad_id}&order=fecha_visita.desc",
                headers=HEADERS
            )
            
            visitas_admin = []
            if visitas_admin_response.status_code == 200:
                for v in visitas_admin_response.json():
                    visitas_admin.append({
                        'id': v.get('id'),
                        'fecha_visita': v.get('fecha_visita'),
                        'tipo_visita': f"Visita a Administrador: {v.get('administrador_fincas', 'N/A')}",
                        'tipo': 'administrador',
                        'observaciones': v.get('observaciones', ''),
                        'persona_contacto': v.get('persona_contacto', '')
                    })
            
            # Combinar y ordenar todas las visitas por fecha
            todas_visitas = visitas_seguimiento + visitas_admin
            todas_visitas.sort(key=lambda x: x['fecha_visita'], reverse=True)
            
            return render_template("ver_oportunidad.html", 
                                 oportunidad=oportunidad,
                                 visitas=todas_visitas)
        else:
            flash("Oportunidad no encontrada", "error")
            return redirect(url_for("oportunidades"))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("oportunidades"))


@app.route("/eliminar_oportunidad/<int:oportunidad_id>")
@helpers.login_required
@helpers.requiere_permiso('oportunidades', 'delete')
def eliminar_oportunidad(oportunidad_id):
    
    try:
        response_get = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=cliente_id",
            headers=HEADERS
        )
        
        if response_get.status_code == 200 and response_get.json():
            cliente_id = response_get.json()[0]["cliente_id"]
            
            response = requests.delete(
                f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
                headers=HEADERS
            )
            
            if response.status_code in [200, 204]:
                flash("Oportunidad eliminada correctamente", "success")
                return redirect(f"/ver_lead/{cliente_id}")
            else:
                flash("Error al eliminar oportunidad", "error")
                return redirect(url_for("oportunidades"))
        else:
            flash("Oportunidad no encontrada", "error")
            return redirect(url_for("oportunidades"))
    except Exception as e:
        flash(f"Error: {str(e)}", "error")
        return redirect(url_for("oportunidades"))


# ============================================
# NUEVAS RUTAS PARA ACCIONES
# ============================================

@app.route('/oportunidad/<int:oportunidad_id>/accion/add', methods=['POST'])
def add_accion(oportunidad_id):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    texto_accion = request.form.get('texto_accion', '').strip()
    
    if not texto_accion:
        flash('Debes escribir una acción', 'error')
        return redirect(url_for('ver_oportunidad', oportunidad_id=oportunidad_id))
    
    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=acciones",
        headers=HEADERS
    )
    
    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])
        if not isinstance(acciones, list):
            acciones = []
        
        # Añadir nueva acción
        acciones.append({
            'texto': texto_accion,
            'completada': False
        })
        
        # Actualizar en BD
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
            headers=HEADERS,
            json={'acciones': acciones}
        )
        
        flash('Acción añadida correctamente', 'success')
    
    return redirect(url_for('ver_oportunidad', oportunidad_id=oportunidad_id))


@app.route('/oportunidad/<int:oportunidad_id>/accion/toggle/<int:index>', methods=['POST'])
def toggle_accion(oportunidad_id, index):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=acciones",
        headers=HEADERS
    )
    
    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])
        
        if 0 <= index < len(acciones):
            # Toggle el estado
            acciones[index]['completada'] = not acciones[index]['completada']
            
            # Actualizar en BD
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
                headers=HEADERS,
                json={'acciones': acciones}
            )
    
    return redirect(url_for('ver_oportunidad', oportunidad_id=oportunidad_id))


@app.route('/oportunidad/<int:oportunidad_id>/accion/delete/<int:index>', methods=['POST'])
def delete_accion(oportunidad_id, index):
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}&select=acciones",
        headers=HEADERS
    )
    
    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])
        
        if 0 <= index < len(acciones):
            # Eliminar acción
            acciones.pop(index)
            
            # Actualizar en BD
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/oportunidades?id=eq.{oportunidad_id}",
                headers=HEADERS,
                json={'acciones': acciones}
            )
            
            flash('Acción eliminada', 'success')
    
    return redirect(url_for('ver_oportunidad', oportunidad_id=oportunidad_id))


# ============================================
# GESTIÓN DE ACCIONES PARA EQUIPOS
# ============================================

@app.route('/equipo/<int:equipo_id>/accion/add', methods=['POST'])
def add_accion_equipo(equipo_id):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    texto_accion = request.form.get('texto_accion', '').strip()

    if not texto_accion:
        flash('Debes escribir una acción', 'error')
        return redirect(url_for('ver_equipo', equipo_id=equipo_id))

    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}&select=acciones",
        headers=HEADERS
    )

    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])
        if not isinstance(acciones, list):
            acciones = []

        # Añadir nueva acción
        acciones.append({
            'texto': texto_accion,
            'completada': False
        })

        # Actualizar en BD
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}",
            headers=HEADERS,
            json={'acciones': acciones}
        )

        flash('Acción añadida', 'success')

    return redirect(url_for('ver_equipo', equipo_id=equipo_id))


@app.route('/equipo/<int:equipo_id>/accion/toggle/<int:index>', methods=['POST'])
def toggle_accion_equipo(equipo_id, index):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}&select=acciones",
        headers=HEADERS
    )

    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])

        if 0 <= index < len(acciones):
            # Toggle el estado
            acciones[index]['completada'] = not acciones[index]['completada']

            # Actualizar en BD
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}",
                headers=HEADERS,
                json={'acciones': acciones}
            )

    return redirect(url_for('ver_equipo', equipo_id=equipo_id))


@app.route('/equipo/<int:equipo_id>/accion/delete/<int:index>', methods=['POST'])
def delete_accion_equipo(equipo_id, index):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # Obtener acciones actuales
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}&select=acciones",
        headers=HEADERS
    )

    if response.status_code == 200 and response.json():
        acciones = response.json()[0].get('acciones', [])

        if 0 <= index < len(acciones):
            # Eliminar acción
            acciones.pop(index)

            # Actualizar en BD
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/equipos?id=eq.{equipo_id}",
                headers=HEADERS,
                json={'acciones': acciones}
            )

            flash('Acción eliminada', 'success')

    return redirect(url_for('ver_equipo', equipo_id=equipo_id))


# ============================================
# MÓDULO DE NOTIFICACIONES POR EMAIL
# ============================================

@app.route('/configuracion_avisos', methods=['GET', 'POST'])
def configuracion_avisos():
    if 'usuario_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['usuario_id']
    
    if request.method == 'POST':
        email = request.form.get('email_destinatario')
        primer_aviso = request.form.get('primer_aviso_despues_ipo')
        segundo_aviso = request.form.get('segundo_aviso_despues_ipo')
        dias_contratos = request.form.get('dias_aviso_antes_contrato')
        sistema_activo = request.form.get('sistema_activo') == 'on'
        frecuencia = request.form.get('frecuencia_chequeo')
        
        # Verificar si ya existe configuración
        config_check = requests.get(
            f"{SUPABASE_URL}/rest/v1/configuracion_avisos?usuario_id=eq.{user_id}",
            headers=HEADERS
        )
        
        data = {
            "email_destinatario": email,
            "primer_aviso_despues_ipo": primer_aviso,
            "segundo_aviso_despues_ipo": segundo_aviso,
            "dias_aviso_antes_contrato": dias_contratos,
            "sistema_activo": sistema_activo,
            "frecuencia_chequeo": frecuencia
        }
        
        if config_check.status_code == 200 and config_check.json():
            # Actualizar
            response = requests.patch(
                f"{SUPABASE_URL}/rest/v1/configuracion_avisos?usuario_id=eq.{user_id}",
                json=data,
                headers=HEADERS
            )
        else:
            # Insertar
            data["usuario_id"] = user_id
            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/configuracion_avisos",
                json=data,
                headers=HEADERS
            )
        
        if response.status_code in [200, 201, 204]:
            flash('✅ Configuración guardada correctamente', 'success')
        else:
            flash(f'❌ Error al guardar configuración: {response.text}', 'error')
        
        return redirect(url_for('configuracion_avisos'))
    
    # GET - Mostrar formulario
    config_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/configuracion_avisos?usuario_id=eq.{user_id}",
        headers=HEADERS
    )
    
    if config_response.status_code == 200 and config_response.json():
        config_data = config_response.json()[0]
        
        # Formatear fecha si existe
        if config_data.get('ultima_ejecucion'):
            try:
                fecha_obj = datetime.fromisoformat(config_data['ultima_ejecucion'].replace('Z', '+00:00'))
                config_data['ultima_ejecucion'] = fecha_obj.strftime('%d/%m/%Y %H:%M')
            except:
                pass
    else:
        # Valores por defecto
        config_data = {
            'email_destinatario': session.get('email', ''),
            'primer_aviso_despues_ipo': 15,
            'segundo_aviso_despues_ipo': 30,
            'dias_aviso_antes_contrato': 30,
            'sistema_activo': True,
            'frecuencia_chequeo': 'diario',
            'ultima_ejecucion': None
        }
    
    return render_template('configuracion_avisos.html', config=config_data)


@app.route('/enviar_avisos_manual')
def enviar_avisos_manual():
    if "usuario" not in session:
        return redirect(url_for("login"))
    
    usuario_id = session.get("usuario_id")
    
    # Obtener configuración del usuario
    config = requests.get(
        f"{SUPABASE_URL}/rest/v1/configuracion_avisos?usuario_id=eq.{usuario_id}",
        headers=HEADERS
    )
    
    if config.status_code != 200 or not config.json():
        return "No hay configuración de avisos", 400
    
    config_data = config.json()[0]
    
    if not config_data.get('sistema_activo'):
        return "Las notificaciones están desactivadas", 400
    
    # Enviar avisos
    resultado = enviar_avisos_email(config_data)
    
    # Actualizar última ejecución
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/configuracion_avisos?usuario_id=eq.{usuario_id}",
        json={"ultima_ejecucion": datetime.now().isoformat()},
        headers=HEADERS
    )
    
    return f"<h3>{resultado}</h3><br><a href='/home'>Volver al inicio</a>"

# ============================================
# ADMINISTRADORES
# ============================================

# Dashboard de Administradores (con pestañas)
@app.route("/administradores_dashboard", methods=["GET"])
def administradores_dashboard():
    if "usuario" not in session:
        return redirect("/")

    # Determinar pestaña activa
    tab = request.args.get("tab", "administradores")  # administradores | visitas

    # ============================================
    # TAB: ADMINISTRADORES
    # ============================================
    if tab == "administradores":
        # Búsqueda
        buscar = request.args.get("buscar", "")

        # Paginación
        try:
            page = int(request.args.get("page", 1))
        except (ValueError, TypeError):
            page = 1

        limit = 10  # Paginación de 10 administradores por página
        offset = (page - 1) * limit

        # Si hay búsqueda, usar RPC para búsqueda sin acentos
        if buscar:
            # Usar función RPC para búsqueda sin acentos
            rpc_url = f"{SUPABASE_URL}/rest/v1/rpc/buscar_administradores_sin_acentos"

            rpc_params = {
                "termino_busqueda": buscar,
                "limite": limit,
                "desplazamiento": offset
            }

            try:
                response = requests.post(rpc_url, json=rpc_params, headers=HEADERS, timeout=10)

                if response.status_code != 200:
                    print(f"Error al buscar administradores: {response.status_code} - {response.text}")
                    flash(f"Error al buscar administradores (Código: {response.status_code})", "error")
                    return render_template(
                        "administradores_dashboard.html",
                        tab=tab,
                        administradores=[],
                        buscar=buscar,
                        page=1,
                        total_pages=1,
                        total_registros=0
                    )

                administradores_data = response.json()

                # Obtener total_count del primer resultado si existe
                total_registros = administradores_data[0].get('total_count', 0) if administradores_data else 0
                total_pages = max(1, (total_registros + limit - 1) // limit)

                # Remover total_count de cada registro (no es parte del modelo)
                administradores_base = [{k: v for k, v in admin.items() if k != 'total_count'} for admin in administradores_data]

                # Obtener IDs de los administradores encontrados
                admin_ids = [admin['id'] for admin in administradores_base]

                if admin_ids:
                    # Obtener conteos de clientes para estos administradores
                    clientes_response = requests.get(
                        f"{SUPABASE_URL}/rest/v1/clientes?select=administrador_id&administrador_id=in.({','.join(map(str, admin_ids))})",
                        headers=HEADERS,
                        timeout=10
                    )
                    clientes_data = clientes_response.json() if clientes_response.status_code == 200 else []

                    # Obtener conteos de oportunidades (a través de clientes)
                    oportunidades_response = requests.get(
                        f"{SUPABASE_URL}/rest/v1/oportunidades?select=cliente_id,clientes!inner(administrador_id)&clientes.administrador_id=in.({','.join(map(str, admin_ids))})",
                        headers=HEADERS,
                        timeout=10
                    )
                    oportunidades_data = oportunidades_response.json() if oportunidades_response.status_code == 200 else []

                    # Contar clientes por administrador
                    clientes_por_admin = {}
                    for cliente in clientes_data:
                        admin_id = cliente.get('administrador_id')
                        if admin_id:
                            clientes_por_admin[admin_id] = clientes_por_admin.get(admin_id, 0) + 1

                    # Contar oportunidades por administrador
                    oportunidades_por_admin = {}
                    for oportunidad in oportunidades_data:
                        cliente_info = oportunidad.get('clientes')
                        if cliente_info:
                            admin_id = cliente_info.get('administrador_id')
                            if admin_id:
                                oportunidades_por_admin[admin_id] = oportunidades_por_admin.get(admin_id, 0) + 1

                    # Agregar conteos a cada administrador
                    for admin in administradores_base:
                        admin_id = admin['id']
                        admin['num_oportunidades'] = oportunidades_por_admin.get(admin_id, 0)
                        admin['num_instalaciones'] = clientes_por_admin.get(admin_id, 0)

                    # Ordenar por: 1) num_oportunidades DESC, 2) num_instalaciones DESC
                    administradores_base.sort(
                        key=lambda x: (x.get('num_oportunidades', 0), x.get('num_instalaciones', 0)),
                        reverse=True
                    )

                administradores = administradores_base

            except Exception as e:
                print(f"Excepción al buscar administradores: {str(e)}")
                flash(f"Error de conexión al buscar administradores", "error")
                administradores = []
                total_registros = 0
                total_pages = 1

        else:
            # Búsqueda normal sin filtro de texto
            # Primero obtener todos los administradores con conteo
            url_count = f"{SUPABASE_URL}/rest/v1/administradores?select=*"
            headers_with_count = HEADERS.copy()
            headers_with_count["Prefer"] = "count=exact"

            try:
                response = requests.get(url_count, headers=headers_with_count, timeout=10)

                # 200 = OK, 206 = Partial Content (respuesta válida con paginación)
                if response.status_code not in [200, 206]:
                    print(f"Error al cargar administradores: {response.status_code} - {response.text}")
                    flash(f"Error al cargar administradores desde la base de datos (Código: {response.status_code})", "error")
                    # Renderizar con datos vacíos
                    return render_template(
                        "administradores_dashboard.html",
                        tab=tab,
                        administradores=[],
                        buscar=buscar,
                        page=1,
                        total_pages=1,
                        total_registros=0
                    )

                # Obtener total de registros del header Content-Range
                content_range = response.headers.get("Content-Range", "*/0")
                total_registros = int(content_range.split("/")[-1])

                # Parsear respuesta JSON
                administradores_base = response.json()

                # Obtener conteos de clientes por administrador
                clientes_response = requests.get(
                    f"{SUPABASE_URL}/rest/v1/clientes?select=administrador_id",
                    headers=HEADERS,
                    timeout=10
                )
                clientes_data = clientes_response.json() if clientes_response.status_code == 200 else []

                # Obtener conteos de oportunidades (a través de clientes)
                oportunidades_response = requests.get(
                    f"{SUPABASE_URL}/rest/v1/oportunidades?select=cliente_id,clientes!inner(administrador_id)",
                    headers=HEADERS,
                    timeout=10
                )
                oportunidades_data = oportunidades_response.json() if oportunidades_response.status_code == 200 else []

                # Contar clientes por administrador
                clientes_por_admin = {}
                for cliente in clientes_data:
                    admin_id = cliente.get('administrador_id')
                    if admin_id:
                        clientes_por_admin[admin_id] = clientes_por_admin.get(admin_id, 0) + 1

                # Contar oportunidades por administrador
                oportunidades_por_admin = {}
                for oportunidad in oportunidades_data:
                    cliente_info = oportunidad.get('clientes')
                    if cliente_info:
                        admin_id = cliente_info.get('administrador_id')
                        if admin_id:
                            oportunidades_por_admin[admin_id] = oportunidades_por_admin.get(admin_id, 0) + 1

                # Agregar conteos a cada administrador
                for admin in administradores_base:
                    admin_id = admin['id']
                    admin['num_oportunidades'] = oportunidades_por_admin.get(admin_id, 0)
                    admin['num_instalaciones'] = clientes_por_admin.get(admin_id, 0)

                # Ordenar por: 1) num_oportunidades DESC, 2) num_instalaciones DESC
                administradores_base.sort(
                    key=lambda x: (x.get('num_oportunidades', 0), x.get('num_instalaciones', 0)),
                    reverse=True
                )

                # Aplicar paginación después del ordenamiento
                start_idx = offset
                end_idx = offset + limit
                administradores = administradores_base[start_idx:end_idx]

            except Exception as e:
                print(f"Excepción al cargar administradores: {str(e)}")
                flash(f"Error de conexión al cargar administradores", "error")
                administradores = []
                total_registros = 0

            # Calcular páginas
            total_pages = max(1, (total_registros + limit - 1) // limit)  # Al menos 1 página

        # Limpiar None en todos los casos
        try:
            administradores = [limpiar_none(admin) for admin in administradores]
        except Exception as e:
            print(f"Error al limpiar datos: {e}")
            administradores = []

        # Renderizar template con los resultados
        return render_template(
            "administradores_dashboard.html",
            tab=tab,
            administradores=administradores,
            buscar=buscar,
            page=page,
            total_pages=total_pages,
            total_registros=total_registros
        )

    # ============================================
    # TAB: VISITAS
    # ============================================
    elif tab == "visitas":
        # Paginación
        try:
            page = int(request.args.get("page", 1))
        except (ValueError, TypeError):
            page = 1

        per_page = 25
        offset = (page - 1) * per_page

        # Obtener registros paginados con JOIN a administradores y conteo
        data_url = f"{SUPABASE_URL}/rest/v1/visitas_administradores?select=*,administradores(nombre_empresa)&order=fecha_visita.desc&limit={per_page}&offset={offset}"

        # Headers para obtener el conteo total
        headers_with_count = HEADERS.copy()
        headers_with_count["Prefer"] = "count=exact"

        try:
            response = requests.get(data_url, headers=headers_with_count, timeout=10)

            # 200 = OK, 206 = Partial Content (respuesta válida con paginación)
            if response.status_code not in [200, 206]:
                print(f"Error al cargar visitas: {response.status_code} - {response.text}")
                flash(f"Error al cargar visitas desde la base de datos (Código: {response.status_code})", "error")
                # Renderizar con datos vacíos
                return render_template(
                    "administradores_dashboard.html",
                    tab=tab,
                    visitas=[],
                    page=1,
                    total_pages=1,
                    total_registros=0
                )

            # Obtener total de registros del header Content-Range
            try:
                content_range = response.headers.get("Content-Range", "*/0")
                total_registros = int(content_range.split("/")[-1])
            except Exception as e:
                print(f"Error al parsear Content-Range: {e}")
                total_registros = 0

            # Parsear respuesta JSON
            try:
                visitas = response.json()
            except Exception as e:
                print(f"Error al parsear JSON: {e}")
                flash(f"Error al procesar datos de visitas", "error")
                return render_template(
                    "administradores_dashboard.html",
                    tab=tab,
                    visitas=[],
                    page=1,
                    total_pages=1,
                    total_registros=0
                )

            # Calcular páginas
            total_pages = max(1, (total_registros + per_page - 1) // per_page)

            # Limpiar None
            try:
                visitas = [limpiar_none(v) for v in visitas]
            except Exception as e:
                print(f"Error al limpiar datos: {e}")
                visitas = []

            return render_template(
                "administradores_dashboard.html",
                tab=tab,
                visitas=visitas,
                page=page,
                total_pages=total_pages,
                total_registros=total_registros
            )

        except requests.exceptions.Timeout:
            print(f"Error de timeout al cargar visitas")
            flash(f"Error de conexión: La base de datos tardó demasiado en responder. Por favor, intente nuevamente.", "error")
            return render_template(
                "administradores_dashboard.html",
                tab=tab,
                visitas=[],
                page=1,
                total_pages=1,
                total_registros=0
            )
        except requests.exceptions.RequestException as e:
            print(f"Error de conexión al cargar visitas: {e}")
            flash(f"Error de conexión al cargar visitas. Por favor, verifique su conexión a internet.", "error")
            return render_template(
                "administradores_dashboard.html",
                tab=tab,
                visitas=[],
                page=1,
                total_pages=1,
                total_registros=0
            )
        except Exception as e:
            print(f"Error inesperado al cargar visitas: {e}")
            flash(f"Error inesperado al cargar visitas. Por favor, contacte al administrador del sistema.", "error")
            return render_template(
                "administradores_dashboard.html",
                tab=tab,
                visitas=[],
                page=1,
                total_pages=1,
                total_registros=0
            )

    # ============================================
    # TAB INVÁLIDO - Redirigir a tab por defecto
    # ============================================
    else:
        print(f"Tab inválido recibido: {tab}")
        return redirect("/administradores_dashboard?tab=administradores")


# Alta de Administrador
@app.route("/nuevo_administrador", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('administradores', 'write')
def nuevo_administrador():

    if request.method == "POST":
        data = {
            "nombre_empresa": request.form.get("nombre_empresa"),
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "direccion": request.form.get("direccion") or None,
            "localidad": request.form.get("localidad") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Validar campo obligatorio
        if not data["nombre_empresa"]:
            flash("El nombre de la empresa es obligatorio", "error")
            return redirect(request.referrer)

        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/administradores",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 201]:
            flash("Administrador creado correctamente", "success")
            return redirect("/administradores_dashboard")
        else:
            flash(f"Error al crear administrador: {response.text}", "error")
            return redirect(request.referrer)

    return render_template("nuevo_administrador.html")


# Ver Administrador
@app.route("/ver_administrador/<int:admin_id>", methods=["GET"])
def ver_administrador(admin_id):
    if "usuario" not in session:
        return redirect("/")

    # Obtener administrador
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{admin_id}",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        return "Administrador no encontrado", 404

    administrador = limpiar_none(response.json()[0])

    # Obtener clientes asociados
    clientes_response = requests.get(
        f"{SUPABASE_URL}/rest/v1/clientes?administrador_id=eq.{admin_id}&select=*",
        headers=HEADERS
    )

    clientes = []
    if clientes_response.status_code == 200:
        clientes = [limpiar_none(c) for c in clientes_response.json()]

    # Obtener oportunidades de todos los clientes asociados
    oportunidades = []
    stats = {
        'total': 0,
        'activas': 0,
        'ganadas': 0,
        'perdidas': 0,
        'valor_total': 0,
        'valor_activas': 0
    }

    if clientes:
        # Obtener IDs de clientes
        cliente_ids = [str(c['id']) for c in clientes]

        # Consultar oportunidades de estos clientes
        oportunidades_response = requests.get(
            f"{SUPABASE_URL}/rest/v1/oportunidades?cliente_id=in.({','.join(cliente_ids)})&select=*,clientes(direccion,localidad)&order=fecha_creacion.desc",
            headers=HEADERS
        )

        if oportunidades_response.status_code == 200:
            oportunidades = oportunidades_response.json()

            # Calcular estadísticas
            for op in oportunidades:
                stats['total'] += 1
                estado = op.get('estado', '').lower()

                if estado == 'activa':
                    stats['activas'] += 1
                    if op.get('valor_estimado'):
                        stats['valor_activas'] += float(op.get('valor_estimado', 0))
                elif estado == 'ganada':
                    stats['ganadas'] += 1
                elif estado == 'perdida':
                    stats['perdidas'] += 1

                if op.get('valor_estimado'):
                    stats['valor_total'] += float(op.get('valor_estimado', 0))

    return render_template(
        "ver_administrador.html",
        administrador=administrador,
        clientes=clientes,
        oportunidades=oportunidades,
        stats=stats
    )


# Editar Administrador
@app.route("/editar_administrador/<int:admin_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('administradores', 'write')
def editar_administrador(admin_id):

    if request.method == "POST":
        data = {
            "nombre_empresa": request.form.get("nombre_empresa"),
            "telefono": request.form.get("telefono") or None,
            "email": request.form.get("email") or None,
            "direccion": request.form.get("direccion") or None,
            "localidad": request.form.get("localidad") or None,
            "observaciones": request.form.get("observaciones") or None
        }

        # Validar campo obligatorio
        if not data["nombre_empresa"]:
            flash("El nombre de la empresa es obligatorio", "error")
            return redirect(request.referrer)

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{admin_id}",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 201, 204]:
            flash("Administrador actualizado correctamente", "success")
            return redirect(f"/ver_administrador/{admin_id}")
        else:
            flash(f"Error al actualizar administrador: {response.text}", "error")
            return redirect(request.referrer)

    # GET - Obtener datos del administrador
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{admin_id}",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        return "Administrador no encontrado", 404

    administrador = limpiar_none(response.json()[0])

    return render_template("editar_administrador.html", administrador=administrador)


# Eliminar Administrador
@app.route("/eliminar_administrador/<int:admin_id>", methods=["GET"])
@helpers.login_required
@helpers.requiere_permiso('administradores', 'delete')
def eliminar_administrador(admin_id):

    # Verificar si tiene clientes asociados
    clientes_check = requests.get(
        f"{SUPABASE_URL}/rest/v1/clientes?administrador_id=eq.{admin_id}&select=count",
        headers=HEADERS
    )

    if clientes_check.status_code == 200 and clientes_check.json():
        num_clientes = len(clientes_check.json())
        if num_clientes > 0:
            flash(f"No se puede eliminar: el administrador tiene {num_clientes} cliente(s) asociado(s)", "error")
            return redirect(f"/ver_administrador/{admin_id}")

    # Eliminar administrador
    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/administradores?id=eq.{admin_id}",
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Administrador eliminado correctamente", "success")
        return redirect("/administradores_dashboard")
    else:
        flash(f"Error al eliminar administrador: {response.text}", "error")
        return redirect(f"/ver_administrador/{admin_id}")


# ============================================
# RUTA DE PRUEBA - DROPDOWN ADMINISTRADORES
# ============================================

@app.route("/test_dropdown_admin")
def test_dropdown_admin():
    if "usuario" not in session:
        return redirect("/")

    # Test 1: Consulta directa (sin caché)
    test_direct = {"success": False, "status": 0, "count": 0, "error": ""}
    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/administradores?select=id,nombre_empresa&order=nombre_empresa.asc",
            headers=HEADERS,
            timeout=10
        )
        test_direct["status"] = response.status_code
        if response.status_code == 200:
            test_direct["success"] = True
            test_direct["data"] = response.json()
            test_direct["count"] = len(test_direct["data"])
        else:
            test_direct["error"] = response.text[:500]
    except Exception as e:
        test_direct["error"] = f"{type(e).__name__}: {str(e)}"

    # Test 2: Obtener desde caché
    administradores_cached = get_administradores_cached()

    # HTML de prueba mejorado
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Test Dropdown Administradores - Diagnóstico</title>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; max-width: 900px; margin: 0 auto; background: #f5f5f5; }}
            .debug-info {{ background: white; padding: 20px; margin: 20px 0; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
            .success {{ color: #28a745; font-weight: bold; }}
            .error {{ color: #dc3545; font-weight: bold; }}
            .warning {{ color: #ffc107; font-weight: bold; }}
            h1 {{ color: #333; }}
            h3 {{ color: #666; margin-top: 0; }}
            pre {{ background: #f8f9fa; padding: 15px; border-radius: 5px; overflow-x: auto; border: 1px solid #dee2e6; }}
            .status-badge {{ display: inline-block; padding: 5px 10px; border-radius: 5px; font-size: 14px; }}
            .badge-success {{ background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }}
            .badge-error {{ background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            td {{ padding: 8px; border-bottom: 1px solid #dee2e6; }}
            td:first-child {{ font-weight: bold; width: 200px; }}
            .back-link {{ display: inline-block; margin-top: 20px; padding: 10px 20px; background: #366092; color: white; text-decoration: none; border-radius: 5px; }}
            .back-link:hover {{ background: #2a4a70; }}
        </style>
    </head>
    <body>
        <h1>🧪 Test Dropdown Administradores - Diagnóstico Completo</h1>

        <div class="debug-info">
            <h3>🔍 Test 1: Consulta Directa a Supabase (sin caché)</h3>
            <table>
                <tr>
                    <td>Estado:</td>
                    <td>
                        {'<span class="status-badge badge-success">✅ Éxito</span>' if test_direct['success'] else '<span class="status-badge badge-error">❌ Error</span>'}
                    </td>
                </tr>
                <tr>
                    <td>Status Code:</td>
                    <td><strong>{test_direct['status']}</strong></td>
                </tr>
                <tr>
                    <td>Administradores encontrados:</td>
                    <td><strong class="{'success' if test_direct['count'] > 0 else 'error'}">{test_direct['count']}</strong></td>
                </tr>
                {'<tr><td>Error:</td><td><pre>' + test_direct['error'] + '</pre></td></tr>' if test_direct['error'] else ''}
            </table>

            {f"<h4>📋 Datos obtenidos:</h4><pre>{test_direct.get('data', [])}</pre>" if test_direct['success'] and test_direct['count'] > 0 else ''}
        </div>

        <div class="debug-info">
            <h3>💾 Test 2: Sistema de Caché</h3>
            <table>
                <tr>
                    <td>Administradores en caché:</td>
                    <td><strong class="{'success' if len(administradores_cached) > 0 else 'error'}">{len(administradores_cached)}</strong></td>
                </tr>
                <tr>
                    <td>Timestamp del caché:</td>
                    <td>{cache_administradores['timestamp'] or 'Sin inicializar'}</td>
                </tr>
            </table>

            {f"<h4>📋 Datos en caché:</h4><pre>{administradores_cached}</pre>" if len(administradores_cached) > 0 else ''}
        </div>

        <div class="debug-info">
            <h3>🔧 Configuración de Supabase</h3>
            <table>
                <tr>
                    <td>URL:</td>
                    <td><code>{SUPABASE_URL}</code></td>
                </tr>
                <tr>
                    <td>API Key configurada:</td>
                    <td>{'✅ Sí' if SUPABASE_KEY else '❌ No'}</td>
                </tr>
            </table>
        </div>

        <div class="debug-info">
            <h3>💡 Diagnóstico</h3>
            {'<p class="success">✅ Todo funciona correctamente. Hay ' + str(test_direct['count']) + ' administradores en la base de datos.</p>' if test_direct['success'] and test_direct['count'] > 0 else ''}
            {'<p class="warning">⚠️ La conexión a Supabase funciona pero <strong>NO HAY ADMINISTRADORES</strong> en la base de datos. Necesitas crear administradores primero en <a href="/nuevo_administrador">/nuevo_administrador</a></p>' if test_direct['success'] and test_direct['count'] == 0 else ''}
            {'<p class="error">❌ Error al conectar con Supabase. Verifica:<br>1. Que la URL de Supabase sea correcta<br>2. Que el API Key esté configurado<br>3. Que la tabla "administradores" exista<br>4. Que tengas permisos de lectura</p>' if not test_direct['success'] else ''}
        </div>

        <a href="/home" class="back-link">← Volver al inicio</a>

        <script>
            console.log('📊 Diagnóstico completo:');
            console.log('Test directo:', {test_direct});
            console.log('Caché:', {administradores_cached});
        </script>
    </body>
    </html>
    """

    return html

@app.route("/admin/clear_cache")
def clear_cache():
    """Endpoint para limpiar manualmente el caché de administradores"""
    if "usuario" not in session:
        return redirect("/")

    # Limpiar el caché
    cache_administradores['data'] = []
    cache_administradores['timestamp'] = None

    # Forzar recarga inmediata
    administradores = get_administradores_cached()

    return f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Caché Actualizado</title>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 40px; max-width: 600px; margin: 0 auto; text-align: center; }}
            .success {{ color: #28a745; font-size: 48px; margin-bottom: 20px; }}
            h1 {{ color: #333; }}
            .count {{ font-size: 32px; font-weight: bold; color: #366092; margin: 20px 0; }}
            .btn {{ display: inline-block; margin: 10px; padding: 12px 24px; background: #366092; color: white; text-decoration: none; border-radius: 5px; }}
            .btn:hover {{ background: #2a4a70; }}
        </style>
    </head>
    <body>
        <div class="success">✅</div>
        <h1>Caché Actualizado</h1>
        <div class="count">{len(administradores)} administradores cargados</div>
        <p>El caché se ha limpiado y recargado correctamente.</p>
        <a href="/test_dropdown_admin" class="btn">Ver Diagnóstico</a>
        <a href="/visita_administrador" class="btn">Probar Dropdown</a>
        <a href="/home" class="btn">Volver al Inicio</a>
    </body>
    </html>
    """

# ============================================
# MÓDULO: GESTIÓN DE INSPECCIONES (IPOs)
# ============================================

# Dashboard de Inspecciones (vista principal)
@app.route("/inspecciones")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'read')
def inspecciones_dashboard():
    """Dashboard principal de inspecciones con alertas y estados"""

    # Obtener todas las inspecciones con información del OCA
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?select=*,oca:ocas(nombre)&order=fecha_inspeccion.desc",
        headers=HEADERS
    )

    inspecciones = []
    if response.status_code == 200:
        inspecciones = response.json()

    # Calcular alertas y estadísticas basadas en inspecciones
    hoy = date.today()
    alertas_criticas = []
    alertas_urgentes = []
    alertas_proximas = []
    alertas_normales = []  # Inspecciones sin urgencia

    # Procesar inspecciones - categorizar según urgencia de segunda inspección
    # La fecha_segunda_inspeccion es la FECHA LÍMITE para tener los defectos solucionados
    # SOLO alertar si NO se ha realizado la 2ª inspección (fecha_segunda_realizada es NULL)
    for inspeccion in inspecciones:
        # Categorizar según urgencia para solucionar defectos antes de la fecha límite
        if inspeccion.get('fecha_segunda_realizada'):
            inspeccion['categoria_segunda'] = 'realizadas'
            # Ya pasó la 2ª inspección, defectos subsanados
            alertas_normales.append(('defectos', inspeccion))
        elif inspeccion.get('fecha_segunda_inspeccion'):
            try:
                fecha_limite = datetime.strptime(inspeccion['fecha_segunda_inspeccion'].split('T')[0], '%Y-%m-%d').date()
                dias_restantes = (fecha_limite - hoy).days

                inspeccion['dias_hasta_segunda'] = dias_restantes

                # Categorizar por urgencia de defectos
                if fecha_limite < hoy:
                    inspeccion['categoria_segunda'] = 'vencidas'
                    # CRÍTICO: Fecha límite vencida, defectos sin subsanar
                    alertas_criticas.append(('defectos', inspeccion))
                elif dias_restantes <= 30:
                    inspeccion['categoria_segunda'] = 'este-mes'
                    # URGENTE: Menos de 30 días para solucionar defectos
                    alertas_urgentes.append(('defectos', inspeccion))
                elif dias_restantes <= 60:
                    inspeccion['categoria_segunda'] = 'proximo-mes'
                    # PRÓXIMO: Entre 30-60 días para solucionar defectos
                    alertas_proximas.append(('defectos', inspeccion))
                else:
                    inspeccion['categoria_segunda'] = 'pendiente'
                    # SIN URGENCIA: Más de 60 días para solucionar defectos
                    alertas_normales.append(('defectos', inspeccion))
            except:
                # Error al parsear fecha, mostrar en normales
                inspeccion['categoria_segunda'] = 'sin-fecha'
                alertas_normales.append(('defectos', inspeccion))
        else:
            # Sin fecha de segunda inspección, mostrar en normales
            inspeccion['categoria_segunda'] = 'sin-fecha'
            alertas_normales.append(('defectos', inspeccion))

    # Ordenar alertas críticas: más antiguas primero (mayor tiempo vencido)
    alertas_criticas.sort(key=lambda x: x[1].get('dias_hasta_segunda', 0))

    # Obtener lista de OCAs para filtros
    response_ocas = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?select=id,nombre&activo=eq.true&order=nombre.asc",
        headers=HEADERS
    )
    ocas = []
    if response_ocas.status_code == 200:
        ocas = response_ocas.json()

    return render_template(
        "inspecciones_dashboard.html",
        inspecciones=inspecciones,
        alertas_criticas=alertas_criticas,
        alertas_urgentes=alertas_urgentes,
        alertas_proximas=alertas_proximas,
        alertas_normales=alertas_normales,
        ocas=ocas
    )

# Dashboard de Defectos (vista dedicada)
@app.route("/defectos_dashboard")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'read')
def defectos_dashboard():
    """Dashboard principal de defectos con estadísticas y filtros"""

    # Obtener todos los defectos con información de urgencia usando la vista
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/v_defectos_con_urgencia?select=*&order=fecha_limite.asc",
        headers=HEADERS
    )

    todos_defectos = []
    if response.status_code == 200:
        todos_defectos = response.json()

    # Calcular estadísticas generales
    total_defectos = len(todos_defectos)
    defectos_pendientes = [d for d in todos_defectos if d.get('estado') == 'PENDIENTE']
    defectos_subsanados = [d for d in todos_defectos if d.get('estado') == 'SUBSANADO']

    # Aplicar filtros
    filtro_tecnico = request.args.get('tecnico', '')
    filtro_material = request.args.get('material', '')
    filtro_stock = request.args.get('stock', '')

    if filtro_tecnico:
        if filtro_tecnico == 'sin_asignar':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('tecnico_asignado')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('tecnico_asignado') == filtro_tecnico]

    if filtro_material:
        if filtro_material == 'sin_definir':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('gestion_material')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('gestion_material') == filtro_material]

    if filtro_stock:
        if filtro_stock == 'sin_definir':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('estado_stock')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('estado_stock') == filtro_stock]

    # Clasificar por urgencia
    defectos_vencidos = [d for d in defectos_pendientes if d.get('nivel_urgencia') == 'VENCIDO']
    defectos_urgentes = [d for d in defectos_pendientes if d.get('nivel_urgencia') == 'URGENTE']
    defectos_proximos = [d for d in defectos_pendientes if d.get('nivel_urgencia') == 'PROXIMO']
    defectos_normales = [d for d in defectos_pendientes if d.get('nivel_urgencia') == 'NORMAL']

    # Estadísticas por calificación (solo pendientes)
    defectos_dl = [d for d in defectos_pendientes if d.get('calificacion') == 'DL']  # Defecto Leve
    defectos_dg = [d for d in defectos_pendientes if d.get('calificacion') == 'DG']  # Defecto Grave
    defectos_dmg = [d for d in defectos_pendientes if d.get('calificacion') == 'DMG']  # Defecto Muy Grave

    # Obtener información adicional de inspecciones para enriquecer datos
    response_insp = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?select=id,maquina,direccion,poblacion,fecha_inspeccion,oca_id,oca:ocas(nombre)",
        headers=HEADERS
    )

    inspecciones_dict = {}
    if response_insp.status_code == 200:
        inspecciones = response_insp.json()
        for insp in inspecciones:
            inspecciones_dict[insp['id']] = insp

    # Enriquecer defectos con información de inspección
    for defecto in todos_defectos:
        insp_id = defecto.get('inspeccion_id')
        if insp_id and insp_id in inspecciones_dict:
            insp = inspecciones_dict[insp_id]
            defecto['direccion'] = insp.get('direccion')
            defecto['poblacion'] = insp.get('poblacion')
            defecto['fecha_inspeccion'] = insp.get('fecha_inspeccion')
            defecto['oca_nombre'] = insp.get('oca', {}).get('nombre') if insp.get('oca') else None

    return render_template(
        "defectos_dashboard.html",
        total_defectos=total_defectos,
        total_pendientes=len(defectos_pendientes),
        total_subsanados=len(defectos_subsanados),
        defectos_vencidos=defectos_vencidos,
        defectos_urgentes=defectos_urgentes,
        defectos_proximos=defectos_proximos,
        defectos_normales=defectos_normales,
        defectos_subsanados=defectos_subsanados,
        defectos_dl=len(defectos_dl),
        defectos_dg=len(defectos_dg),
        defectos_dmg=len(defectos_dmg),
        todos_defectos=todos_defectos
    )

# Exportar Defectos a PDF
@app.route("/exportar_defectos_pdf")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'read')
def exportar_defectos_pdf():
    """Exporta defectos a PDF en formato horizontal, agrupados por máquina"""

    # Obtener todos los defectos con información de urgencia usando la vista
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/v_defectos_con_urgencia?select=*&order=fecha_limite.asc",
        headers=HEADERS
    )

    todos_defectos = []
    if response.status_code == 200:
        todos_defectos = response.json()

    # Aplicar filtros (mismo código que el dashboard)
    defectos_pendientes = [d for d in todos_defectos if d.get('estado') == 'PENDIENTE']

    filtro_tecnico = request.args.get('tecnico', '')
    filtro_material = request.args.get('material', '')
    filtro_stock = request.args.get('stock', '')

    if filtro_tecnico:
        if filtro_tecnico == 'sin_asignar':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('tecnico_asignado')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('tecnico_asignado') == filtro_tecnico]

    if filtro_material:
        if filtro_material == 'sin_definir':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('gestion_material')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('gestion_material') == filtro_material]

    if filtro_stock:
        if filtro_stock == 'sin_definir':
            defectos_pendientes = [d for d in defectos_pendientes if not d.get('estado_stock')]
        else:
            defectos_pendientes = [d for d in defectos_pendientes if d.get('estado_stock') == filtro_stock]

    # Obtener información adicional de inspecciones para enriquecer datos
    response_insp = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?select=id,maquina,direccion,poblacion,fecha_inspeccion,oca_id,oca:ocas(nombre)",
        headers=HEADERS
    )

    inspecciones_dict = {}
    if response_insp.status_code == 200:
        inspecciones = response_insp.json()
        for insp in inspecciones:
            inspecciones_dict[insp['id']] = insp

    # Enriquecer defectos con información de inspección
    for defecto in defectos_pendientes:
        insp_id = defecto.get('inspeccion_id')
        if insp_id and insp_id in inspecciones_dict:
            insp = inspecciones_dict[insp_id]
            defecto['direccion'] = insp.get('direccion')
            defecto['poblacion'] = insp.get('poblacion')
            defecto['fecha_inspeccion'] = insp.get('fecha_inspeccion')
            defecto['oca_nombre'] = insp.get('oca', {}).get('nombre') if insp.get('oca') else None

    # Agrupar defectos por máquina
    defectos_por_maquina = {}
    for defecto in defectos_pendientes:
        maquina = defecto.get('maquina', 'Sin especificar')
        if maquina not in defectos_por_maquina:
            defectos_por_maquina[maquina] = []
        defectos_por_maquina[maquina].append(defecto)

    # Generar PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    # Elementos del PDF
    elementos = []

    # Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo-fedes-ascensores.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=4*cm, height=1.6*cm)
        elementos.append(logo)
        elementos.append(Spacer(1, 0.3*cm))

    # Título
    styles = getSampleStyleSheet()
    titulo = Paragraph("<b>LISTADO DE DEFECTOS PENDIENTES</b>", styles['Title'])
    elementos.append(titulo)

    # Fecha de generación
    fecha_hoy = datetime.now().strftime('%d/%m/%Y %H:%M')
    fecha_texto = Paragraph(f"<i>Generado el: {fecha_hoy}</i>", styles['Normal'])
    elementos.append(fecha_texto)
    elementos.append(Spacer(1, 0.5*cm))

    # Crear tabla para cada máquina
    for maquina, defectos in sorted(defectos_por_maquina.items()):
        # Encabezado de máquina
        maquina_texto = Paragraph(f"<b>Máquina: {maquina}</b>", styles['Heading2'])
        elementos.append(maquina_texto)
        elementos.append(Spacer(1, 0.2*cm))

        # Datos de la tabla
        data = [['Descripción', 'Calificación', 'Vencimiento', 'Estado', 'Técnico', 'Estado Material']]

        for defecto in defectos:
            # Formatear valores
            descripcion = defecto.get('descripcion', '')  # Descripción completa
            calificacion = defecto.get('calificacion', '-')

            # Formatear fecha de vencimiento
            fecha_limite = defecto.get('fecha_limite', '')
            if fecha_limite:
                try:
                    # Formato: YYYY-MM-DD -> DD/MM/YYYY
                    fecha_limpia = fecha_limite.split('T')[0] if 'T' in str(fecha_limite) else str(fecha_limite)
                    fecha_obj = datetime.strptime(fecha_limpia, '%Y-%m-%d')
                    vencimiento = fecha_obj.strftime('%d/%m/%Y')
                except:
                    vencimiento = '-'
            else:
                vencimiento = '-'

            estado = defecto.get('estado', '-')
            tecnico = defecto.get('tecnico_asignado', '-') or '-'
            estado_material = defecto.get('estado_stock', '-') or '-'

            data.append([descripcion, calificacion, vencimiento, estado, tecnico, estado_material])

        # Crear tabla (anchos: Descripción, Calificación, Vencimiento, Estado, Técnico, Estado Material)
        tabla = Table(data, colWidths=[13*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])

        # Estilo de la tabla
        estilo = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Centrar todo excepto descripción
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),     # Descripción alineada a la izquierda
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ])

        # Aplicar colores según calificación
        for i, row in enumerate(data[1:], start=1):
            calificacion = row[1]
            if calificacion == 'DMG':  # Muy Grave
                estilo.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#ff4444'))
                estilo.add('TEXTCOLOR', (1, i), (1, i), colors.white)
            elif calificacion == 'DG':  # Grave
                estilo.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#ff9900'))
                estilo.add('TEXTCOLOR', (1, i), (1, i), colors.white)
            elif calificacion == 'DL':  # Leve
                estilo.add('BACKGROUND', (1, i), (1, i), colors.HexColor('#ffdd44'))

        tabla.setStyle(estilo)
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.5*cm))

    # Si no hay defectos
    if not defectos_por_maquina:
        mensaje = Paragraph("<i>No hay defectos pendientes con los filtros aplicados.</i>", styles['Normal'])
        elementos.append(mensaje)

    # Construir PDF
    doc.build(elementos)
    pdf_buffer.seek(0)

    # Generar nombre de archivo
    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
    nombre_archivo = f"defectos_pendientes_{fecha_archivo}.pdf"

    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=nombre_archivo
    )

# Nueva Inspección - Formulario
@app.route("/inspecciones/nueva", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def nueva_inspeccion():
    """Crear una nueva inspección"""

    if request.method == "POST":
        # Recoger datos del formulario
        fecha_inspeccion_str = request.form.get("fecha_inspeccion")

        # Calcular fecha de segunda inspección (6 meses después)
        fecha_segunda_inspeccion = None
        if fecha_inspeccion_str:
            try:
                fecha_insp = datetime.strptime(fecha_inspeccion_str, '%Y-%m-%d')
                # Sumar 6 meses
                mes_segunda = fecha_insp.month + 6
                anio_segunda = fecha_insp.year + (mes_segunda - 1) // 12
                mes_segunda = ((mes_segunda - 1) % 12) + 1
                fecha_segunda = fecha_insp.replace(year=anio_segunda, month=mes_segunda)
                fecha_segunda_inspeccion = fecha_segunda.strftime('%Y-%m-%d')
            except:
                pass

        data = {
            # Campos principales
            "maquina": request.form.get("maquina"),
            "fecha_inspeccion": fecha_inspeccion_str,
            "fecha_segunda_inspeccion": fecha_segunda_inspeccion,
            "presupuesto": request.form.get("presupuesto") or "PENDIENTE",

            # OCA
            "oca_id": int(request.form.get("oca_id")) if request.form.get("oca_id") else None,

            # Usuario que crea
            "created_by": session.get("usuario")
        }

        # Validaciones mínimas
        if not data["maquina"] or not data["fecha_inspeccion"]:
            flash("Los campos Máquina y Fecha de Inspección son obligatorios", "error")
            return redirect(request.referrer)

        # Crear inspección
        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/inspecciones?select=id",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 201]:
            nueva_id = response.json()[0]["id"]
            flash("Inspección registrada correctamente", "success")
            return redirect(f"/inspecciones/ver/{nueva_id}")
        else:
            flash(f"Error al crear inspección: {response.text}", "error")
            return redirect(request.referrer)

    # GET - Mostrar formulario
    # Obtener lista de OCAs
    response_ocas = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?select=id,nombre&activo=eq.true&order=nombre.asc",
        headers=HEADERS
    )
    ocas = []
    if response_ocas.status_code == 200:
        ocas = response_ocas.json()

    return render_template("nueva_inspeccion.html", ocas=ocas)

# Ver Detalle de Inspección
@app.route("/inspecciones/ver/<int:inspeccion_id>")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'read')
def ver_inspeccion(inspeccion_id):
    """Ver detalle completo de una inspección"""

    # Obtener inspección con información del OCA
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=*,oca:ocas(id,nombre)",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Inspección no encontrada", "error")
        return redirect("/inspecciones")

    inspeccion = response.json()[0]

    # Obtener defectos de esta inspección
    response_defectos = requests.get(
        f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?inspeccion_id=eq.{inspeccion_id}&order=calificacion.desc,fecha_limite.asc",
        headers=HEADERS
    )

    defectos = []
    if response_defectos.status_code == 200:
        defectos = response_defectos.json()

    # Calcular urgencia de cada defecto
    hoy = date.today()
    for defecto in defectos:
        if defecto.get('fecha_limite') and defecto.get('estado') == 'PENDIENTE':
            try:
                fecha_limite = datetime.strptime(defecto['fecha_limite'].split('T')[0], '%Y-%m-%d').date()
                dias_restantes = (fecha_limite - hoy).days
                defecto['dias_restantes'] = dias_restantes

                if dias_restantes < 0:
                    defecto['urgencia'] = 'VENCIDO'
                elif dias_restantes <= 15:
                    defecto['urgencia'] = 'URGENTE'
                elif dias_restantes <= 30:
                    defecto['urgencia'] = 'PROXIMO'
                else:
                    defecto['urgencia'] = 'NORMAL'
            except:
                defecto['urgencia'] = 'NORMAL'
        else:
            defecto['urgencia'] = 'COMPLETADO'

    # Calcular días hasta segunda inspección
    if inspeccion.get('fecha_segunda_inspeccion'):
        try:
            fecha_segunda = datetime.strptime(inspeccion['fecha_segunda_inspeccion'].split('T')[0], '%Y-%m-%d').date()
            dias_hasta_segunda = (fecha_segunda - hoy).days
            inspeccion['dias_hasta_segunda'] = dias_hasta_segunda
        except:
            pass

    return render_template(
        "ver_inspeccion.html",
        inspeccion=inspeccion,
        defectos=defectos
    )

# Editar Inspección
@app.route("/inspecciones/editar/<int:inspeccion_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def editar_inspeccion(inspeccion_id):
    """Editar una inspección existente"""

    if request.method == "POST":
        # Recoger datos del formulario
        fecha_inspeccion_str = request.form.get("fecha_inspeccion")

        # Calcular fecha de segunda inspección (6 meses después)
        fecha_segunda_inspeccion = None
        if fecha_inspeccion_str:
            try:
                fecha_insp = datetime.strptime(fecha_inspeccion_str, '%Y-%m-%d')
                # Sumar 6 meses
                mes_segunda = fecha_insp.month + 6
                anio_segunda = fecha_insp.year + (mes_segunda - 1) // 12
                mes_segunda = ((mes_segunda - 1) % 12) + 1
                fecha_segunda = fecha_insp.replace(year=anio_segunda, month=mes_segunda)
                fecha_segunda_inspeccion = fecha_segunda.strftime('%Y-%m-%d')
            except:
                pass

        data = {
            "maquina": request.form.get("maquina"),
            "fecha_inspeccion": fecha_inspeccion_str,
            "fecha_segunda_inspeccion": fecha_segunda_inspeccion,
            "presupuesto": request.form.get("presupuesto") or "PENDIENTE",
            "oca_id": int(request.form.get("oca_id")) if request.form.get("oca_id") else None
        }

        # Actualizar inspección
        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 204]:
            flash("Inspección actualizada correctamente", "success")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")
        else:
            flash(f"Error al actualizar inspección: {response.text}", "error")
            return redirect(request.referrer)

    # GET - Cargar datos para editar
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=*",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Inspección no encontrada", "error")
        return redirect("/inspecciones")

    inspeccion = response.json()[0]

    # Limpiar None para formulario
    def limpiar_none(data):
        if isinstance(data, dict):
            return {k: (v if v is not None else '') for k, v in data.items()}
        return data

    inspeccion = limpiar_none(inspeccion)

    # Obtener lista de OCAs
    response_ocas = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?select=id,nombre&activo=eq.true&order=nombre.asc",
        headers=HEADERS
    )
    ocas = []
    if response_ocas.status_code == 200:
        ocas = response_ocas.json()

    return render_template("editar_inspeccion.html", inspeccion=inspeccion, ocas=ocas)

# Cambiar Estado de Presupuesto
@app.route("/inspecciones/estado_presupuesto/<int:inspeccion_id>", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def cambiar_estado_presupuesto(inspeccion_id):
    """Cambiar el estado del presupuesto de una inspección"""

    nuevo_estado = request.form.get("presupuesto")

    data = {
        "presupuesto": nuevo_estado
    }

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
        json=data,
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash(f"Presupuesto cambiado a {nuevo_estado}", "success")
    else:
        flash("Error al cambiar estado", "error")

    return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Marcar Segunda Inspección como Realizada
@app.route("/inspecciones/marcar_segunda_realizada/<int:inspeccion_id>", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def marcar_segunda_realizada(inspeccion_id):
    """Marcar que se realizó la segunda inspección (a los 6 meses)"""

    # Establecer fecha de hoy como fecha_segunda_realizada
    hoy = date.today().strftime('%Y-%m-%d')

    data = {
        "fecha_segunda_realizada": hoy
    }

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
        json=data,
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Segunda inspección marcada como realizada", "success")
    else:
        flash("Error al marcar segunda inspección", "error")

    return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Subir PDF de Acta
@app.route("/inspecciones/<int:inspeccion_id>/subir_acta", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def subir_acta_pdf(inspeccion_id):
    """Subir PDF del acta de inspección a Supabase Storage"""

    if 'acta_pdf' not in request.files:
        flash("No se seleccionó ningún archivo", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    file = request.files['acta_pdf']

    if file.filename == '':
        flash("No se seleccionó ningún archivo", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    # Validar que sea PDF
    if not file.filename.lower().endswith('.pdf'):
        flash("Solo se permiten archivos PDF", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    try:
        # Generar nombre de archivo único
        file_name = f"inspeccion_{inspeccion_id}_acta.pdf"
        file_path = f"inspecciones/{file_name}"

        # Leer contenido del archivo
        file_content = file.read()

        # Headers para operaciones de storage con Content-Type
        storage_headers = {
            **STORAGE_HEADERS,
            "Content-Type": "application/pdf",
        }

        # Primero intentar eliminar el archivo existente (si existe)
        requests.delete(
            f"{SUPABASE_URL}/storage/v1/object/inspecciones-pdfs/{file_path}",
            headers=storage_headers
        )

        # Subir nuevo archivo
        upload_response = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/inspecciones-pdfs/{file_path}",
            data=file_content,
            headers=storage_headers
        )

        if upload_response.status_code not in [200, 201]:
            flash(f"Error al subir archivo: {upload_response.text}", "error")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")

        # Obtener URL pública del archivo
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/inspecciones-pdfs/{file_path}"

        # Actualizar base de datos con la URL
        data = {"acta_pdf_url": public_url}

        db_response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
            json=data,
            headers=HEADERS
        )

        if db_response.status_code in [200, 204]:
            flash("Acta PDF subida correctamente", "success")
        else:
            error_detail = db_response.text
            flash(f"Archivo subido pero error al guardar en base de datos: {error_detail}", "error")

    except Exception as e:
        flash(f"Error al procesar archivo: {str(e)}", "error")

    return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Subir PDF de Presupuesto
@app.route("/inspecciones/<int:inspeccion_id>/subir_presupuesto", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def subir_presupuesto_pdf(inspeccion_id):
    """Subir PDF del presupuesto de inspección a Supabase Storage"""

    if 'presupuesto_pdf' not in request.files:
        flash("No se seleccionó ningún archivo", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    file = request.files['presupuesto_pdf']

    if file.filename == '':
        flash("No se seleccionó ningún archivo", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    # Validar que sea PDF
    if not file.filename.lower().endswith('.pdf'):
        flash("Solo se permiten archivos PDF", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    try:
        # Generar nombre de archivo único
        file_name = f"inspeccion_{inspeccion_id}_presupuesto.pdf"
        file_path = f"inspecciones/{file_name}"

        # Leer contenido del archivo
        file_content = file.read()

        # Headers para operaciones de storage con Content-Type
        storage_headers = {
            **STORAGE_HEADERS,
            "Content-Type": "application/pdf",
        }

        # Primero intentar eliminar el archivo existente (si existe)
        requests.delete(
            f"{SUPABASE_URL}/storage/v1/object/inspecciones-pdfs/{file_path}",
            headers=storage_headers
        )

        # Subir nuevo archivo
        upload_response = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/inspecciones-pdfs/{file_path}",
            data=file_content,
            headers=storage_headers
        )

        if upload_response.status_code not in [200, 201]:
            flash(f"Error al subir archivo: {upload_response.text}", "error")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")

        # Obtener URL pública del archivo
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/inspecciones-pdfs/{file_path}"

        # Actualizar base de datos con la URL
        data = {"presupuesto_pdf_url": public_url}

        db_response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
            json=data,
            headers=HEADERS
        )

        if db_response.status_code in [200, 204]:
            flash("Presupuesto PDF subido correctamente", "success")
        else:
            error_detail = db_response.text
            flash(f"Archivo subido pero error al guardar en base de datos: {error_detail}", "error")

    except Exception as e:
        flash(f"Error al procesar archivo: {str(e)}", "error")

    return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Eliminar Inspección
@app.route("/inspecciones/eliminar/<int:inspeccion_id>")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'delete')
def eliminar_inspeccion(inspeccion_id):
    """Eliminar una inspección (y sus defectos en cascada)"""

    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}",
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Inspección eliminada correctamente", "success")
        return redirect("/inspecciones")
    else:
        flash("Error al eliminar inspección", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Función auxiliar para extraer descripciones de PDF de presupuesto
def extraer_descripciones_pdf(pdf_content):
    """
    Extrae las descripciones de la tabla del PDF de presupuesto FEDES.
    Retorna una lista de diccionarios con: codigo, descripcion
    """
    descripciones = []

    try:
        # Crear objeto PDF desde bytes
        pdf_file = io.BytesIO(pdf_content)

        with pdfplumber.open(pdf_file) as pdf:
            for page_num, page in enumerate(pdf.pages):
                logger.info(f" Procesando página {page_num + 1}")

                # Estrategia 1: Extraer tablas con configuración específica
                tables = page.extract_tables(table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                })

                if tables:
                    logger.info(f" Encontradas {len(tables)} tablas en página {page_num + 1}")

                    for table_idx, table in enumerate(tables):
                        logger.info(f" Tabla {table_idx + 1} tiene {len(table)} filas")

                        # Buscar índice de columnas Código y Descripción
                        codigo_idx = None
                        descripcion_idx = None
                        data_start = 0

                        for i, row in enumerate(table):
                            if not row:
                                continue

                            # Buscar header
                            for j, cell in enumerate(row):
                                cell_str = str(cell).strip().lower() if cell else ""
                                if 'cód' in cell_str or 'codigo' in cell_str:
                                    codigo_idx = j
                                if 'descripción' in cell_str or 'descripcion' in cell_str:
                                    descripcion_idx = j

                            if codigo_idx is not None and descripcion_idx is not None:
                                data_start = i + 1
                                logger.info(f" Header encontrado en fila {i}, Código col={codigo_idx}, Descripción col={descripcion_idx}")
                                break

                        # Si no encontramos header, asumir primeras 2 columnas
                        if codigo_idx is None or descripcion_idx is None:
                            logger.info(" Header no encontrado, asumiendo columnas 0 y 1")
                            codigo_idx = 0
                            descripcion_idx = 1
                            data_start = 1  # Saltar primera fila que probablemente sea header

                        # Extraer datos
                        for row in table[data_start:]:
                            if not row or len(row) <= max(codigo_idx, descripcion_idx):
                                continue

                            codigo = str(row[codigo_idx]).strip() if row[codigo_idx] else ""
                            descripcion_raw = str(row[descripcion_idx]).strip() if row[descripcion_idx] else ""

                            # Limpiar descripción
                            # Eliminar líneas que empiezan con "ORDEN:" pero mantener el resto
                            lineas = descripcion_raw.split('\n')
                            descripcion_limpia = []
                            for linea in lineas:
                                linea = linea.strip()
                                if not linea.startswith('ORDEN:') and linea:
                                    descripcion_limpia.append(linea)

                            descripcion = ' '.join(descripcion_limpia)

                            # ===== LIMPIEZA AGRESIVA DE NÚMEROS =====
                            import re

                            logger.info(f" Antes de limpiar: {descripcion[:100]}")

                            # Estrategia 1: Eliminar TODO después del primer número (con o sin decimales)
                            # Busca patrones como: "1,00" o "100" o "1.300,00"
                            match = re.search(r'\s+\d+[,\.]?\d*', descripcion)
                            if match:
                                descripcion = descripcion[:match.start()].strip()
                                logger.info(f" Después de cortar: {descripcion[:100]}")

                            # Estrategia 2: Si aún quedan números al final, eliminarlos
                            descripcion = re.sub(r'\s+[\d\.,\s]+$', '', descripcion)
                            descripcion = descripcion.strip()

                            logger.info(f" Final limpio: {descripcion[:100]}")

                            # ===== VALIDACIONES ESTRICTAS =====
                            # 1. Longitud mínima
                            if not descripcion or len(descripcion) < 10:
                                continue

                            # 2. Filtrar headers repetidos
                            if any(x in descripcion for x in ['Descripción', 'PRESUPUESTO', 'Cód.', 'Cant.', 'Precio', 'Total', '% Igic', 'INSTALACIÓN']):
                                continue

                            # 3. Filtrar direcciones (empiezan con C/, Calle, Avenida, etc.)
                            if descripcion.startswith(('C/', 'Calle', 'Avenida', 'Avda', 'c/', 'calle', 'C /')):
                                continue

                            # 4. Filtrar líneas que son mayormente números (totales, impuestos)
                            palabras = descripcion.split()
                            if len(palabras) <= 4:  # Líneas cortas tipo "7% 339,50 5.189,47"
                                numeros_count = sum(1 for p in palabras if any(c.isdigit() for c in p) or '%' in p)
                                if numeros_count >= len(palabras) * 0.6:  # 60%+ son números
                                    continue

                            # 5. Volver a validar longitud después de limpiar
                            if len(descripcion) < 10:
                                continue

                            # 6. La descripción debe contener letras (no solo números/símbolos)
                            if not any(c.isalpha() for c in descripcion):
                                continue

                            # 7. Filtrar líneas con palabras clave que NO son defectos
                            palabras_invalidas = ['NIF', 'CIF', 'Tel.', 'Tel', 'Email', '@', 'www', 'http',
                                                 'GRAN CANARIA', 'LAS PALMAS', 'Municipio', 'Serie']
                            if any(palabra in descripcion for palabra in palabras_invalidas):
                                continue

                            # 8. Debe tener al menos 3 palabras (descripciones reales son más largas)
                            if len(descripcion.split()) < 3:
                                continue

                            # 9. Filtrar si toda la descripción es MAYÚSCULAS cortas (probablemente headers/títulos)
                            if descripcion.isupper() and len(descripcion) < 30:
                                continue

                            # NOTA: No validamos el código ya que no es necesario
                            # Guardamos código si existe, o vacío si no
                            logger.info(f" ✓ Añadiendo: {descripcion[:70]}...")
                            descripciones.append({
                                'codigo': codigo if codigo else "",
                                'descripcion': descripcion
                            })

                # Estrategia 2: Si no encontró tablas, intentar extracción de texto
                if not tables:
                    logger.info(" No se encontraron tablas, intentando extracción de texto")
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            line = line.strip()

                            if not line or len(line) < 10:
                                continue

                            # Buscar líneas que empiezan con código numérico (8-10 dígitos)
                            parts = line.split(None, 1)
                            if len(parts) < 2:
                                continue

                            possible_code = parts[0]
                            descripcion = parts[1]

                            # Código debe ser numérico de 8-11 dígitos
                            if not (possible_code.isdigit() and 8 <= len(possible_code) <= 11):
                                continue

                            # Aplicar MISMA LIMPIEZA que Estrategia 1
                            import re

                            logger.info(f" Antes de limpiar: {descripcion[:100]}")

                            # Cortar en el primer número decimal
                            match = re.search(r'\s+\d+[,\.]?\d*', descripcion)
                            if match:
                                descripcion = descripcion[:match.start()].strip()
                                logger.info(f" Después de cortar: {descripcion[:100]}")

                            # Eliminar números al final
                            descripcion = re.sub(r'\s+[\d\.,\s]+$', '', descripcion)
                            descripcion = descripcion.strip()

                            # APLICAR MISMAS VALIDACIONES que Estrategia 1
                            if len(descripcion) < 10:
                                continue

                            if any(x in descripcion for x in ['Descripción', 'Cód.', 'Cant.', 'Precio', 'Total', 'INSTALACIÓN']):
                                continue

                            if descripcion.startswith(('C/', 'Calle', 'Avda')):
                                continue

                            if not any(c.isalpha() for c in descripcion):
                                continue

                            if len(descripcion.split()) < 3:
                                continue

                            logger.info(f" ✓ Añadiendo: {descripcion[:70]}...")
                            descripciones.append({
                                'codigo': possible_code,
                                'descripcion': descripcion
                            })

        logger.info(f" Total descripciones extraídas: {len(descripciones)}")

    except Exception as e:
        logger.error(f"Error al extraer descripciones del PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

    return descripciones

# Extraer defectos de PDF de presupuesto
@app.route("/inspecciones/<int:inspeccion_id>/extraer_defectos_pdf", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def extraer_defectos_pdf(inspeccion_id):
    """Extraer defectos del PDF de presupuesto y mostrar preview para clasificación"""

    # Obtener información de la inspección
    response_insp = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=*",
        headers=HEADERS
    )

    if response_insp.status_code != 200 or not response_insp.json():
        flash("Inspección no encontrada", "error")
        return redirect("/inspecciones")

    inspeccion = response_insp.json()[0]

    # Verificar que existe un PDF de presupuesto
    presupuesto_pdf_url = inspeccion.get('presupuesto_pdf_url')

    if not presupuesto_pdf_url:
        flash("No hay PDF de presupuesto subido. Por favor, sube el PDF primero.", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    try:
        # Descargar el PDF desde Supabase Storage
        pdf_response = requests.get(presupuesto_pdf_url)

        if pdf_response.status_code != 200:
            flash("Error al descargar el PDF de presupuesto", "error")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")

        # Extraer descripciones del PDF
        descripciones = extraer_descripciones_pdf(pdf_response.content)

        if not descripciones:
            flash("No se encontraron descripciones en el PDF. Verifica el formato del archivo.", "warning")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")

        # Guardar descripciones en la sesión para el siguiente paso
        session[f'defectos_extraidos_{inspeccion_id}'] = descripciones

        # Renderizar template de preview
        return render_template(
            "importar_defectos_preview.html",
            inspeccion=inspeccion,
            descripciones=descripciones
        )

    except Exception as e:
        flash(f"Error al procesar PDF: {str(e)}", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

# Guardar defectos importados desde PDF
@app.route("/inspecciones/<int:inspeccion_id>/guardar_defectos_importados", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def guardar_defectos_importados(inspeccion_id):
    """Guardar defectos clasificados manualmente después de la extracción del PDF"""

    logger.info(f" Guardando defectos para inspección {inspeccion_id}")

    # Recuperar descripciones de la sesión
    descripciones = session.get(f'defectos_extraidos_{inspeccion_id}')

    logger.info(f" Descripciones en sesión: {len(descripciones) if descripciones else 'None'}")

    if not descripciones:
        flash("No hay defectos pendientes de importar", "error")
        return redirect(f"/inspecciones/ver/{inspeccion_id}")

    # Obtener fecha de inspección para calcular fechas límite
    response_insp = requests.get(
        f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=*",
        headers=HEADERS
    )

    logger.info(f" Respuesta BD: status={response_insp.status_code}")

    if response_insp.status_code != 200:
        logger.error(f" Error BD: {response_insp.text}")
        flash("Error al consultar inspección", "error")
        return redirect("/inspecciones")

    if not response_insp.json():
        flash("Inspección no encontrada", "error")
        logger.error(f" Inspección {inspeccion_id} no tiene datos")
        return redirect("/inspecciones")

    insp_data = response_insp.json()[0]
    fecha_inspeccion = insp_data.get('fecha_inspeccion')

    # Procesar formulario
    defectos_guardados = 0
    defectos_omitidos = 0

    for i, descripcion_data in enumerate(descripciones):
        # Verificar si este defecto fue seleccionado (checkbox)
        seleccionado = request.form.get(f'seleccionar_{i}') == 'on'

        if not seleccionado:
            defectos_omitidos += 1
            continue

        # Obtener clasificación del usuario
        calificacion = request.form.get(f'calificacion_{i}')
        plazo_meses = int(request.form.get(f'plazo_{i}', 6))
        es_cortina = request.form.get(f'es_cortina_{i}') == 'on'
        es_pesacarga = request.form.get(f'es_pesacarga_{i}') == 'on'

        # Validar que tenga calificación
        if not calificacion:
            continue

        # Calcular fecha límite
        fecha_limite = None
        if fecha_inspeccion:
            try:
                fecha_insp_dt = datetime.strptime(fecha_inspeccion.split('T')[0], '%Y-%m-%d')
                mes_limite = fecha_insp_dt.month + plazo_meses
                anio_limite = fecha_insp_dt.year + (mes_limite - 1) // 12
                mes_limite = ((mes_limite - 1) % 12) + 1
                fecha_limite_dt = fecha_insp_dt.replace(year=anio_limite, month=mes_limite)
                fecha_limite = fecha_limite_dt.strftime('%Y-%m-%d')
            except:
                pass

        # Crear defecto
        defecto_data = {
            "inspeccion_id": inspeccion_id,
            "descripcion": descripcion_data.get('descripcion'),
            "calificacion": calificacion,
            "plazo_meses": plazo_meses,
            "fecha_limite": fecha_limite,
            "estado": "PENDIENTE",
            "es_cortina": es_cortina,
            "es_pesacarga": es_pesacarga,
            "observaciones": f"Importado desde PDF de presupuesto"
        }

        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/defectos_inspeccion",
            json=defecto_data,
            headers=HEADERS
        )

        if response.status_code in [200, 201]:
            defectos_guardados += 1

    # Limpiar sesión
    session.pop(f'defectos_extraidos_{inspeccion_id}', None)

    # Mostrar resultado
    if defectos_guardados > 0:
        flash(f"Se importaron {defectos_guardados} defecto(s) correctamente", "success")

    if defectos_omitidos > 0:
        flash(f"Se omitieron {defectos_omitidos} defecto(s)", "info")

    return redirect(f"/inspecciones/ver/{inspeccion_id}")

# ============================================
# DEFECTOS DE INSPECCIÓN
# ============================================

# Añadir Defecto a Inspección
@app.route("/inspecciones/<int:inspeccion_id>/defectos/nuevo", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def nuevo_defecto(inspeccion_id):
    """Añadir un nuevo defecto a una inspección"""

    if request.method == "POST":
        # Obtener fecha de inspección para calcular fecha límite
        response_insp = requests.get(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=fecha_inspeccion",
            headers=HEADERS
        )

        if response_insp.status_code != 200 or not response_insp.json():
            flash("Inspección no encontrada", "error")
            return redirect("/inspecciones")

        fecha_inspeccion = response_insp.json()[0].get('fecha_inspeccion')
        plazo_meses = int(request.form.get("plazo_meses", 6))

        # Calcular fecha límite
        fecha_limite = None
        if fecha_inspeccion:
            try:
                fecha_insp_dt = datetime.strptime(fecha_inspeccion.split('T')[0], '%Y-%m-%d')
                # Sumar plazo en meses
                mes_limite = fecha_insp_dt.month + plazo_meses
                anio_limite = fecha_insp_dt.year + (mes_limite - 1) // 12
                mes_limite = ((mes_limite - 1) % 12) + 1

                fecha_limite_dt = fecha_insp_dt.replace(year=anio_limite, month=mes_limite)
                fecha_limite = fecha_limite_dt.strftime('%Y-%m-%d')
            except:
                flash("Error al calcular fecha límite", "error")
                return redirect(request.referrer)

        # Crear defecto
        data = {
            "inspeccion_id": inspeccion_id,
            "descripcion": request.form.get("descripcion"),
            "calificacion": request.form.get("calificacion"),
            "plazo_meses": plazo_meses,
            "fecha_limite": fecha_limite,
            "estado": "PENDIENTE",
            "es_cortina": request.form.get("es_cortina") == "true",
            "es_pesacarga": request.form.get("es_pesacarga") == "true",
            "observaciones": request.form.get("observaciones") or None
        }

        # Validar
        if not data["descripcion"] or not data["calificacion"]:
            flash("Descripción y Calificación son obligatorios", "error")
            return redirect(request.referrer)

        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/defectos_inspeccion",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 201]:
            flash("Defecto añadido correctamente", "success")
            return redirect(f"/inspecciones/ver/{inspeccion_id}")
        else:
            flash(f"Error al añadir defecto: {response.text}", "error")
            return redirect(request.referrer)

    # GET - Mostrar formulario
    return render_template("nuevo_defecto.html", inspeccion_id=inspeccion_id)

# Marcar Defecto como Subsanado
@app.route("/defectos/<int:defecto_id>/subsanar", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def subsanar_defecto(defecto_id):
    """Marcar un defecto como subsanado"""

    data = {
        "estado": "SUBSANADO",
        "fecha_subsanacion": date.today().isoformat()
    }

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}",
        json=data,
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Defecto marcado como subsanado", "success")
    else:
        flash("Error al actualizar defecto", "error")

    return redirect(request.referrer)

# Eliminar Defecto
@app.route("/defectos/<int:defecto_id>/eliminar")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'delete')
def eliminar_defecto(defecto_id):
    """Eliminar un defecto"""

    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}",
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Defecto eliminado correctamente", "success")
    else:
        flash("Error al eliminar defecto", "error")

    return redirect(request.referrer)

# Ver detalle de un defecto
@app.route("/defectos/<int:defecto_id>")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'read')
def ver_defecto(defecto_id):
    """Ver detalle completo de un defecto"""

    # Obtener el defecto con información de inspección
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}&select=*",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Defecto no encontrado", "error")
        return redirect("/defectos_dashboard")

    defecto = response.json()[0]

    # Obtener información de la inspección asociada
    if defecto.get('inspeccion_id'):
        response_insp = requests.get(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{defecto['inspeccion_id']}&select=*",
            headers=HEADERS
        )

        if response_insp.status_code == 200 and response_insp.json():
            inspeccion = response_insp.json()[0]

            # Obtener información del OCA si existe
            if inspeccion.get('oca_id'):
                response_oca = requests.get(
                    f"{SUPABASE_URL}/rest/v1/ocas?id=eq.{inspeccion['oca_id']}&select=nombre",
                    headers=HEADERS
                )
                if response_oca.status_code == 200 and response_oca.json():
                    inspeccion['oca_nombre'] = response_oca.json()[0].get('nombre')
                else:
                    inspeccion['oca_nombre'] = None
            else:
                inspeccion['oca_nombre'] = None

            defecto['inspeccion'] = inspeccion
        else:
            defecto['inspeccion'] = None
    else:
        defecto['inspeccion'] = None

    # Calcular días restantes
    if defecto.get('fecha_limite') and defecto.get('estado') == 'PENDIENTE':
        try:
            fecha_limite = datetime.strptime(defecto['fecha_limite'].split('T')[0], '%Y-%m-%d').date()
            hoy = date.today()
            dias_restantes = (fecha_limite - hoy).days
            defecto['dias_restantes'] = dias_restantes

            if dias_restantes < 0:
                defecto['nivel_urgencia'] = 'VENCIDO'
            elif dias_restantes <= 15:
                defecto['nivel_urgencia'] = 'URGENTE'
            elif dias_restantes <= 30:
                defecto['nivel_urgencia'] = 'PROXIMO'
            else:
                defecto['nivel_urgencia'] = 'NORMAL'
        except:
            defecto['nivel_urgencia'] = 'NORMAL'
    else:
        defecto['nivel_urgencia'] = 'COMPLETADO'

    return render_template(
        "ver_defecto.html",
        defecto=defecto
    )

# Editar defecto
@app.route("/defectos/<int:defecto_id>/editar", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def editar_defecto(defecto_id):
    """Editar un defecto existente"""

    if request.method == "POST":
        # Obtener datos del formulario
        descripcion = request.form.get("descripcion")
        calificacion = request.form.get("calificacion")
        plazo_meses = request.form.get("plazo_meses", type=int)
        fecha_limite_str = request.form.get("fecha_limite")
        estado = request.form.get("estado")
        fecha_subsanacion_str = request.form.get("fecha_subsanacion")
        es_cortina = request.form.get("es_cortina") == "on"
        es_pesacarga = request.form.get("es_pesacarga") == "on"
        observaciones = request.form.get("observaciones")

        # Nuevos campos de gestión operativa
        tecnico_asignado = request.form.get("tecnico_asignado") or None
        gestion_material = request.form.get("gestion_material") or None
        estado_stock = request.form.get("estado_stock") or None

        # Validaciones
        if not descripcion or not calificacion:
            flash("Descripción y calificación son obligatorios", "error")
            return redirect(f"/defectos/{defecto_id}/editar")

        # Obtener el defecto actual para verificar si cambió el plazo
        response_defecto = requests.get(
            f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}&select=plazo_meses,inspeccion_id",
            headers=HEADERS
        )

        if response_defecto.status_code == 200 and response_defecto.json():
            defecto_actual = response_defecto.json()[0]
            plazo_anterior = defecto_actual.get('plazo_meses')
            inspeccion_id = defecto_actual.get('inspeccion_id')

            # Si cambió el plazo, recalcular la fecha límite
            if plazo_meses != plazo_anterior and inspeccion_id:
                # Obtener fecha de inspección
                response_insp = requests.get(
                    f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{inspeccion_id}&select=fecha_inspeccion",
                    headers=HEADERS
                )

                if response_insp.status_code == 200 and response_insp.json():
                    fecha_inspeccion = response_insp.json()[0].get('fecha_inspeccion')

                    if fecha_inspeccion:
                        try:
                            fecha_insp_dt = datetime.strptime(fecha_inspeccion.split('T')[0], '%Y-%m-%d')
                            # Sumar plazo en meses
                            mes_limite = fecha_insp_dt.month + plazo_meses
                            anio_limite = fecha_insp_dt.year + (mes_limite - 1) // 12
                            mes_limite = ((mes_limite - 1) % 12) + 1

                            fecha_limite_dt = fecha_insp_dt.replace(year=anio_limite, month=mes_limite)
                            fecha_limite_str = fecha_limite_dt.strftime('%Y-%m-%d')
                        except Exception as e:
                            flash(f"Error al calcular fecha límite: {str(e)}", "error")

        # Preparar datos para actualizar
        datos_actualizacion = {
            "descripcion": descripcion,
            "calificacion": calificacion,
            "plazo_meses": plazo_meses,
            "fecha_limite": fecha_limite_str,
            "estado": estado,
            "es_cortina": es_cortina,
            "es_pesacarga": es_pesacarga,
            "observaciones": observaciones,
            "tecnico_asignado": tecnico_asignado,
            "gestion_material": gestion_material,
            "estado_stock": estado_stock
        }

        # Si el estado es subsanado y hay fecha, incluirla
        if estado == "SUBSANADO" and fecha_subsanacion_str:
            datos_actualizacion["fecha_subsanacion"] = fecha_subsanacion_str
        elif estado == "PENDIENTE":
            datos_actualizacion["fecha_subsanacion"] = None

        # Actualizar en la base de datos
        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}",
            headers=HEADERS,
            json=datos_actualizacion
        )

        if response.status_code in [200, 204]:
            flash("Defecto actualizado correctamente", "success")
            return redirect(f"/defectos/{defecto_id}")
        else:
            flash(f"Error al actualizar defecto: {response.text}", "error")
            return redirect(f"/defectos/{defecto_id}/editar")

    # GET: Mostrar formulario de edición
    # Obtener el defecto
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}&select=*",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Defecto no encontrado", "error")
        return redirect("/defectos_dashboard")

    defecto = response.json()[0]

    # Obtener información de la inspección asociada
    if defecto.get('inspeccion_id'):
        response_insp = requests.get(
            f"{SUPABASE_URL}/rest/v1/inspecciones?id=eq.{defecto['inspeccion_id']}&select=id,rae,maquina,direccion,fecha_inspeccion",
            headers=HEADERS
        )

        if response_insp.status_code == 200 and response_insp.json():
            defecto['inspeccion'] = response_insp.json()[0]
        else:
            defecto['inspeccion'] = {'maquina': 'N/A', 'direccion': 'N/A'}
    else:
        defecto['inspeccion'] = {'maquina': 'N/A', 'direccion': 'N/A'}

    return render_template(
        "editar_defecto.html",
        defecto=defecto
    )


# Endpoint para actualización rápida de gestión operativa (AJAX)
@app.route("/defectos/<int:defecto_id>/actualizar_gestion", methods=["POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def actualizar_gestion_defecto(defecto_id):
    """Endpoint para actualización rápida de campos de gestión operativa desde el dashboard"""
    try:
        data = request.json
        campo = data.get("campo")
        valor = data.get("valor")

        if not campo:
            return {"error": "Campo requerido"}, 400

        # Mapear campos del frontend a campos de base de datos
        campos_map = {
            "tecnico": "tecnico_asignado",
            "material": "gestion_material",
            "stock": "estado_stock"
        }

        if campo not in campos_map:
            return {"error": "Campo no válido"}, 400

        campo_db = campos_map[campo]

        # Preparar datos para actualizar
        datos_actualizacion = {
            campo_db: valor if valor else None
        }

        # Actualizar en la base de datos
        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/defectos_inspeccion?id=eq.{defecto_id}",
            headers=HEADERS,
            json=datos_actualizacion
        )

        if response.status_code in [200, 204]:
            return {"success": True, "campo": campo, "valor": valor}, 200
        else:
            return {"error": f"Error al actualizar: {response.text}"}, 500

    except Exception as e:
        return {"error": str(e)}, 500


# ============================================
# ============================================
# GESTIÓN DE OCAs (Organismos de Control)
# ============================================

# Listado de OCAs
@app.route("/ocas")
def lista_ocas():
    """Listado de todos los OCAs"""
    if "usuario" not in session:
        return redirect("/")

    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?select=*&order=nombre.asc",
        headers=HEADERS
    )

    ocas = []
    if response.status_code == 200:
        ocas = response.json()

    # Contar inspecciones por OCA
    for oca in ocas:
        response_count = requests.get(
            f"{SUPABASE_URL}/rest/v1/inspecciones?oca_id=eq.{oca['id']}&select=id",
            headers=HEADERS
        )
        if response_count.status_code == 200:
            oca['total_inspecciones'] = len(response_count.json())
        else:
            oca['total_inspecciones'] = 0

    return render_template("lista_ocas.html", ocas=ocas)

# Nuevo OCA
@app.route("/ocas/nuevo", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def nuevo_oca():
    """Crear un nuevo OCA"""

    if request.method == "POST":
        data = {
            "nombre": request.form.get("nombre"),
            "contacto_nombre": request.form.get("contacto_nombre") or None,
            "contacto_email": request.form.get("contacto_email") or None,
            "contacto_telefono": request.form.get("contacto_telefono") or None,
            "direccion": request.form.get("direccion") or None,
            "observaciones": request.form.get("observaciones") or None,
            "activo": True
        }

        if not data["nombre"]:
            flash("El nombre del OCA es obligatorio", "error")
            return redirect(request.referrer)

        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/ocas",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 201]:
            flash("OCA creado correctamente", "success")
            return redirect("/ocas")
        else:
            flash(f"Error al crear OCA: {response.text}", "error")
            return redirect(request.referrer)

    return render_template("nuevo_oca.html")

# Editar OCA
@app.route("/ocas/editar/<int:oca_id>", methods=["GET", "POST"])
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'write')
def editar_oca(oca_id):
    """Editar un OCA existente"""

    if request.method == "POST":
        data = {
            "nombre": request.form.get("nombre"),
            "contacto_nombre": request.form.get("contacto_nombre") or None,
            "contacto_email": request.form.get("contacto_email") or None,
            "contacto_telefono": request.form.get("contacto_telefono") or None,
            "direccion": request.form.get("direccion") or None,
            "observaciones": request.form.get("observaciones") or None,
            "activo": request.form.get("activo") == "true"
        }

        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/ocas?id=eq.{oca_id}",
            json=data,
            headers=HEADERS
        )

        if response.status_code in [200, 204]:
            flash("OCA actualizado correctamente", "success")
            return redirect("/ocas")
        else:
            flash(f"Error al actualizar OCA: {response.text}", "error")
            return redirect(request.referrer)

    # GET
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/ocas?id=eq.{oca_id}",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("OCA no encontrado", "error")
        return redirect("/ocas")

    oca = response.json()[0]

    # Limpiar None
    def limpiar_none(data):
        if isinstance(data, dict):
            return {k: (v if v is not None else '') for k, v in data.items()}
        return data

    oca = limpiar_none(oca)

    return render_template("editar_oca.html", oca=oca)

# Eliminar OCA
@app.route("/ocas/eliminar/<int:oca_id>")
@helpers.login_required
@helpers.requiere_permiso('inspecciones', 'delete')
def eliminar_oca(oca_id):
    """Eliminar un OCA"""

    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/ocas?id=eq.{oca_id}",
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("OCA eliminado correctamente", "success")
    else:
        flash("Error al eliminar OCA", "error")

    return redirect("/ocas")

# ============================================
# ADMINISTRACIÓN DE USUARIOS (Solo Admin)
# ============================================

@app.route("/admin/usuarios")
@helpers.login_required
@helpers.solo_admin
def admin_usuarios():
    """Panel de administración de usuarios - Solo para admin"""

    # Obtener todos los usuarios
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/usuarios?select=*&order=id",
        headers=HEADERS
    )

    usuarios = []
    if response.status_code == 200:
        usuarios = response.json()

    return render_template("admin_usuarios.html", usuarios=usuarios)


@app.route("/admin/usuarios/crear", methods=["POST"])
@helpers.login_required
@helpers.solo_admin
def admin_crear_usuario():
    """Crear un nuevo usuario"""

    nombre_usuario = request.form.get("nombre_usuario", "").strip()
    email = request.form.get("email", "").strip()
    contrasena = request.form.get("contrasena", "").strip()
    perfil = request.form.get("perfil", "visualizador")

    # Validaciones
    if not nombre_usuario or not contrasena:
        flash("Nombre de usuario y contraseña son obligatorios", "error")
        return redirect("/admin/usuarios")

    if perfil not in ['admin', 'gestor', 'visualizador']:
        flash("Perfil inválido", "error")
        return redirect("/admin/usuarios")

    # Crear usuario
    data = {
        "nombre_usuario": nombre_usuario,
        "email": email or None,
        "contrasena": contrasena,  # NOTA: En producción deberías usar hash
        "perfil": perfil
    }

    response = requests.post(
        f"{SUPABASE_URL}/rest/v1/usuarios",
        json=data,
        headers=HEADERS
    )

    if response.status_code in [200, 201]:
        flash(f"Usuario '{nombre_usuario}' creado exitosamente con perfil '{perfil}'", "success")
    else:
        flash(f"Error al crear usuario: {response.text}", "error")

    return redirect("/admin/usuarios")


@app.route("/admin/usuarios/editar/<int:usuario_id>", methods=["POST"])
@helpers.login_required
@helpers.solo_admin
def admin_editar_usuario(usuario_id):
    """Editar perfil de un usuario"""

    perfil = request.form.get("perfil", "visualizador")

    if perfil not in ['admin', 'gestor', 'visualizador']:
        flash("Perfil inválido", "error")
        return redirect("/admin/usuarios")

    # Actualizar perfil
    data = {"perfil": perfil}

    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/usuarios?id=eq.{usuario_id}",
        json=data,
        headers=HEADERS
    )

    if response.status_code == 200:
        flash("Perfil actualizado correctamente", "success")
    else:
        flash(f"Error al actualizar perfil: {response.text}", "error")

    return redirect("/admin/usuarios")


@app.route("/admin/usuarios/eliminar/<int:usuario_id>")
@helpers.login_required
@helpers.solo_admin
def admin_eliminar_usuario(usuario_id):
    """Eliminar un usuario"""

    # Evitar que el admin se elimine a sí mismo
    if usuario_id == session.get("usuario_id"):
        flash("No puedes eliminar tu propio usuario", "error")
        return redirect("/admin/usuarios")

    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/usuarios?id=eq.{usuario_id}",
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Usuario eliminado correctamente", "success")
    else:
        flash(f"Error al eliminar usuario: {response.text}", "error")

    return redirect("/admin/usuarios")


# ============================================
# CARTERA Y ANÁLISIS
# ============================================

@app.route("/cartera")
@helpers.login_required
def cartera_dashboard():
    """Dashboard principal de Cartera y Análisis"""

    # Obtener estadísticas generales
    stats = {}

    # Total de instalaciones
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/instalaciones?select=count",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['total_instalaciones'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # Total de máquinas
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/maquinas_cartera?select=count",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['total_maquinas'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # Total de partes
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?select=count",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['total_partes'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # Recomendaciones pendientes (no revisadas y sin oportunidad creada)
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?select=count&tiene_recomendacion=eq.true&recomendacion_revisada=eq.false&oportunidad_creada=eq.false",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['recomendaciones_pendientes'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # KPIs adicionales de análisis
    # Averías último año
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?select=count&tipo_parte_normalizado=eq.AVERIA&fecha_parte=gte.{(datetime.now() - timedelta(days=365)).isoformat()}",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['averias_anio'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # Mantenimientos último año
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?select=count&tipo_parte_normalizado=eq.MANTENIMIENTO&fecha_parte=gte.{(datetime.now() - timedelta(days=365)).isoformat()}",
        headers={**HEADERS, "Prefer": "count=exact"}
    )
    stats['mantenimientos_anio'] = response.headers.get('Content-Range', '0').split('/')[-1]

    # Top 10 máquinas problemáticas (usando la vista)
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/v_maquinas_problematicas?select=*&order=indice_problema.desc&limit=10",
        headers=HEADERS
    )
    maquinas_problematicas = response.json() if response.status_code == 200 else []

    # Recomendaciones pendientes (últimas 20)
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/v_partes_con_recomendaciones?select=*&order=fecha_parte.desc&limit=20",
        headers=HEADERS
    )
    recomendaciones = response.json() if response.status_code == 200 else []

    # Distribución de tipos de parte (último año)
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?select=tipo_parte_normalizado&fecha_parte=gte.{(datetime.now() - timedelta(days=365)).isoformat()}",
        headers=HEADERS
    )
    if response.status_code == 200:
        partes_data = response.json()
        tipos_distribucion = {}
        for parte in partes_data:
            tipo = parte.get('tipo_parte_normalizado', 'OTRO')
            tipos_distribucion[tipo] = tipos_distribucion.get(tipo, 0) + 1
    else:
        tipos_distribucion = {}

    return render_template(
        "cartera/dashboard.html",
        stats=stats,
        maquinas_problematicas=maquinas_problematicas,
        recomendaciones=recomendaciones,
        tipos_distribucion=tipos_distribucion
    )


@app.route("/cartera/importar")
@helpers.login_required
def cartera_importar():
    """Interfaz de importación de datos"""
    return render_template("cartera/importar.html")


@app.route("/cartera/importar_equipos", methods=["POST"])
@helpers.login_required
def cartera_importar_equipos():
    """Importar instalaciones y máquinas desde Excel"""

    if 'archivo_equipos' not in request.files:
        flash("No se seleccionó ningún archivo", "error")
        return redirect("/cartera/importar")

    file = request.files['archivo_equipos']

    if file.filename == '':
        flash("No se seleccionó ningún archivo", "error")
        return redirect("/cartera/importar")

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("El archivo debe ser formato Excel (.xlsx o .xls)", "error")
        return redirect("/cartera/importar")

    try:
        # Leer Excel
        df = pd.read_excel(file)

        # Log de columnas encontradas
        logger.info(f"Columnas en Excel: {list(df.columns)}")
        logger.info(f"Total de filas: {len(df)}")

        # Mapeo de columnas (flexible con mayúsculas/minúsculas y acentos)
        column_mapping = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            if 'instalación' in col_lower or 'instalacion' in col_lower:
                if 'cód' in col_lower or 'cod' in col_lower:
                    column_mapping[col] = 'cod_instalacion'
                else:
                    column_mapping[col] = 'instalacion'
            elif 'máquina' in col_lower or 'maquina' in col_lower:
                if 'cód' in col_lower or 'cod' in col_lower:
                    column_mapping[col] = 'cod_maquina'
                else:
                    column_mapping[col] = 'maquina'
            elif 'tecnico' in col_lower or 'técnico' in col_lower:
                column_mapping[col] = 'tecnico'

        df.rename(columns=column_mapping, inplace=True)
        logger.info(f"Columnas después de mapeo: {list(df.columns)}")

        # Verificar columnas requeridas
        required = ['cod_instalacion', 'instalacion', 'cod_maquina', 'maquina']
        missing = [col for col in required if col not in df.columns]

        if missing:
            logger.error(f"Faltan columnas: {missing}. Columnas después de mapeo: {list(df.columns)}")
            flash(f"Faltan columnas requeridas: {', '.join(missing)}", "error")
            flash(f"Columnas encontradas en Excel: {', '.join(df.columns)}", "info")
            return redirect("/cartera/importar")

        # Procesar instalaciones únicas
        instalaciones_unicas = df[['cod_instalacion', 'instalacion']].drop_duplicates('cod_instalacion')
        instalaciones_map = {}
        stats = {
            'instalaciones_nuevas': 0,
            'instalaciones_existentes': 0,
            'maquinas_nuevas': 0,
            'maquinas_existentes': 0,
            'errores': 0
        }

        # Municipios de Gran Canaria
        municipios_gc = [
            'LAS PALMAS', 'TELDE', 'SANTA LUCIA', 'AGÜIMES', 'INGENIO',
            'MOGAN', 'SAN BARTOLOME', 'SANTA BRIGIDA', 'ARUCAS', 'TEROR',
            'GALDAR', 'AGAETE', 'VALLESECO', 'FIRGAS', 'MOYA',
            'SANTA MARIA DE GUIA', 'VALSEQUILLO', 'VEGA DE SAN MATEO',
            'TEJEDA', 'ALDEA DE SAN NICOLAS'
        ]

        for idx, row in instalaciones_unicas.iterrows():
            # Limpiar nombre (quitar dirección después del guion)
            nombre = row['instalacion'].split(' - ')[0].strip() if ' - ' in row['instalacion'] else row['instalacion'].strip()

            # Extraer municipio
            texto_upper = row['instalacion'].upper()
            municipio = "Las Palmas de Gran Canaria"  # Default
            for mun in municipios_gc:
                if mun in texto_upper:
                    municipio = mun.title()
                    break

            # Verificar si ya existe
            response = requests.get(
                f"{SUPABASE_URL}/rest/v1/instalaciones?nombre=eq.{urllib.parse.quote(nombre)}",
                headers=HEADERS
            )

            if response.status_code == 200 and len(response.json()) > 0:
                instalacion_id = response.json()[0]['id']
                stats['instalaciones_existentes'] += 1
            else:
                # Crear nueva
                data = {"nombre": nombre, "municipio": municipio}
                response = requests.post(
                    f"{SUPABASE_URL}/rest/v1/instalaciones",
                    json=data,
                    headers=HEADERS
                )

                if response.status_code == 201:
                    instalacion_id = response.json()[0]['id']
                    stats['instalaciones_nuevas'] += 1
                else:
                    logger.error(f"Error creando instalación '{nombre}': {response.status_code} - {response.text}")
                    stats['errores'] += 1
                    continue

            instalaciones_map[row['cod_instalacion']] = instalacion_id

        # Procesar máquinas
        for idx, row in df.iterrows():
            cod_instalacion = row['cod_instalacion']
            identificador = row['maquina'].strip()
            codigo_maquina = row['cod_maquina'].strip() if pd.notna(row['cod_maquina']) else None

            instalacion_id = instalaciones_map.get(cod_instalacion)
            if not instalacion_id:
                stats['errores'] += 1
                continue

            # Verificar si la máquina ya existe
            response = requests.get(
                f"{SUPABASE_URL}/rest/v1/maquinas_cartera?identificador=eq.{urllib.parse.quote(identificador)}",
                headers=HEADERS
            )

            if response.status_code == 200 and len(response.json()) > 0:
                stats['maquinas_existentes'] += 1
                continue

            # Crear nueva máquina
            data = {
                "instalacion_id": instalacion_id,
                "identificador": identificador,
                "codigo_maquina": codigo_maquina
            }

            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/maquinas_cartera",
                json=data,
                headers=HEADERS
            )

            if response.status_code == 201:
                stats['maquinas_nuevas'] += 1
            else:
                logger.error(f"Error creando máquina '{identificador}': {response.status_code} - {response.text}")
                stats['errores'] += 1

        # Mostrar resumen
        flash(f"Importación completada: {stats['instalaciones_nuevas']} instalaciones nuevas, "
              f"{stats['maquinas_nuevas']} máquinas nuevas", "success")

        if stats['instalaciones_existentes'] > 0:
            flash(f"{stats['instalaciones_existentes']} instalaciones ya existían", "info")

        if stats['maquinas_existentes'] > 0:
            flash(f"{stats['maquinas_existentes']} máquinas ya existían", "info")

        if stats['errores'] > 0:
            flash(f"{stats['errores']} registros con errores", "warning")

    except Exception as e:
        flash(f"Error al procesar archivo: {str(e)}", "error")
        logger.error(f"Error importando equipos: {str(e)}")

    return redirect("/cartera/importar")


@app.route("/cartera/importar_partes", methods=["POST"])
@helpers.login_required
def cartera_importar_partes():
    """Importar partes de trabajo desde Excel"""

    if 'archivo_partes' not in request.files:
        flash("No se seleccionó ningún archivo", "error")
        return redirect("/cartera/importar")

    file = request.files['archivo_partes']

    if file.filename == '':
        flash("No se seleccionó ningún archivo", "error")
        return redirect("/cartera/importar")

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("El archivo debe ser formato Excel (.xlsx o .xls)", "error")
        return redirect("/cartera/importar")

    try:
        # Leer Excel
        df = pd.read_excel(file)

        # Palabras clave para detectar recomendaciones
        palabras_clave_recomendacion = [
            'RECOMENDACIÓN', 'RECOMENDACION', 'RECOMIENDO', 'RECOMENDAMOS',
            'CONVENDRÍA', 'CONVIENE', 'SERÍA CONVENIENTE', 'SE RECOMIENDA',
            'IMPORTANTE', 'URGENTE', 'NECESARIO', 'CAMBIAR', 'SUSTITUIR',
            'MODERNIZAR', 'REVISAR', 'PRÓXIMAMENTE', 'PROXIMAMENTE'
        ]

        # Cargar mapeo de tipos
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/tipos_parte_mapeo?select=*",
            headers=HEADERS
        )
        mapeo_tipos = {}
        if response.status_code == 200:
            for row in response.json():
                mapeo_tipos[row['tipo_original'].upper()] = row['tipo_normalizado']

        # Cargar máquinas
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/maquinas_cartera?select=id,identificador",
            headers=HEADERS
        )
        maquinas_map = {}
        if response.status_code == 200:
            for row in response.json():
                maquinas_map[row['identificador']] = row['id']

        stats = {
            'total': len(df),
            'insertados': 0,
            'duplicados': 0,
            'sin_maquina': 0,
            'errores': 0,
            'recomendaciones_detectadas': 0
        }

        partes_batch = []
        batch_size = 100

        for idx, row in df.iterrows():
            # Validar columnas requeridas
            if pd.isna(row.get('PARTE')) or pd.isna(row.get('MÁQUINA')):
                stats['errores'] += 1
                continue

            # Mapear máquina
            identificador_maquina = str(row['MÁQUINA']).strip()
            maquina_id = maquinas_map.get(identificador_maquina)

            if not maquina_id:
                stats['sin_maquina'] += 1

            # Mapear tipo de parte
            tipo_original = str(row.get('TIPO PARTE', '')).strip().upper()
            tipo_normalizado = mapeo_tipos.get(tipo_original, 'OTRO')

            # Parsear fecha
            fecha_parte = row.get('FECHA')
            if pd.isna(fecha_parte):
                stats['errores'] += 1
                continue

            if isinstance(fecha_parte, datetime):
                fecha_parte_iso = fecha_parte.isoformat()
            else:
                try:
                    fecha_parte_iso = pd.to_datetime(str(fecha_parte)).isoformat()
                except:
                    stats['errores'] += 1
                    continue

            # Detectar recomendaciones
            resolucion = row.get('RESOLUCIÓN', '')
            tiene_recomendacion = False
            recomendacion_extraida = None

            if pd.notna(resolucion):
                texto_upper = str(resolucion).upper()
                for palabra in palabras_clave_recomendacion:
                    if palabra in texto_upper:
                        tiene_recomendacion = True
                        recomendacion_extraida = str(resolucion)
                        stats['recomendaciones_detectadas'] += 1
                        break

            # Preparar datos para inserción
            parte_data = {
                "numero_parte": str(row['PARTE']),
                "tipo_parte_original": str(row.get('TIPO PARTE', '')).strip(),
                "codigo_maquina": str(row.get('CÓD. MÁQUINA', '')) if pd.notna(row.get('CÓD. MÁQUINA')) else None,
                "maquina_texto": identificador_maquina,
                "fecha_parte": fecha_parte_iso,
                "codificacion_adicional": str(row.get('CODIFICACIÓN ADICIONAL', '')) if pd.notna(row.get('CODIFICACIÓN ADICIONAL')) else None,
                "resolucion": str(resolucion) if pd.notna(resolucion) else None,
                "maquina_id": maquina_id,
                "tipo_parte_normalizado": tipo_normalizado,
                "tiene_recomendacion": tiene_recomendacion,
                "recomendaciones_extraidas": recomendacion_extraida if tiene_recomendacion else None,
                "estado": "COMPLETADO",
                "importado": True
            }

            partes_batch.append(parte_data)

            # Insertar por lotes
            if len(partes_batch) >= batch_size:
                response = requests.post(
                    f"{SUPABASE_URL}/rest/v1/partes_trabajo",
                    json=partes_batch,
                    headers={**HEADERS, "Prefer": "return=representation,resolution=ignore-duplicates"}
                )

                if response.status_code in [200, 201]:
                    stats['insertados'] += len(partes_batch)
                else:
                    stats['errores'] += len(partes_batch)

                partes_batch = []

        # Insertar lote final
        if partes_batch:
            response = requests.post(
                f"{SUPABASE_URL}/rest/v1/partes_trabajo",
                json=partes_batch,
                headers={**HEADERS, "Prefer": "return=representation,resolution=ignore-duplicates"}
            )

            if response.status_code in [200, 201]:
                stats['insertados'] += len(partes_batch)
            else:
                stats['errores'] += len(partes_batch)

        # Mostrar resumen
        flash(f"Importación completada: {stats['insertados']} partes insertados de {stats['total']} procesados", "success")

        if stats['recomendaciones_detectadas'] > 0:
            flash(f"{stats['recomendaciones_detectadas']} recomendaciones detectadas automáticamente", "success")

        if stats['sin_maquina'] > 0:
            flash(f"{stats['sin_maquina']} partes sin máquina asignada (revisar identificadores)", "warning")

        if stats['errores'] > 0:
            flash(f"{stats['errores']} registros con errores", "warning")

    except Exception as e:
        flash(f"Error al procesar archivo: {str(e)}", "error")
        logger.error(f"Error importando partes: {str(e)}")

    return redirect("/cartera/importar")


# ============================================
# OPORTUNIDADES DE FACTURACIÓN
# ============================================

@app.route("/cartera/oportunidades")
@helpers.login_required
def cartera_oportunidades():
    """Dashboard de oportunidades de facturación"""

    # Filtro por estado (opcional)
    estado_filtro = request.args.get('estado', '')

    # Query base
    query_params = []
    if estado_filtro:
        query_params.append(f"estado=eq.{estado_filtro}")

    query_string = "&".join(query_params) if query_params else ""
    url = f"{SUPABASE_URL}/rest/v1/oportunidades_facturacion?select=*,maquinas_cartera(identificador,instalaciones(nombre))&order=created_at.desc"
    if query_string:
        url += f"&{query_string}"

    response = requests.get(url, headers=HEADERS)
    oportunidades = response.json() if response.status_code == 200 else []

    # Agrupar por estado para vista Kanban
    oportunidades_por_estado = {
        'DETECTADA': [],
        'PRESUPUESTO_ENVIADO': [],
        'ACEPTADO': [],
        'PENDIENTE_REPUESTO': [],
        'LISTO_EJECUTAR': [],
        'COMPLETADO': [],
        'FACTURADO': [],
        'RECHAZADO': []
    }

    for opp in oportunidades:
        estado = opp.get('estado', 'DETECTADA')
        if estado in oportunidades_por_estado:
            oportunidades_por_estado[estado].append(opp)

    # Estadísticas
    stats = {
        'total': len(oportunidades),
        'detectadas': len(oportunidades_por_estado['DETECTADA']),
        'en_proceso': len(oportunidades_por_estado['PRESUPUESTO_ENVIADO']) + len(oportunidades_por_estado['ACEPTADO']) + len(oportunidades_por_estado['PENDIENTE_REPUESTO']) + len(oportunidades_por_estado['LISTO_EJECUTAR']),
        'completadas': len(oportunidades_por_estado['COMPLETADO']),
        'facturadas': len(oportunidades_por_estado['FACTURADO']),
        'rechazadas': len(oportunidades_por_estado['RECHAZADO'])
    }

    return render_template(
        "cartera/oportunidades.html",
        oportunidades_por_estado=oportunidades_por_estado,
        stats=stats,
        estado_filtro=estado_filtro
    )


@app.route("/cartera/oportunidades/crear/<int:parte_id>", methods=["GET", "POST"])
@helpers.login_required
def cartera_crear_oportunidad(parte_id):
    """Crear oportunidad desde una recomendación"""

    if request.method == "GET":
        # Obtener datos del parte
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/partes_trabajo?id=eq.{parte_id}&select=*,maquinas_cartera(id,identificador,instalaciones(nombre))",
            headers=HEADERS
        )

        if response.status_code != 200 or not response.json():
            flash("Parte no encontrado", "error")
            return redirect("/cartera")

        parte = response.json()[0]

        return render_template("cartera/crear_oportunidad.html", parte=parte)

    else:  # POST
        # Obtener datos del formulario
        titulo = request.form.get('titulo')
        descripcion = request.form.get('descripcion')
        tipo = request.form.get('tipo')
        prioridad = request.form.get('prioridad', 'MEDIA')
        repuestos = request.form.get('repuestos')

        # Obtener maquina_id del parte
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/partes_trabajo?id=eq.{parte_id}&select=maquina_id",
            headers=HEADERS
        )

        if response.status_code != 200 or not response.json():
            flash("Error al obtener datos del parte", "error")
            return redirect("/cartera/oportunidades")

        maquina_id = response.json()[0]['maquina_id']

        # Crear oportunidad
        oportunidad_data = {
            "maquina_id": maquina_id,
            "parte_origen_id": parte_id,
            "titulo": titulo,
            "descripcion_tecnica": descripcion,
            "tipo": tipo,
            "estado": "DETECTADA",
            "prioridad_comercial": prioridad,
            "repuestos_necesarios": repuestos if repuestos else None,
            "created_by": session.get("usuario_email", "sistema")
        }

        response = requests.post(
            f"{SUPABASE_URL}/rest/v1/oportunidades_facturacion",
            json=oportunidad_data,
            headers=HEADERS
        )

        if response.status_code == 201:
            oportunidad_id = response.json()[0]['id']

            # Marcar parte como oportunidad creada
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/partes_trabajo?id=eq.{parte_id}",
                json={"oportunidad_creada": True, "oportunidad_id": oportunidad_id, "recomendacion_revisada": True},
                headers=HEADERS
            )

            flash("Oportunidad creada exitosamente", "success")
            return redirect(f"/cartera/oportunidades/{oportunidad_id}")
        else:
            logger.error(f"Error creando oportunidad: {response.status_code} - {response.text}")
            flash("Error al crear oportunidad", "error")
            return redirect("/cartera/oportunidades")


@app.route("/cartera/recomendaciones/<int:parte_id>/descartar", methods=["POST"])
@helpers.login_required
def cartera_descartar_recomendacion(parte_id):
    """Descartar una recomendación sin crear oportunidad"""

    # Marcar recomendación como revisada pero sin crear oportunidad
    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/partes_trabajo?id=eq.{parte_id}",
        json={
            "recomendacion_revisada": True,
            "oportunidad_creada": False,
            "oportunidad_id": None
        },
        headers=HEADERS
    )

    if response.status_code in [200, 204]:
        flash("Recomendación descartada", "success")
    else:
        logger.error(f"Error descartando recomendación: {response.status_code} - {response.text}")
        flash("Error al descartar recomendación", "error")

    return redirect("/cartera")


@app.route("/cartera/oportunidades/<int:oportunidad_id>")
@helpers.login_required
def cartera_ver_oportunidad(oportunidad_id):
    """Ver detalle de oportunidad"""

    # Obtener oportunidad con datos relacionados
    response = requests.get(
        f"{SUPABASE_URL}/rest/v1/oportunidades_facturacion?id=eq.{oportunidad_id}&select=*,maquinas_cartera(identificador,codigo_maquina,instalaciones(nombre,municipio)),partes_trabajo(numero_parte,fecha_parte,resolucion,recomendaciones_extraidas)",
        headers=HEADERS
    )

    if response.status_code != 200 or not response.json():
        flash("Oportunidad no encontrada", "error")
        return redirect("/cartera/oportunidades")

    oportunidad = response.json()[0]

    return render_template("cartera/ver_oportunidad.html", oportunidad=oportunidad)


@app.route("/cartera/oportunidades/<int:oportunidad_id>/actualizar", methods=["POST"])
@helpers.login_required
def cartera_actualizar_oportunidad(oportunidad_id):
    """Actualizar estado y datos de oportunidad"""

    # Obtener datos del formulario
    accion = request.form.get('accion')

    update_data = {}

    if accion == 'cambiar_estado':
        nuevo_estado = request.form.get('estado')
        update_data['estado'] = nuevo_estado

        # Actualizar fechas según el estado
        if nuevo_estado == 'PRESUPUESTO_ENVIADO':
            update_data['fecha_envio_presupuesto'] = date.today().isoformat()
            update_data['numero_presupuesto_erp'] = request.form.get('numero_presupuesto')
            update_data['importe_presupuestado'] = request.form.get('importe')
        elif nuevo_estado == 'ACEPTADO':
            update_data['fecha_aceptacion'] = date.today().isoformat()
            update_data['fecha_respuesta_cliente'] = date.today().isoformat()
        elif nuevo_estado == 'RECHAZADO':
            update_data['fecha_rechazo'] = date.today().isoformat()
            update_data['fecha_respuesta_cliente'] = date.today().isoformat()
            update_data['motivo_rechazo'] = request.form.get('motivo_rechazo')
        elif nuevo_estado == 'COMPLETADO':
            update_data['fecha_completado'] = date.today().isoformat()
            update_data['importe_final'] = request.form.get('importe_final')
        elif nuevo_estado == 'FACTURADO':
            update_data['facturado'] = True
            update_data['fecha_factura'] = date.today().isoformat()
            update_data['numero_factura'] = request.form.get('numero_factura')

    elif accion == 'actualizar_repuestos':
        update_data['estado_repuestos'] = request.form.get('estado_repuestos')
        update_data['proveedor'] = request.form.get('proveedor')
        update_data['coste_repuestos'] = request.form.get('coste_repuestos')

        if request.form.get('estado_repuestos') == 'SOLICITADO':
            update_data['fecha_solicitud_repuesto'] = date.today().isoformat()
        elif request.form.get('estado_repuestos') == 'RECIBIDO':
            update_data['fecha_recepcion_repuesto'] = date.today().isoformat()

    elif accion == 'actualizar_notas':
        update_data['notas'] = request.form.get('notas')

    # Actualizar timestamp
    update_data['updated_at'] = datetime.now().isoformat()

    # Ejecutar actualización
    response = requests.patch(
        f"{SUPABASE_URL}/rest/v1/oportunidades_facturacion?id=eq.{oportunidad_id}",
        json=update_data,
        headers=HEADERS
    )

    if response.status_code == 200:
        flash("Oportunidad actualizada correctamente", "success")
    else:
        logger.error(f"Error actualizando oportunidad: {response.status_code} - {response.text}")
        flash("Error al actualizar oportunidad", "error")

    return redirect(f"/cartera/oportunidades/{oportunidad_id}")


# ============================================
# CIERRE
# ============================================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
